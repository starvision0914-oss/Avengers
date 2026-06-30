"""
상품 조회/수정 → 엑셀 일괄작업 드롭다운 열기 → 엑셀 다운로드
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    path = f'{DOWNLOAD_DIR}/{name}.png'
    driver.save_screenshot(path)
    print(f'스크린샷: {path}')

def wait_download(timeout=30):
    for _ in range(timeout):
        files = glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls')
        if files:
            return files[0]
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)

    # 팝업 닫기
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    # 상품 조회/수정 이동
    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(5)

    # 팝업 닫기 (일괄등록 페이지에서 뜬 팝업 유사)
    try:
        driver.find_element(By.XPATH, '//*[contains(@class,"close") and not(contains(@class,"icon"))] | //button[contains(text(),"닫기")] | //*[@aria-label="닫기"]').click()
        time.sleep(1)
    except: pass

    # 검색 실행
    search_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    )
    search_btn.click()
    time.sleep(5)
    ss('10_search_done')

    # 엑셀 일괄작업 버튼 찾기
    print('엑셀 일괄작업 버튼 탐색')
    excel_btn = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"엑셀 일괄작업") or contains(text(),"엑셀일괄작업")]'))
    )
    print('버튼 텍스트:', excel_btn.text)
    excel_btn.click()
    time.sleep(2)
    ss('11_excel_dropdown')

    # 드롭다운 옵션 출력
    options = driver.find_elements(By.XPATH, '//ul[@class and contains(@class,"dropdown")]//li | //div[contains(@class,"dropdown-menu")]//a | //*[@role="menuitem"] | //*[@role="option"]')
    print('엑셀 드롭다운 옵션들:')
    for o in options:
        txt = o.text.strip()
        if txt:
            print(f'  "{txt}"')

    # 다운로드 옵션 클릭
    for keyword in ['다운로드', '상품정보 다운', '엑셀 다운']:
        try:
            opt = driver.find_element(By.XPATH, f'//*[contains(text(),"{keyword}")]')
            print(f'"{keyword}" 클릭')
            opt.click()
            time.sleep(3)
            ss('12_after_download_click')
            break
        except: pass

    # 추가 팝업/확인창 처리
    try:
        ok_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"확인") or contains(text(),"다운로드")]'))
        )
        ok_btn.click()
        time.sleep(3)
    except: pass

    # 파일 대기
    dl_file = wait_download(30)
    if dl_file:
        print(f'다운로드 완료: {dl_file}')
        import openpyxl
        wb = openpyxl.load_workbook(dl_file)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            print(f'[{sheet_name}] 헤더({len(headers)}개):', headers)
    else:
        print('파일 다운로드 실패')
        ss('13_no_download')

finally:
    time.sleep(2)
    driver.quit()
    print('완료')
