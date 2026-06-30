"""
Angular triggerHandler + 좌표 클릭으로 상품 체크박스 선택 → 수정양식 다운로드
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=90):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 100]
        if done:
            return done[0]
        if i % 10 == 0:
            print(f'  대기 {i}s')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(7)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    # 스크롤 최상단으로
    driver.execute_script("window.scrollTo(0, 0)")
    time.sleep(1)

    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(7)

    # 스크롤 없이 체크박스 위치 재확인
    chk_pos = driver.execute_script("""
        var inputs = document.querySelectorAll('input[type="checkbox"]');
        var result = [];
        inputs.forEach(function(el, i) {
            var rect = el.getBoundingClientRect();
            if (el.offsetParent !== null && rect.width > 0) {
                result.push({i: i, ng: el.getAttribute('ng-model')||'', name: el.name,
                    top: Math.round(rect.top), left: Math.round(rect.left)});
            }
        });
        return result;
    """)
    print('화면에 보이는 체크박스:')
    for c in chk_pos:
        print(f'  [{c["i"]}] ng={c["ng"]} name={c["name"]} pos=({c["left"]},{c["top"]})')

    # ng-model 없는 체크박스 = 상품 행 체크박스
    product_chks = [c for c in chk_pos if not c['ng'] and not c['name']]
    print(f'\n상품 행 체크박스: {len(product_chks)}개')
    for c in product_chks:
        print(f'  [{c["i"]}] pos=({c["left"]},{c["top"]})')

    if product_chks:
        # 첫 번째 상품 체크박스 Angular triggerHandler로 클릭
        idx = product_chks[0]['i']
        result = driver.execute_script(f"""
            var inputs = document.querySelectorAll('input[type="checkbox"]');
            var el = inputs[{idx}];
            try {{
                // Angular triggerHandler
                if (typeof angular !== 'undefined') {{
                    angular.element(el).triggerHandler('click');
                    return 'angular click';
                }}
            }} catch(e) {{}}
            // 일반 클릭
            el.click();
            el.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true}}));
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
            return 'dom click';
        """)
        print(f'체크박스[{idx}] 클릭: {result}')
        time.sleep(1)

        # 확인
        is_checked = driver.execute_script(f"""
            var inputs = document.querySelectorAll('input[type="checkbox"]');
            return inputs[{idx}].checked;
        """)
        print(f'체크 상태: {is_checked}')
    else:
        # ng-model 없는 것 없으면 좌표 직접 클릭 (scroll=0 상태에서)
        print('직접 좌표 클릭: (289, 693)')
        driver.execute_script("window.scrollTo(0, 0)")
        time.sleep(0.5)
        action = ActionChains(driver)
        action.move_by_offset(289, 693).click().perform()
        time.sleep(1)

    ss('a0_after_select')

    # 수정양식 다운로드
    driver.execute_script("""
        var btns = document.querySelectorAll('button, div.item, li, a, span');
        for (var b of btns) {
            if (b.textContent.trim() === '엑셀 일괄작업' && b.offsetParent) {
                b.click(); return;
            }
        }
    """)
    time.sleep(2)
    ss('a1_dropdown')

    msg = driver.execute_script("""
        var items = document.querySelectorAll('a, li, div, span, button');
        for (var b of items) {
            if (b.textContent.trim() === '수정양식 다운로드' && b.offsetParent) {
                b.click();
                return '클릭: ' + b.textContent.trim();
            }
        }
        return '없음';
    """)
    print('수정양식:', msg)
    time.sleep(3)

    try:
        alert = driver.find_element(By.XPATH, '//*[contains(text(),"없습니다") or contains(text(),"선택") or contains(text(),"생성")]')
        print('알림:', alert.text[:100])
    except: pass

    ss('a2_final')
    dl = wait_download(90)

    if dl:
        print(f'\n수정양식 성공: {dl} ({os.path.getsize(dl):,}bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                h = next(csv.reader(fp))
            print(f'컬럼 {len(h)}개:')
            for i,v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for ri in range(1, 4):
                    row = [ws.cell(ri,c).value for c in range(1, ws.max_column+1)]
                    print(f'행{ri}: {[v for v in row if v][:15]}')
                h2 = [ws.cell(2,c).value for c in range(1, ws.max_column+1)]
                print(f'\n전체 컬럼:')
                for i,v in enumerate(h2):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        print('수정양식 다운로드 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
