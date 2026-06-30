"""
상품목록 다운로드 (선택 불필요) → 엑셀 헤더 확인
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'스크린샷: {name}')

def wait_download(timeout=180):
    for i in range(timeout):
        files = glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls')
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 1000]
        if done:
            return done[0]
        if i % 20 == 0:
            print(f'  대기 {i}초...')
            # 다운로드 진행 팝업 텍스트 출력
            try:
                pg = driver.find_element(By.XPATH, '//*[contains(text(),"생성중") or contains(text(),"/997")]')
                print(f'  진행: {pg.text[:60]}')
            except: pass
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)

    # 팝업 닫기
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    # 상품 조회/수정 이동
    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(6)

    # 팝업 닫기
    for kw in ['하루동안', '닫기']:
        try:
            driver.find_element(By.XPATH, f'//*[contains(text(),"{kw}")]').click()
            time.sleep(0.5)
        except: pass

    # 검색
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(5)

    # 팝업 한 번 더 닫기
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
    except: pass

    ss('40_before_excel')

    # 엑셀 일괄작업 클릭
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"엑셀 일괄작업")]'))
    ).click()
    time.sleep(2)
    ss('41_dropdown')

    # 상품목록 다운로드 클릭
    opt = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"상품목록 다운로드")]'))
    )
    print('상품목록 다운로드 클릭')
    opt.click()
    time.sleep(3)
    ss('42_after_click')

    # 진행 팝업 확인
    try:
        pg = driver.find_element(By.XPATH, '//*[contains(text(),"생성중") or contains(text(),"/997") or contains(text(),"다운로드")]')
        print('진행 팝업:', pg.text[:80])
    except: pass

    # 파일 대기 (최대 3분)
    dl_file = wait_download(180)
    ss('43_final')

    if dl_file:
        print(f'\n다운로드 완료: {dl_file} ({os.path.getsize(dl_file):,} bytes)')
        import openpyxl
        wb = openpyxl.load_workbook(dl_file, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            print(f'\n[{sheet_name}] 컬럼 {ws.max_column}개:')
            for i, h in enumerate(row1):
                print(f'  {i+1:3d}: {h}')
        wb.close()
    else:
        print('다운로드 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
