"""
엑셀 일괄작업 → 수정양식 다운로드 (속성 컬럼 확인용)
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    path = f'{DOWNLOAD_DIR}/{name}.png'
    driver.save_screenshot(path)
    print(f'스크린샷: {path}')

def wait_download(timeout=120):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        # .crdownload 제외 (다운로드 중인 파일)
        done = [f for f in files if not f.endswith('.crdownload')]
        if done:
            return done[0]
        if i % 10 == 0:
            print(f'  다운로드 대기 {i}초...')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)

    # 팝업 닫기
    for kw in ['하루동안', '닫기', 'close']:
        try:
            driver.find_element(By.XPATH, f'//*[contains(text(),"{kw}")]').click()
            time.sleep(0.5)
        except: pass

    # 상품 조회/수정 이동
    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(5)

    # 팝업 닫기 재시도
    try:
        driver.find_element(By.XPATH, '//button[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    # 검색
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(5)

    # 엑셀 일괄작업 드롭다운 클릭
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"엑셀 일괄작업")]'))
    ).click()
    time.sleep(2)
    ss('20_dropdown_open')

    # 드롭다운 옵션 확인
    all_visible = driver.find_elements(By.XPATH, '//*[contains(@class,"dropdown") or contains(@class,"popup")]//li | //*[@class and contains(@class,"option")]//span')
    print('드롭다운 내 항목:')
    for e in all_visible[:20]:
        t = e.text.strip()
        if t and len(t) < 20:
            print(f'  "{t}"')

    # "수정양식 다운로드" 클릭 우선, 없으면 "상품목록 다운로드"
    clicked = False
    for keyword in ['수정양식 다운로드', '수정 양식 다운로드', '상품목록 다운로드', '상품 목록 다운로드']:
        try:
            opt = driver.find_element(By.XPATH, f'//*[normalize-space(text())="{keyword}"] | //*[contains(text(),"{keyword}")]')
            print(f'클릭: "{keyword}"')
            opt.click()
            clicked = True
            time.sleep(3)
            break
        except: pass

    if not clicked:
        print('수정양식/상품목록 다운로드 버튼 못 찾음')
        ss('20_fail')
        sys.exit(1)

    ss('21_after_click')

    # 다운로드 진행 팝업 대기
    try:
        progress = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[contains(text(),"생성중") or contains(text(),"다운로드")]'))
        )
        print('진행 팝업:', progress.text[:80])
    except: pass

    # 파일 완성 대기 (120초)
    dl_file = wait_download(120)
    ss('22_after_download')

    if dl_file:
        print(f'\n다운로드 완료: {dl_file}')
        import openpyxl
        wb = openpyxl.load_workbook(dl_file, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row1 = [ws.cell(1, c).value for c in range(1, min(ws.max_column+1, 100))]
            print(f'\n[{sheet_name}] 컬럼 {ws.max_column}개:')
            for i, h in enumerate(row1):
                if h:
                    print(f'  {i+1}: {h}')
        wb.close()
    else:
        print('파일 다운로드 실패 (120초 초과)')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
