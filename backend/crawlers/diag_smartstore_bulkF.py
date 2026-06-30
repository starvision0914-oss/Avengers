"""
ag-selection-checkbox 클릭 → 수정양식 다운로드 (240초 대기)
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=240):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 100]
        if done:
            return done[0]
        if i % 20 == 0:
            print(f'  대기 {i}s')
            try:
                pg = driver.find_element(By.XPATH, '//*[contains(text(),"생성중") or contains(text(),"/100")]')
                print(f'  진행: {pg.text[:60]}')
            except: pass
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(7)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(7)
    ss('c0_searched')

    # ag-selection-checkbox (헤더 전체 선택) 찾아서 ActionChains 클릭
    all_chks = driver.find_elements(By.CSS_SELECTOR, 'input[type="checkbox"]')
    chk12 = all_chks[12]
    print(f'ag-selection-checkbox: displayed={chk12.is_displayed()}, class={chk12.get_attribute("class")}')

    driver.execute_script("arguments[0].scrollIntoView({block:'center'})", chk12)
    time.sleep(0.5)
    ActionChains(driver).move_to_element(chk12).click().perform()
    time.sleep(1)

    print(f'클릭 후 checked: {chk12.is_selected()}')
    ss('c1_selected')

    # 선택 확인
    sel_info = driver.execute_script("""
        var agChks = document.querySelectorAll('.ag-selection-checkbox:checked');
        var allAgChks = document.querySelectorAll('.ag-selection-checkbox');
        return {checked: agChks.length, total: allAgChks.length};
    """)
    print(f'ag 체크박스 상태: {sel_info}')

    # 엑셀 일괄작업 드롭다운 클릭
    driver.execute_script("""
        var els = document.querySelectorAll('button, div, li, a, span');
        for (var e of els) {
            if (e.textContent.trim() === '엑셀 일괄작업' && e.offsetParent) {
                e.click(); break;
            }
        }
    """)
    time.sleep(2)
    ss('c2_dropdown')

    # 수정양식 다운로드 클릭
    msg = driver.execute_script("""
        var els = document.querySelectorAll('a, li, div, span, button, ul li');
        for (var e of els) {
            if (e.textContent.trim() === '수정양식 다운로드' && e.offsetParent) {
                e.click();
                return e.textContent.trim();
            }
        }
        return '없음';
    """)
    print('클릭:', msg)
    time.sleep(3)
    ss('c3_after_click')

    # 알림/진행 팝업 확인
    try:
        alert = driver.find_element(By.XPATH, '//*[contains(text(),"없습니다") or contains(text(),"선택") or contains(text(),"생성") or contains(text(),"다운로드")]')
        print('팝업:', alert.text[:100])
    except: pass

    # 240초 대기
    dl = wait_download(240)
    ss('c4_final')

    if dl:
        print(f'\n성공! {dl} ({os.path.getsize(dl):,}bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                h = next(csv.reader(fp))
            print(f'컬럼 {len(h)}개:')
            for i,v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for ri in range(1,4):
                    r = [ws.cell(ri,c).value for c in range(1,ws.max_column+1)]
                    vals = [v for v in r if v]
                    if vals: print(f'행{ri}: {vals[:15]}')
                h2 = [ws.cell(2,c).value for c in range(1,ws.max_column+1)]
                print(f'\n[{sn}] 전체 컬럼:')
                for i,v in enumerate(h2):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        # 다른 위치 탐색
        import subprocess
        result = subprocess.run(['find', '/home/rejoice888', '-name', '*.xlsx', '-newer', f'{DOWNLOAD_DIR}/c0_searched.png'], capture_output=True, text=True)
        print('다른 xlsx:', result.stdout.strip()[:300])
        print('다운로드 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
