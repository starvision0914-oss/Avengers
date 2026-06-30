"""
tbody 체크박스 클릭 → 수정양식 다운로드 → 속성 컬럼 확인
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=90):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 100]
        if done:
            return done[0]
        if i % 10 == 0:
            print(f'  대기 {i}s')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)

    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(7)

    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    # 검색
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(7)
    ss('80_searched')

    # tbody 내 체크박스 정보 상세 조회
    chk_detail = driver.execute_script("""
        // tbody 직접 탐색
        var tbodies = document.querySelectorAll('tbody');
        var result = {tbody_count: tbodies.length, inputs: []};

        tbodies.forEach(function(tb, ti) {
            var inputs = tb.querySelectorAll('input[type="checkbox"]');
            inputs.forEach(function(el, i) {
                if (result.inputs.length < 5) {
                    result.inputs.push({
                        tbody_idx: ti,
                        input_idx: i,
                        name: el.name,
                        id: el.id,
                        cls: el.className.substring(0,50),
                        ng_model: el.getAttribute('ng-model') || '',
                        visible: el.offsetParent !== null
                    });
                }
            });
        });

        // table 탐색
        var tables = document.querySelectorAll('table');
        result.table_count = tables.length;

        // ng-repeat 탐색
        var ngRepeat = document.querySelectorAll('[ng-repeat]');
        result.ng_repeat_count = ngRepeat.length;
        if (ngRepeat.length > 0) {
            result.ng_repeat_first = ngRepeat[0].getAttribute('ng-repeat');
        }

        return result;
    """)
    print('DOM 탐색 결과:')
    import json
    print(json.dumps(chk_detail, ensure_ascii=False, indent=2))

    # tbody input 클릭 시도
    clicked = driver.execute_script("""
        var tbodies = document.querySelectorAll('tbody');
        var clicked = 0;
        tbodies.forEach(function(tb) {
            var inputs = tb.querySelectorAll('input[type="checkbox"]');
            inputs.forEach(function(el, i) {
                if (i < 3) {
                    el.click();
                    // 또는 dispatchEvent
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                    el.dispatchEvent(new MouseEvent('click', {bubbles: true}));
                    clicked++;
                }
            });
        });
        return clicked;
    """)
    print(f'tbody 체크박스 {clicked}개 클릭')
    time.sleep(2)
    ss('81_selected')

    # 선택된 상품 수 확인
    sel_count = driver.execute_script("""
        var selected = document.querySelectorAll('tbody input[type="checkbox"]:checked');
        return selected.length;
    """)
    print(f'선택된 체크박스: {sel_count}개')

    # 엑셀 일괄작업 → 수정양식 다운로드
    driver.execute_script("""
        var btns = document.querySelectorAll('button, a, div');
        for (var b of btns) {
            if (b.textContent.trim() === '엑셀 일괄작업') {
                b.click();
                return;
            }
        }
    """)
    time.sleep(2)
    ss('82_dropdown')

    driver.execute_script("""
        var btns = document.querySelectorAll('a, li, button, span');
        for (var b of btns) {
            if (b.textContent.trim() === '수정양식 다운로드') {
                b.click();
                return;
            }
        }
    """)
    time.sleep(3)

    try:
        msg = driver.find_element(By.XPATH, '//*[contains(text(),"없습니다") or contains(text(),"선택")]')
        print('메시지:', msg.text[:80])
    except: pass

    ss('83_after_click')
    dl = wait_download(90)

    if dl:
        print(f'\n수정양식 성공: {dl} ({os.path.getsize(dl):,} bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                h = next(csv.reader(fp))
            print(f'컬럼 {len(h)}개:')
            for i,v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                rows = list(ws.rows)
                for ri, row in enumerate(rows[:3]):
                    vals = [c.value for c in row if c.value]
                    print(f'행{ri+1}: {vals[:10]}')
                h = [ws.cell(2,c).value for c in range(1, ws.max_column+1)]
                print(f'\n[{sn}] 전체 컬럼 {ws.max_column}개:')
                for i,v in enumerate(h):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        print('수정양식 다운로드 실패')
        # 현재 화면 상태
        cur_url = driver.current_url
        print('현재 URL:', cur_url)

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
