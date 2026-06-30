"""
체크박스 HTML 구조 분석 + Selenium WebElement.click() 직접 클릭
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=90):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 100]
        if done:
            return done[0]
        if i % 10 == 0:
            print(f'  대기 {i}s')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(7)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(7)

    # 체크박스 12번 HTML 구조 분석
    html_info = driver.execute_script("""
        var inputs = document.querySelectorAll('input[type="checkbox"]');
        var el = inputs[12];
        if (!el) return 'not found';
        return {
            html: el.outerHTML.substring(0, 200),
            parentHtml: el.parentElement ? el.parentElement.outerHTML.substring(0, 400) : 'no parent',
            grandpaHtml: el.parentElement && el.parentElement.parentElement ?
                el.parentElement.parentElement.outerHTML.substring(0, 500) : 'no grandpa',
            ngClick: el.getAttribute('ng-click') || '',
            parentNgClick: el.parentElement ? (el.parentElement.getAttribute('ng-click') || '') : ''
        };
    """)
    import json
    print('체크박스 12 HTML 구조:')
    print(json.dumps(html_info, ensure_ascii=False, indent=2))

    # Selenium WebElement.click() 직접 시도
    all_chks = driver.find_elements(By.CSS_SELECTOR, 'input[type="checkbox"]')
    print(f'\n전체 체크박스: {len(all_chks)}개')

    if len(all_chks) > 12:
        chk12 = all_chks[12]
        print(f'체크박스 12: displayed={chk12.is_displayed()}, enabled={chk12.is_enabled()}')

        # scrollIntoView 후 클릭
        driver.execute_script("arguments[0].scrollIntoView({block:'center'})", chk12)
        time.sleep(1)
        ss('b0_before_click')

        try:
            ActionChains(driver).move_to_element(chk12).click().perform()
            print('ActionChains 클릭 완료')
        except Exception as e:
            print(f'ActionChains 실패: {e}')
            try:
                driver.execute_script("arguments[0].click()", chk12)
                print('JS click 완료')
            except Exception as e2:
                print(f'JS click 실패: {e2}')

        time.sleep(1)
        print(f'클릭 후 checked: {chk12.is_selected()}')
        ss('b1_after_click')

    # 부모 요소의 label 또는 click 영역 찾기
    parent_info = driver.execute_script("""
        var inputs = document.querySelectorAll('input[type="checkbox"]');
        var el = inputs[12];
        if (!el) return null;

        // 형제나 부모에서 클릭 가능한 label 찾기
        var parent = el.parentElement;
        var label = null;

        // label 찾기
        if (parent) {
            label = parent.querySelector('label') || parent.closest('label');
        }
        if (!label) {
            label = document.querySelector('label[for="'+el.id+'"]');
        }

        return {
            hasLabel: !!label,
            labelHtml: label ? label.outerHTML.substring(0, 200) : 'no label',
            parentTag: parent ? parent.tagName : 'no parent',
            parentClass: parent ? parent.className.substring(0, 50) : ''
        };
    """)
    print('\n부모/레이블 정보:')
    print(json.dumps(parent_info, ensure_ascii=False, indent=2))

    # 헤더 체크박스 (전체 선택) 찾기 시도
    # 스크린샷에서 (248, 605) 위치 좌표 클릭
    print('\n헤더 체크박스 좌표 클릭 (248, 605)')
    driver.execute_script("window.scrollTo(0, 0)")
    time.sleep(0.5)

    # body 기준 좌표 클릭
    body = driver.find_element(By.TAG_NAME, 'body')
    try:
        ActionChains(driver).move_to_element_with_offset(body, 248 - body.size['width']//2, 605 - body.size['height']//2).click().perform()
        print('좌표 클릭 완료')
    except Exception as e:
        print(f'좌표 클릭 오류: {e}')

    time.sleep(1)
    ss('b2_after_coord')

    # 선택된 상품 확인 (화면에서 체크된 항목 수)
    sel_info = driver.execute_script("""
        var checked = document.querySelectorAll('input[type="checkbox"]:checked');
        var allInputs = document.querySelectorAll('input[type="checkbox"]');
        return {checked: checked.length, total: allInputs.length};
    """)
    print(f'체크 상태: {sel_info}')

    # 수정양식 다운로드
    driver.execute_script("""
        var els = document.querySelectorAll('button, div.item, li, a, span, div');
        for (var e of els) {
            if (e.textContent.trim() === '엑셀 일괄작업' && e.offsetParent) {
                e.click(); break;
            }
        }
    """)
    time.sleep(2)
    ss('b3_dropdown')

    msg = driver.execute_script("""
        var els = document.querySelectorAll('a, li, div, span, button, ul li');
        for (var e of els) {
            if (e.textContent.trim() === '수정양식 다운로드' && e.offsetParent) {
                e.click();
                return e.textContent.trim();
            }
        }
        return '없음';
    """)
    print('수정양식 클릭:', msg)
    time.sleep(3)

    try:
        alert = driver.find_element(By.XPATH, '//*[contains(text(),"없습니다") or contains(text(),"생성중")]')
        print('알림:', alert.text[:100])
    except: pass

    ss('b4_final')
    dl = wait_download(90)
    if dl:
        print(f'\n성공! {dl} ({os.path.getsize(dl):,}bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                h = next(csv.reader(fp))
            for i,v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for ri in range(1,4):
                    r = [ws.cell(ri,c).value for c in range(1,ws.max_column+1)]
                    print(f'행{ri}:', [v for v in r if v][:15])
                h2 = [ws.cell(2,c).value for c in range(1,ws.max_column+1)]
                for i,v in enumerate(h2):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        print('수정양식 다운로드 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
