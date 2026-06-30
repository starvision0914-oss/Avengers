"""
팝업 닫기 → 체크박스 선택 → 수정양식 다운로드
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    path = f'{DOWNLOAD_DIR}/{name}.png'
    driver.save_screenshot(path)
    print(f'스크린샷: {path}')

def wait_download(timeout=180):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.xls'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 1000]
        if done:
            return done[0]
        if i % 15 == 0:
            print(f'  다운로드 대기 {i}초...')
        time.sleep(1)
    return None

def close_popups():
    for xpath in [
        '//label[contains(text(),"하루동안")]',
        '//button[contains(text(),"하루동안")]',
        '//*[contains(@class,"modal")]//button[contains(@class,"close")]',
        '//*[@aria-label="닫기"]',
        '//*[contains(@class,"btn-close")]',
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            el.click()
            time.sleep(0.5)
        except: pass

    # X 버튼들 클릭
    try:
        closes = driver.find_elements(By.XPATH, '//*[contains(@class,"modal") and contains(@style,"display: block")]//button[@type="button"]')
        for c in closes:
            try:
                c.click()
                time.sleep(0.3)
            except: pass
    except: pass

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)

    # 첫 팝업 닫기
    time.sleep(2)
    close_popups()
    time.sleep(1)

    # 상품 조회/수정 이동
    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(6)
    close_popups()
    time.sleep(1)
    ss('30_after_close')

    # 검색
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(6)
    close_popups()
    ss('31_search_result')

    # 체크박스 찾기 - JavaScript로 모든 checkbox 정보 출력
    chk_info = driver.execute_script("""
        var inputs = document.querySelectorAll('input[type="checkbox"]');
        var result = [];
        inputs.forEach(function(el, idx) {
            if (idx < 10) {
                result.push({
                    idx: idx,
                    id: el.id,
                    name: el.name,
                    class: el.className,
                    checked: el.checked,
                    visible: el.offsetParent !== null
                });
            }
        });
        return result;
    """)
    print(f'체크박스 {len(chk_info)}개 발견:')
    for c in chk_info:
        print(f'  {c}')

    # JS로 첫 번째 데이터 행 체크박스 클릭 (인덱스 1 = 헤더 제외)
    clicked = driver.execute_script("""
        var inputs = document.querySelectorAll('input[type="checkbox"]');
        if (inputs.length > 1) {
            inputs[1].click();  // 첫 번째 상품
            inputs[2].click();  // 두 번째 상품
            inputs[3].click();  // 세 번째 상품
            return inputs.length;
        }
        return 0;
    """)
    print(f'JS 체크박스 클릭 완료 (전체 {clicked}개)')
    time.sleep(1)
    ss('32_selected')

    # 엑셀 일괄작업 클릭
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"엑셀 일괄작업")]'))
    ).click()
    time.sleep(2)
    ss('33_dropdown')

    # 수정양식 다운로드 클릭
    for keyword in ['수정양식 다운로드', '수정 양식', '상품목록 다운로드']:
        try:
            opt = driver.find_element(By.XPATH, f'//*[contains(text(),"{keyword}")]')
            print(f'클릭: "{keyword}"')
            opt.click()
            time.sleep(3)
            break
        except: pass

    ss('34_after_click')

    # 팝업 확인
    try:
        msg = driver.find_element(By.XPATH, '//*[contains(@class,"modal-body") or contains(@class,"toast") or contains(@class,"alert")]')
        print('메시지:', msg.text[:100])
    except: pass

    # 파일 대기
    dl_file = wait_download(180)
    ss('35_final')

    if dl_file:
        print(f'\n다운로드 완료: {dl_file} ({os.path.getsize(dl_file)} bytes)')
        import openpyxl
        wb = openpyxl.load_workbook(dl_file, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row1 = [ws.cell(1, c).value for c in range(1, min(ws.max_column+1, 200))]
            print(f'\n[{sheet_name}] 컬럼 {ws.max_column}개:')
            for i, h in enumerate(row1):
                print(f'  {i+1:3d}: {h}')
        wb.close()
    else:
        print('파일 다운로드 실패 (180초 초과)')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
