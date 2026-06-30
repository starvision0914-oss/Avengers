"""
방법 A: 상품 행 체크박스(인덱스10+) 클릭 → 수정양식 다운로드
방법 B: 상품 일괄등록 → 양식 다운로드
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=120):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 500]
        if done:
            return done[0]
        if i % 20 == 0:
            print(f'  대기 {i}s')
        time.sleep(1)
    return None

def close_popups():
    for xpath in ['//*[contains(text(),"하루동안")]', '//button[contains(text(),"닫기")]',
                  '//*[@class and contains(@class,"close") and @type="button"]']:
        try:
            driver.find_element(By.XPATH, xpath).click()
            time.sleep(0.5)
        except: pass

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)
    close_popups()

    # ── 방법 B: 상품 일괄등록 → 양식 다운로드 ──
    print('\n=== 방법B: 상품 일괄등록 양식 다운로드 ===')
    driver.get('https://sell.smartstore.naver.com/#/products/bulkadd')
    time.sleep(6)
    close_popups()
    time.sleep(1)
    ss('50_bulkadd')

    try:
        btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[contains(text(),"양식 다운로드")]'))
        )
        print('양식 다운로드 클릭')
        btn.click()
        time.sleep(5)
        dl = wait_download(60)
        if dl:
            print(f'방법B 성공: {dl}')
            import openpyxl, csv
            if dl.endswith('.csv'):
                with open(dl, encoding='utf-8-sig') as fp:
                    h = next(csv.reader(fp))
                print(f'컬럼 {len(h)}개:', h[:30])
            else:
                wb = openpyxl.load_workbook(dl, read_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    h = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                    print(f'[{sn}] 컬럼 {len(h)}개:')
                    for i, v in enumerate(h):
                        if v: print(f'  {i+1}: {v}')
                wb.close()
            sys.exit(0)
        else:
            print('방법B 실패, 방법A 시도')
    except Exception as e:
        print(f'방법B 오류: {e}')

    # ── 방법 A: 수정양식 다운로드 ──
    print('\n=== 방법A: 수정양식 다운로드 ===')
    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(6)
    close_popups()

    # 검색
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(6)
    close_popups()

    # 모든 체크박스 정보 출력
    chk_info = driver.execute_script("""
        var inputs = document.querySelectorAll('input[type="checkbox"]');
        return Array.from(inputs).map(function(el, i) {
            return {i: i, name: el.name, id: el.id, cls: el.className.substring(0,40)};
        });
    """)
    print(f'전체 체크박스 {len(chk_info)}개:')
    for c in chk_info:
        print(f'  [{c["i"]}] name={c["name"]} id={c["id"]} cls={c["cls"][:30]}')

    # 상품 행 체크박스 = name이 비어있거나 다른 것
    clicked = driver.execute_script("""
        var inputs = Array.from(document.querySelectorAll('input[type="checkbox"]'));
        // productStatusTypes, channelServiceTypes 제외
        var productChk = inputs.filter(function(el) {
            return el.name !== 'productStatusTypes' && el.name !== 'channelServiceTypes';
        });
        var clicked = 0;
        productChk.slice(0, 5).forEach(function(el) {
            el.dispatchEvent(new MouseEvent('click', {bubbles: true}));
            clicked++;
        });
        return {total: productChk.length, clicked: clicked};
    """)
    print(f'상품 체크박스: {clicked}')
    time.sleep(2)
    ss('51_selected')

    # 엑셀 일괄작업 → 수정양식 다운로드
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"엑셀 일괄작업")]'))
    ).click()
    time.sleep(2)

    opt = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '//*[contains(text(),"수정양식 다운로드")]'))
    )
    opt.click()
    time.sleep(3)
    ss('52_after_click')

    try:
        msg = driver.find_element(By.XPATH, '//*[contains(text(),"없습니다") or contains(text(),"선택")]')
        print('메시지:', msg.text[:80])
    except: pass

    dl = wait_download(120)
    if dl:
        print(f'방법A 성공: {dl} ({os.path.getsize(dl):,}bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                h = next(csv.reader(fp))
            print(f'컬럼 {len(h)}개:')
            for i,v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                h = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                print(f'[{sn}] {len(h)}개:')
                for i,v in enumerate(h):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        print('방법A도 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
