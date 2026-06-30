"""
수정양식 다운로드 팝업 → 다운로드 버튼 클릭
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl, csv

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# 기존 파일 목록 저장 (새 파일만 감지)
existing = set(glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls'))

driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_new_download(timeout=120):
    for i in range(timeout):
        files = glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls')
        new = [f for f in files if f not in existing and not f.endswith('.crdownload') and os.path.getsize(f) > 1000]
        if new:
            return new[0]
        if i % 15 == 0:
            print(f'  대기 {i}s...')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)
    for _ in range(2):
        try:
            driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
            time.sleep(1)
        except: pass

    driver.get('https://sell.smartstore.naver.com/#/products/origin-list')
    time.sleep(7)
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//button[normalize-space(text())="검색"]'))
    ).click()
    time.sleep(7)
    ss('g0_searched')

    # 전체 선택 체크박스 (ag-selection-checkbox)
    all_chks = driver.find_elements(By.CSS_SELECTOR, 'input[type="checkbox"]')
    chk12 = all_chks[12]
    driver.execute_script("arguments[0].scrollIntoView({block:'center'})", chk12)
    time.sleep(0.5)
    ActionChains(driver).move_to_element(chk12).click().perform()
    time.sleep(1.5)

    checked = driver.execute_script("return document.querySelectorAll('input[type=\"checkbox\"]:checked').length")
    print(f'체크 수: {checked}')
    ss('g1_selected')

    # 엑셀 일괄작업 드롭다운
    driver.execute_script("""
        var els = document.querySelectorAll('button, div, li, a, span');
        for (var e of els) {
            if (e.textContent.trim() === '엑셀 일괄작업' && e.offsetParent) {
                e.click(); break;
            }
        }
    """)
    time.sleep(2)
    ss('g2_dropdown')

    # 수정양식 다운로드 메뉴 클릭
    driver.execute_script("""
        var els = document.querySelectorAll('a, li, div, span, button, ul li');
        for (var e of els) {
            if (e.textContent.trim() === '수정양식 다운로드' && e.offsetParent) {
                e.click(); break;
            }
        }
    """)
    time.sleep(3)
    ss('g3_popup')

    # 팝업 확인 - "다운로드" 버튼 찾아 클릭
    popup_text = driver.execute_script("""
        var modal = document.querySelector('.modal, [role="dialog"], .popup, .dialog, .layer');
        return modal ? modal.textContent.substring(0, 200) : '팝업 없음';
    """)
    print(f'팝업 텍스트: {popup_text[:100]}')

    # 팝업 내 "다운로드" 버튼 클릭
    result = driver.execute_script("""
        var btns = document.querySelectorAll('button, a');
        for (var b of btns) {
            var t = b.textContent.trim();
            if (t === '다운로드' && b.offsetParent) {
                b.click();
                return '클릭: ' + t;
            }
        }
        // 대소문자 무관 재시도
        for (var b of btns) {
            if (b.offsetParent && b.textContent.includes('다운로드')) {
                b.click();
                return '클릭(포함): ' + b.textContent.trim().substring(0, 30);
            }
        }
        return '버튼 없음';
    """)
    print(f'팝업 버튼: {result}')
    time.sleep(3)
    ss('g4_after_download_btn')

    # 파일 대기
    dl = wait_new_download(120)
    ss('g5_final')

    if dl:
        size = os.path.getsize(dl)
        print(f'\n성공! {dl} ({size:,}bytes)')
        wb = openpyxl.load_workbook(dl, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f'\n[{sn}] 크기: {ws.max_row}행 × {ws.max_column}열')
            # 1-3행 출력
            for ri in range(1, 4):
                r = [ws.cell(ri, c).value for c in range(1, min(ws.max_column+1, 200))]
                vals = [str(v)[:20] for v in r if v is not None]
                if vals:
                    print(f'  행{ri}: {vals[:20]}')
            # 2행 전체 컬럼 (헤더)
            h = [ws.cell(2, c).value for c in range(1, ws.max_column+1)]
            print(f'\n  전체 컬럼 ({len([v for v in h if v])}개):')
            for i, v in enumerate(h):
                if v is not None:
                    print(f'    {i+1}: {v}')
        wb.close()
    else:
        print('다운로드 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
