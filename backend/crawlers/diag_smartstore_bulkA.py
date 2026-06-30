"""
JS click으로 양식 다운로드
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=90):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 100]
        if done:
            return done[0]
        if i % 10 == 0:
            print(f'  대기 {i}s')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)

    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    driver.get('https://sell.smartstore.naver.com/#/products/bulkadd')
    time.sleep(7)

    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    ss('70_bulkadd')

    # JS click으로 양식 다운로드
    result = driver.execute_script("""
        var btns = document.querySelectorAll('button');
        for (var b of btns) {
            if (b.textContent.trim().includes('양식 다운로드')) {
                b.click();
                return '클릭완료: ' + b.textContent.trim();
            }
        }
        return '버튼없음';
    """)
    print('JS click 결과:', result)
    time.sleep(3)
    ss('71_after_click')

    # 모달이 떴는지 확인
    try:
        modal = driver.find_element(By.XPATH, '//*[contains(@class,"modal") and contains(@style,"display: block")] | //*[contains(@class,"modal-open")]')
        print('모달 텍스트:', modal.text[:200])
        ss('72_modal')

        # 모달 내 다운로드/확인 버튼 클릭
        for kw in ['다운로드', '확인', 'OK']:
            try:
                btn = modal.find_element(By.XPATH, f'.//button[contains(text(),"{kw}")]')
                print(f'모달 버튼 클릭: {btn.text}')
                btn.click()
                time.sleep(3)
                break
            except: pass
    except:
        print('모달 없음')

    dl = wait_download(90)
    ss('73_final')

    if dl:
        print(f'\n성공: {dl} ({os.path.getsize(dl):,} bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                h = next(csv.reader(fp))
            print(f'컬럼 {len(h)}개:')
            for i,v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                h = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                print(f'\n[{sn}] {ws.max_column}컬럼:')
                for i,v in enumerate(h):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        print('다운로드 실패')

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
