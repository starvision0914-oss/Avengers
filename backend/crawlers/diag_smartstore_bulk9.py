"""
상품 일괄등록 → 양식 다운로드 단독 실행
"""
import os, sys, time, glob
sys.path.insert(0, '/home/rejoice888/Avengers/backend')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
import django; django.setup()

from crawlers.browser import create_driver
from crawlers.smartstore_crawler import login_smartstore
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

LOGIN_ID = 'dlrmsgh01234@gmail.com'
from apps.smartstore.models import SmartStoreAccount
acc = SmartStoreAccount.objects.get(login_id=LOGIN_ID)

DOWNLOAD_DIR = '/tmp/smartstore_excel'
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
driver = create_driver(download_dir=DOWNLOAD_DIR)

def ss(name):
    driver.save_screenshot(f'{DOWNLOAD_DIR}/{name}.png')
    print(f'SS: {name}')

def wait_download(timeout=90):
    for i in range(timeout):
        files = (glob.glob(f'{DOWNLOAD_DIR}/*.xlsx') + glob.glob(f'{DOWNLOAD_DIR}/*.xls') +
                 glob.glob(f'{DOWNLOAD_DIR}/*.csv'))
        done = [f for f in files if not f.endswith('.crdownload') and os.path.getsize(f) > 100]
        if done:
            return done[0]
        if i % 10 == 0:
            print(f'  대기 {i}s')
        time.sleep(1)
    return None

try:
    ok = login_smartstore(driver, acc.login_id, acc.login_pw, print)
    if not ok:
        sys.exit(1)
    time.sleep(2)

    # 팝업 닫기
    try:
        driver.find_element(By.XPATH, '//*[contains(text(),"하루동안")]').click()
        time.sleep(1)
    except: pass

    # 상품 일괄등록
    driver.get('https://sell.smartstore.naver.com/#/products/bulkadd')
    time.sleep(7)

    # 팝업 닫기
    for xpath in ['//*[contains(text(),"하루동안")]', '//button[text()="닫기"]',
                  '//*[@aria-label="닫기"]', '//*[contains(@class,"close")][@type="button"]']:
        try:
            driver.find_element(By.XPATH, xpath).click()
            time.sleep(0.5)
        except: pass

    ss('60_bulkadd_clean')

    # 모든 버튼 텍스트 출력
    btns = driver.find_elements(By.TAG_NAME, 'button')
    print('버튼 목록:')
    for b in btns:
        t = b.text.strip()
        if t:
            print(f'  "{t}"')

    # 양식 다운로드 버튼 찾아서 클릭
    btn = None
    for xpath in [
        '//button[contains(text(),"양식 다운로드")]',
        '//button[normalize-space(text())="양식 다운로드"]',
        '//*[contains(@class,"download") and contains(text(),"양식")]',
        '//a[contains(text(),"양식 다운로드")]',
    ]:
        try:
            btn = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            print(f'버튼 발견: {xpath}')
            break
        except: pass

    if not btn:
        print('양식 다운로드 버튼 없음')
        ss('61_no_btn')
        sys.exit(1)

    # 클릭 전 스크린샷
    ss('61_before_click')

    # 클릭
    driver.execute_script("arguments[0].scrollIntoView();", btn)
    time.sleep(0.5)
    btn.click()
    print('클릭 완료')
    time.sleep(3)
    ss('62_after_click')

    # 팝업 확인/닫기
    try:
        confirm = driver.find_element(By.XPATH, '//button[contains(text(),"확인") or contains(text(),"다운")]')
        print('확인 팝업:', confirm.text)
        confirm.click()
        time.sleep(2)
    except: pass

    # 다운로드 대기
    dl = wait_download(90)
    ss('63_final')

    if dl:
        print(f'\n성공: {dl} ({os.path.getsize(dl):,} bytes)')
        import openpyxl, csv
        if dl.endswith('.csv'):
            with open(dl, encoding='utf-8-sig') as fp:
                r = csv.reader(fp)
                h = next(r)
            print(f'컬럼 {len(h)}개:')
            for i, v in enumerate(h): print(f'  {i+1}: {v}')
        else:
            wb = openpyxl.load_workbook(dl, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                h = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
                print(f'\n[{sn}] {ws.max_column}컬럼:')
                for i, v in enumerate(h):
                    if v: print(f'  {i+1}: {v}')
            wb.close()
    else:
        print('다운로드 실패')
        # 화면에 뭔가 있는지 확인
        links = driver.find_elements(By.XPATH, '//a[@href and contains(@href,".xlsx")]')
        print('엑셀 링크:', [l.get_attribute('href') for l in links[:3]])

finally:
    time.sleep(2)
    driver.quit()
    print('\n완료')
