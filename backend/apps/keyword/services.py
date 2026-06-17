import io
import os
import csv
import sys
import subprocess
import tempfile
import zipfile
import logging
from datetime import datetime

import openpyxl
from django.conf import settings
from django.db import connections

from .models import KeywordTask

logger = logging.getLogger(__name__)

# 엑셀 컬럼 인덱스 (0-based) → DB 필드명 매핑
EXCEL_COL_MAP = {
    0: 'seller_code1',
    1: 'seller_code2',
    # 2 = product_code (별도 처리)
    3: 'category_code',
    4: 'category_name',
    5: 'market_category',
    6: 'product_name',
    7: 'market_product_name',
    8: 'ownerclan_price',
    9: 'consumer_price',
    10: 'market_price',
    11: 'shipping_fee',
    12: 'shipping_type',
    13: 'min_qty',
    14: 'max_qty',
    15: 'company_notice',
    16: 'special_notice',
    17: 'option1_name',
    18: 'option1_values',
    19: 'option2_name',
    20: 'option2_values',
    21: 'combined_option',
    22: 'product_attribute',
    23: 'product_grade',
    24: 'tax_type',
    25: 'compliance',
    26: 'age_restriction',
    27: 'return_possible',
    28: 'image_large',
    29: 'image_medium',
    30: 'image_small',
    31: 'manufacturer',
    32: 'brand',
    33: 'model_name',
    34: 'origin',
    35: 'keywords',
    36: 'registered_at',
    37: 'modified_at',
    38: 'header_text',
    39: 'detail_html',
    40: 'notice_code',
    41: 'notice_category',
    42: 'notice_info',
    43: 'notice_html',
    44: 'market_gmarket',
    45: 'market_auction',
    46: 'market_11st',
    47: 'market_coupang',
    48: 'market_smartstore',
    49: 'market_promo',
    50: 'market_gift',
    51: 'certification_type',
    52: 'certification_info',
    53: 'return_fee',
    54: 'independent_option',
    55: 'combined_option_detail',
    # 56 = 상품상태 (BE열, 빈값이지만 일단 무시)
}

# INT 타입 필드
INT_FIELDS = {
    'ownerclan_price', 'consumer_price', 'market_price',
    'shipping_fee', 'min_qty', 'max_qty', 'return_fee',
}

# DATETIME 타입 필드
DATETIME_FIELDS = {'registered_at', 'modified_at'}

# 비교 대상 필드 목록 (orig_ 접두사로 비교)
TRACKABLE_FIELDS = list(EXCEL_COL_MAP.values())


def _dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _safe_str(val):
    if val is None:
        return ''
    return str(val).strip()


def _safe_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _safe_datetime(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_excel_row(row_values):
    """엑셀 1행(0-based list) → dict 변환"""
    data = {}
    for col_idx, field_name in EXCEL_COL_MAP.items():
        raw = row_values[col_idx] if col_idx < len(row_values) else None
        if field_name in INT_FIELDS:
            data[field_name] = _safe_int(raw)
        elif field_name in DATETIME_FIELDS:
            data[field_name] = _safe_datetime(raw)
        else:
            data[field_name] = _safe_str(raw)
    return data


def _extract_excel_from_upload(uploaded_file):
    """업로드 파일(xlsx 또는 zip)에서 openpyxl Workbook 반환"""
    name = uploaded_file.name.lower()
    content = uploaded_file.read()

    if name.endswith('.zip'):
        zf = zipfile.ZipFile(io.BytesIO(content))
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith('.xlsx')]
        if not xlsx_names:
            raise ValueError('ZIP 안에 .xlsx 파일이 없습니다.')
        xlsx_bytes = zf.read(xlsx_names[0])
        return openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    elif name.endswith('.xlsx'):
        return openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    else:
        raise ValueError('xlsx 또는 zip 파일만 업로드 가능합니다.')


def _field_changed(old_val, new_val, field_name):
    """두 값이 다른지 비교 (NULL/빈문자열 통일)"""
    if field_name in INT_FIELDS:
        return _safe_int(old_val) != _safe_int(new_val)
    if field_name in DATETIME_FIELDS:
        a = _safe_datetime(old_val)
        b = _safe_datetime(new_val)
        if a and b:
            return a.strftime('%Y-%m-%d %H:%M:%S') != b.strftime('%Y-%m-%d %H:%M:%S')
        return (a is None) != (b is None)
    return _safe_str(old_val) != _safe_str(new_val)


def upload_excel(uploaded_file):
    """엑셀/ZIP 업로드 → DB UPSERT, 결과 반환"""
    wb = _extract_excel_from_upload(uploaded_file)
    ws = wb.active
    now = datetime.now()

    inserted = 0
    updated = 0
    skipped = 0

    # 기존 상품 전체 조회 (product_code → id 매핑)
    existing = {}
    with connections['default'].cursor() as cur:
        cur.execute("SELECT id, product_code FROM keyword_product")
        for row in cur.fetchall():
            existing[row[1]] = row[0]

    rows_to_process = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 3:
            continue
        product_code = _safe_str(row[2])
        if not product_code:
            continue
        data = _parse_excel_row(list(row))
        rows_to_process.append((product_code, data))

    wb.close()

    # 모든 현재 필드명
    fields = list(EXCEL_COL_MAP.values())
    orig_fields = [f'orig_{f}' for f in fields]

    with connections['default'].cursor() as cur:
        for product_code, data in rows_to_process:
            if product_code in existing:
                # UPDATE: 현재값만 업데이트, orig_ 안 건드림
                # 먼저 기존 현재값 읽어서 변경 여부 확인
                pid = existing[product_code]
                cur.execute(
                    f"SELECT {', '.join(fields)} FROM keyword_product WHERE id=%s",
                    [pid],
                )
                old_row = cur.fetchone()
                old_data = dict(zip(fields, old_row))

                # 현재값 변경 여부 체크
                any_current_changed = False
                for f in fields:
                    if _field_changed(old_data[f], data[f], f):
                        any_current_changed = True
                        break

                if not any_current_changed:
                    skipped += 1
                    continue

                # 현재값 업데이트
                set_parts = [f"{f}=%s" for f in fields]
                set_parts.append("uploaded_at=%s")
                vals = [data[f] for f in fields] + [now]

                cur.execute(
                    f"UPDATE keyword_product SET {', '.join(set_parts)} WHERE id=%s",
                    vals + [pid],
                )

                # orig vs 새 현재값 비교 → is_synced 결정
                cur.execute(
                    f"SELECT {', '.join(orig_fields)} FROM keyword_product WHERE id=%s",
                    [pid],
                )
                orig_row = cur.fetchone()
                orig_data = dict(zip(fields, orig_row))  # orig_ 접두사 제거한 키

                is_synced = 1
                for f in fields:
                    if _field_changed(orig_data[f], data[f], f):
                        is_synced = 0
                        break

                cur.execute(
                    "UPDATE keyword_product SET is_synced=%s WHERE id=%s",
                    [is_synced, pid],
                )
                updated += 1
            else:
                # INSERT: 현재값 = orig값 (동일하게)
                all_fields = ['product_code'] + fields + orig_fields + [
                    'sale_status', 'is_synced', 'uploaded_at',
                ]
                placeholders = ', '.join(['%s'] * len(all_fields))
                vals = (
                    [product_code]
                    + [data[f] for f in fields]
                    + [data[f] for f in fields]  # orig = 현재와 동일
                    + [1, 1, now]
                )
                cur.execute(
                    f"INSERT INTO keyword_product ({', '.join(all_fields)}) "
                    f"VALUES ({placeholders})",
                    vals,
                )
                existing[product_code] = cur.lastrowid
                inserted += 1

    return {
        'inserted': inserted,
        'updated': updated,
        'skipped': skipped,
        'total': inserted + updated + skipped,
    }


def upload_excel_async(uploaded_file):
    suffix = '.zip' if uploaded_file.name.lower().endswith('.zip') else '.xlsx'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, prefix='keyword_')
    for chunk in uploaded_file.chunks():
        tmp.write(chunk)
    tmp.close()

    task = KeywordTask.objects.create(
        task_type='keyword_upload',
        status='pending',
        input_data={'file_path': tmp.name, 'filename': uploaded_file.name},
    )

    manage_py = os.path.join(settings.BASE_DIR, 'manage.py')
    proc = subprocess.Popen(
        [sys.executable, manage_py, 'keyword_upload', str(task.id)],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    task.pid = proc.pid
    task.save(update_fields=['pid'])

    return {'task_id': task.id, 'status': 'pending'}


def upload_csv_status(uploaded_file):
    """CSV 업로드 → sale_status 업데이트"""
    content = uploaded_file.read()
    # EUC-KR 또는 UTF-8
    for enc in ('euc-kr', 'utf-8', 'cp949'):
        try:
            text = content.decode(enc, errors='replace')
            break
        except Exception:
            continue
    else:
        text = content.decode('utf-8', errors='replace')

    reader = csv.reader(io.StringIO(text))
    next(reader, None)  # 헤더 라인
    next(reader, None)  # 컬럼 헤더

    STATUS_MAP = {
        '품절': 2,
        '단종': 3,
        '유통금지': 3,
        '옵션 품절': 2,
        '옵션 단종': 3,
        '재입고': 1,
        '옵션 재입고': 1,
    }

    updated = 0
    with connections['default'].cursor() as cur:
        for row in reader:
            if len(row) < 8:
                continue
            code = row[1].strip()
            status_text = row[7].strip()
            if not code:
                continue

            sale_status = STATUS_MAP.get(status_text)
            if sale_status is not None:
                cur.execute(
                    "UPDATE keyword_product SET sale_status=%s WHERE product_code=%s",
                    [sale_status, code],
                )
                if cur.rowcount > 0:
                    updated += 1

    return {'updated': updated}


def sync_products(product_ids=None):
    """동기화: orig_ = 현재값, is_synced = 1"""
    fields = list(EXCEL_COL_MAP.values())
    set_parts = [f"orig_{f} = {f}" for f in fields]
    set_parts.append("is_synced = 1")
    set_parts.append("synced_at = %s")

    now = datetime.now()

    with connections['default'].cursor() as cur:
        if product_ids:
            placeholders = ','.join(['%s'] * len(product_ids))
            cur.execute(
                f"UPDATE keyword_product SET {', '.join(set_parts)} "
                f"WHERE id IN ({placeholders})",
                [now] + product_ids,
            )
        else:
            cur.execute(
                f"UPDATE keyword_product SET {', '.join(set_parts)} "
                f"WHERE is_synced = 0",
                [now],
            )
        count = cur.rowcount

    return {'synced': count}


def _changed_field_condition(field_name):
    """필드별 orig vs 현재 비교 SQL 조건"""
    if field_name in INT_FIELDS:
        return f"COALESCE({field_name},0) != COALESCE(orig_{field_name},0)"
    elif field_name in DATETIME_FIELDS:
        return f"COALESCE(CAST({field_name} AS CHAR),'') != COALESCE(CAST(orig_{field_name} AS CHAR),'')"
    else:
        return f"COALESCE({field_name},'') != COALESCE(orig_{field_name},'')"


def _build_where(sale_status=None, is_synced=None, search=None, changed_field=None,
                 filter_col=None, filter_vals=None, codes=None):
    """공통 WHERE 절 생성"""
    where = ['1=1']
    params = []

    if sale_status is not None:
        where.append('sale_status = %s')
        params.append(int(sale_status))
    if is_synced is not None:
        where.append('is_synced = %s')
        params.append(int(is_synced))
    if search:
        where.append('(product_code LIKE %s OR product_name LIKE %s)')
        like = f'%{search}%'
        params.extend([like, like])
    if changed_field:
        if changed_field in TRACKABLE_FIELDS:
            where.append(_changed_field_condition(changed_field))
        elif changed_field == '__any__':
            where.append('is_synced = 0')
    if filter_col and filter_vals and filter_col in FILTERABLE_COLUMNS:
        empty_marker = '(빈값)'
        clean = [v for v in filter_vals if v != empty_marker]
        include_empty = empty_marker in filter_vals
        clauses = []
        if clean:
            placeholders = ','.join(['%s'] * len(clean))
            clauses.append(f'{filter_col} IN ({placeholders})')
            params.extend(clean)
        if include_empty:
            clauses.append(f'({filter_col} IS NULL OR {filter_col} = "")')
        if clauses:
            where.append('(' + ' OR '.join(clauses) + ')')
    if codes:
        clean_codes = [c.strip() for c in codes if c and c.strip()]
        if clean_codes:
            placeholders = ','.join(['%s'] * len(clean_codes))
            where.append(f'product_code IN ({placeholders})')
            params.extend(clean_codes)

    return ' AND '.join(where), params


SORTABLE_COLUMNS = {
    'product_code', 'category_code', 'ownerclan_price',
    'market_price', 'shipping_fee', 'uploaded_at',
}


def get_products(page=1, per_page=50, sale_status=None, is_synced=None,
                 search=None, changed_field=None, sort=None, order='asc',
                 filter_col=None, filter_vals=None, codes=None):
    """상품 목록 조회"""
    where_sql, params = _build_where(
        sale_status, is_synced, search, changed_field,
        filter_col=filter_col, filter_vals=filter_vals, codes=codes,
    )

    sort_col = sort if sort in SORTABLE_COLUMNS else 'product_code'
    sort_dir = 'DESC' if str(order).lower() == 'desc' else 'ASC'

    offset = (page - 1) * per_page

    with connections['default'].cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM keyword_product WHERE {where_sql}", params)
        total = cur.fetchone()[0]

        cur.execute(
            f"SELECT keyword_product.id, keyword_product.product_code, "
            f"keyword_product.product_name, keyword_product.orig_product_name, "
            f"keyword_product.market_product_name, keyword_product.orig_market_product_name, "
            f"keyword_product.ownerclan_price, keyword_product.orig_ownerclan_price, "
            f"keyword_product.market_price, keyword_product.orig_market_price, "
            f"keyword_product.shipping_fee, keyword_product.orig_shipping_fee, "
            f"keyword_product.return_fee, keyword_product.orig_return_fee, "
            f"keyword_product.image_large, keyword_product.orig_image_large, "
            f"keyword_product.image_small, "
            f"keyword_product.sale_status, keyword_product.is_synced, "
            f"keyword_product.category_code, keyword_product.category_name, "
            f"keyword_product.manufacturer, keyword_product.origin, "
            f"keyword_product.uploaded_at, keyword_product.synced_at, keyword_product.created_at, "
            f"COALESCE(d.cnt, 0) AS dup_count "
            f"FROM keyword_product "
            f"LEFT JOIN ("
            f"  SELECT product_name AS dup_name, COUNT(*) AS cnt FROM keyword_product "
            f"  WHERE product_name IS NOT NULL AND product_name != '' "
            f"  GROUP BY product_name HAVING COUNT(*) > 1"
            f") d ON keyword_product.product_name = d.dup_name "
            f"WHERE {where_sql} "
            f"ORDER BY (d.cnt IS NOT NULL) DESC, {sort_col} {sort_dir}, keyword_product.product_code ASC "
            f"LIMIT %s OFFSET %s",
            params + [per_page, offset],
        )
        rows = _dictfetchall(cur)

    # datetime 직렬화
    for r in rows:
        for k in ('uploaded_at', 'synced_at', 'created_at'):
            if r.get(k) and isinstance(r[k], datetime):
                r[k] = r[k].isoformat()

    return {
        'items': rows,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page if total > 0 else 0,
    }


def get_product_detail(product_id):
    """상품 상세 (전 필드 orig + 현재)"""
    fields = list(EXCEL_COL_MAP.values())
    select_parts = ['id', 'product_code', 'sale_status', 'is_synced',
                     'synced_at', 'uploaded_at', 'created_at', 'updated_at']
    for f in fields:
        select_parts.append(f)
        select_parts.append(f'orig_{f}')

    with connections['default'].cursor() as cur:
        cur.execute(
            f"SELECT {', '.join(select_parts)} FROM keyword_product WHERE id=%s",
            [product_id],
        )
        rows = _dictfetchall(cur)
        if not rows:
            return None
        row = rows[0]

    # datetime 직렬화
    for k, v in row.items():
        if isinstance(v, datetime):
            row[k] = v.isoformat()

    # 변경된 필드 목록
    changed_fields = []
    for f in fields:
        if _field_changed(row.get(f'orig_{f}'), row.get(f), f):
            changed_fields.append(f)

    row['changed_fields'] = changed_fields
    return row


def get_stats():
    """통계"""
    with connections['default'].cursor() as cur:
        cur.execute(
            "SELECT "
            "COUNT(*) as total, "
            "SUM(sale_status=1) as selling, "
            "SUM(sale_status=2) as soldout, "
            "SUM(sale_status=3) as discontinued, "
            "SUM(is_synced=0) as changed "
            "FROM keyword_product"
        )
        row = _dictfetchall(cur)[0]
    return {k: int(v or 0) for k, v in row.items()}


def get_changed_field_counts():
    """변경 항목별 카운트 (필터 드롭다운용)"""
    counts = {}
    with connections['default'].cursor() as cur:
        for f in TRACKABLE_FIELDS:
            cond = _changed_field_condition(f)
            cur.execute(f"SELECT COUNT(*) FROM keyword_product WHERE {cond}")
            cnt = cur.fetchone()[0]
            if cnt > 0:
                counts[f] = cnt
    return counts


def get_products_for_export(sale_status=None, is_synced=None, search=None, changed_field=None):
    """엑셀 내보내기용 전체 조회 (페이지네이션 없음)"""
    where_sql, params = _build_where(sale_status, is_synced, search, changed_field)

    fields = list(EXCEL_COL_MAP.values())
    orig_fields = [f'orig_{f}' for f in fields]

    select_cols = (
        ['product_code', 'sale_status', 'is_synced']
        + fields + orig_fields
        + ['uploaded_at', 'synced_at']
    )

    with connections['default'].cursor() as cur:
        cur.execute(
            f"SELECT {', '.join(select_cols)} "
            f"FROM keyword_product WHERE {where_sql} "
            f"ORDER BY product_code",
            params,
        )
        rows = _dictfetchall(cur)

    return rows


def upload_soldout_txt(uploaded_file):
    """품절/단종 TXT 업로드 → keyword_product sale_status=2 + smartstore_product ownerclan_soldout=1"""
    content = uploaded_file.read()
    for enc in ('utf-8', 'euc-kr', 'cp949'):
        try:
            text = content.decode(enc)
            break
        except Exception:
            continue
    else:
        text = content.decode('utf-8', errors='replace')

    soldout_codes = {line.strip() for line in text.splitlines() if line.strip()}

    # keyword_product 전체 product_code 로드
    with connections['default'].cursor() as cur:
        cur.execute("SELECT product_code FROM keyword_product")
        db_codes = {r[0] for r in cur.fetchall()}

    matched = soldout_codes & db_codes
    today = datetime.now().strftime('%Y-%m-%d')

    # keyword_product 배치 UPDATE (1000개씩)
    ownerclan_updated = 0
    batch = list(matched)
    with connections['default'].cursor() as cur:
        for i in range(0, len(batch), 1000):
            chunk = batch[i:i + 1000]
            placeholders = ','.join(['%s'] * len(chunk))
            cur.execute(
                f"UPDATE keyword_product SET sale_status=2, modified_at=%s "
                f"WHERE product_code IN ({placeholders}) AND sale_status != 2",
                [today] + chunk,
            )
            ownerclan_updated += cur.rowcount

    # smartstore_product 외부 DB 연동은 Avengers 단일DB 정책으로 비활성화
    smartstore_updated = 0

    return {
        'total_codes': len(soldout_codes),
        'ownerclan_matched': len(matched),
        'ownerclan_updated': ownerclan_updated,
        'smartstore_updated': smartstore_updated,
    }


def get_w_codes(sale_status=None, is_synced=None, search=None, changed_field=None):
    """필터 조건에 맞는 W코드(product_code) 목록"""
    where_sql, params = _build_where(sale_status, is_synced, search, changed_field)

    with connections['default'].cursor() as cur:
        cur.execute(
            f"SELECT product_code FROM keyword_product "
            f"WHERE {where_sql} ORDER BY product_code",
            params,
        )
        return [row[0] for row in cur.fetchall()]


def delete_all_products():
    """모든 상품 삭제"""
    with connections['default'].cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM keyword_product")
        before = cur.fetchone()[0]
        cur.execute("DELETE FROM keyword_product")
    return {'deleted': before}


def delete_products_by_ids(ids):
    """id 리스트로 일괄 삭제"""
    ids = [int(i) for i in ids if i]
    if not ids:
        return {'deleted': 0}
    placeholders = ','.join(['%s'] * len(ids))
    with connections['default'].cursor() as cur:
        cur.execute(
            f"DELETE FROM keyword_product WHERE id IN ({placeholders})",
            ids,
        )
        deleted = cur.rowcount
    return {'deleted': deleted}


def dedupe_by_product_name():
    """상품명이 같은 상품 중 ownerclan_price가 더 높은 것을 삭제.
    같은 가격이면 id가 큰 것을 삭제(가장 오래된/작은 id를 보존).
    """
    with connections['default'].cursor() as cur:
        cur.execute("""
            DELETE FROM keyword_product
            WHERE id IN (
                SELECT id FROM (
                    SELECT p.id
                    FROM keyword_product p
                    INNER JOIN (
                        SELECT product_name,
                               MIN(ownerclan_price) AS min_price
                        FROM keyword_product
                        WHERE product_name IS NOT NULL AND product_name != ''
                        GROUP BY product_name
                        HAVING COUNT(*) > 1
                    ) m ON p.product_name = m.product_name
                    WHERE p.ownerclan_price > m.min_price
                ) t
            )
        """)
        deleted_high = cur.rowcount

        cur.execute("""
            DELETE FROM keyword_product
            WHERE id IN (
                SELECT id FROM (
                    SELECT p.id
                    FROM keyword_product p
                    INNER JOIN (
                        SELECT product_name, ownerclan_price, MIN(id) AS keep_id
                        FROM keyword_product
                        WHERE product_name IS NOT NULL AND product_name != ''
                        GROUP BY product_name, ownerclan_price
                        HAVING COUNT(*) > 1
                    ) m
                      ON p.product_name = m.product_name
                     AND p.ownerclan_price = m.ownerclan_price
                    WHERE p.id > m.keep_id
                ) t
            )
        """)
        deleted_dup = cur.rowcount

    return {
        'deleted': deleted_high + deleted_dup,
        'higher_price_removed': deleted_high,
        'same_price_removed': deleted_dup,
    }


FILTERABLE_COLUMNS = {
    'category_name', 'category_code', 'manufacturer',
    'origin', 'shipping_type', 'brand',
}


def get_distinct_values(column):
    """화이트리스트 컬럼의 distinct 값 + 각 값의 상품 수 반환"""
    if column not in FILTERABLE_COLUMNS:
        raise ValueError(f'필터 불가 컬럼: {column}')
    with connections['default'].cursor() as cur:
        cur.execute(
            f"SELECT COALESCE(NULLIF({column}, ''), '(빈값)') AS v, COUNT(*) AS c "
            f"FROM keyword_product GROUP BY v ORDER BY c DESC, v ASC"
        )
        return [{'value': r[0], 'count': r[1]} for r in cur.fetchall()]
