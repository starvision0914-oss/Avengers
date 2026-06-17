import io
import os
import traceback
import zipfile
from datetime import datetime

import openpyxl
from django.core.management.base import BaseCommand
from django.db import connections

from apps.keyword.models import KeywordTask
from apps.keyword.services import (
    EXCEL_COL_MAP, INT_FIELDS, DATETIME_FIELDS,
    _parse_excel_row, _safe_str, _field_changed,
)

PROGRESS_INTERVAL = 1000


class Command(BaseCommand):
    help = '키워드추출기 상품대장 비동기 업로드 워커'

    def add_arguments(self, parser):
        parser.add_argument('task_id', type=int)

    def handle(self, *args, **options):
        task_id = options['task_id']
        try:
            task = KeywordTask.objects.get(pk=task_id)
        except KeywordTask.DoesNotExist:
            return

        task.status = 'running'
        task.pid = os.getpid()
        task.save(update_fields=['status', 'pid'])

        file_path = task.input_data.get('file_path', '')

        try:
            result = _process_upload(file_path, task)
            task.result_data = result
            task.status = 'done'
        except Exception:
            task.result_data = {'error': traceback.format_exc()}
            task.status = 'error'
        finally:
            task.save(update_fields=['status', 'result_data'])
            try:
                os.unlink(file_path)
            except Exception:
                pass


def _load_workbooks(file_path):
    name = file_path.lower()
    if name.endswith('.zip'):
        with open(file_path, 'rb') as f:
            zf = zipfile.ZipFile(io.BytesIO(f.read()))
        all_names = zf.namelist()
        xlsx_names = sorted(n for n in all_names if n.lower().endswith(('.xlsx', '.xlsm')))
        if not xlsx_names:
            preview = ', '.join(all_names[:10]) or '(empty)'
            raise ValueError(
                f'ZIP 안에 .xlsx/.xlsm 파일이 없습니다. '
                f'발견된 파일({len(all_names)}개): {preview}'
            )
        wbs = []
        for xn in xlsx_names:
            xlsx_bytes = zf.read(xn)
            wbs.append((xn, openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)))
        return wbs
    elif name.endswith(('.xlsx', '.xlsm')):
        return [(os.path.basename(file_path), openpyxl.load_workbook(file_path, read_only=True))]
    else:
        raise ValueError(
            f'xlsx 또는 zip 파일만 업로드 가능합니다. (받은 파일: {os.path.basename(file_path)})'
        )


def _process_upload(file_path, task):
    workbooks = _load_workbooks(file_path)
    now = datetime.now()

    inserted = 0
    updated = 0
    skipped = 0

    existing = {}
    with connections['default'].cursor() as cur:
        cur.execute("SELECT id, product_code FROM keyword_product")
        for row in cur.fetchall():
            existing[row[1]] = row[0]

    rows_to_process = []
    for wb_name, wb in workbooks:
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 3:
                continue
            product_code = _safe_str(row[2])
            if not product_code:
                continue
            data = _parse_excel_row(list(row))
            rows_to_process.append((product_code, data))
        wb.close()

    total_rows = len(rows_to_process)
    task.result_data = {
        'progress': 0, 'inserted': 0, 'updated': 0,
        'skipped': 0, 'total_rows': total_rows,
    }
    task.save(update_fields=['result_data'])

    fields = list(EXCEL_COL_MAP.values())
    orig_fields = [f'orig_{f}' for f in fields]

    with connections['default'].cursor() as cur:
        for idx, (product_code, data) in enumerate(rows_to_process, 1):
            if product_code in existing:
                pid = existing[product_code]
                cur.execute(
                    f"SELECT {', '.join(fields)} FROM keyword_product WHERE id=%s",
                    [pid],
                )
                old_row = cur.fetchone()
                old_data = dict(zip(fields, old_row))

                any_current_changed = False
                for f in fields:
                    if _field_changed(old_data[f], data[f], f):
                        any_current_changed = True
                        break

                if not any_current_changed:
                    skipped += 1
                else:
                    set_parts = [f"{f}=%s" for f in fields]
                    set_parts.append("uploaded_at=%s")
                    vals = [data[f] for f in fields] + [now]

                    cur.execute(
                        f"UPDATE keyword_product SET {', '.join(set_parts)} WHERE id=%s",
                        vals + [pid],
                    )

                    cur.execute(
                        f"SELECT {', '.join(orig_fields)} FROM keyword_product WHERE id=%s",
                        [pid],
                    )
                    orig_row = cur.fetchone()
                    orig_data = dict(zip(fields, orig_row))

                    is_synced = 1
                    for f in fields:
                        if _field_changed(orig_data[f], data[f], f):
                            is_synced = 0
                            break

                    cur.execute(
                        "UPDATE keyword_product SET is_synced=%s WHERE id=%s",
                        [is_synced, pid],
                    )
                    updated += 1
            else:
                all_fields = ['product_code'] + fields + orig_fields + [
                    'sale_status', 'is_synced', 'uploaded_at',
                ]
                placeholders = ', '.join(['%s'] * len(all_fields))
                vals = (
                    [product_code]
                    + [data[f] for f in fields]
                    + [data[f] for f in fields]
                    + [1, 1, now]
                )
                cur.execute(
                    f"INSERT INTO keyword_product ({', '.join(all_fields)}) "
                    f"VALUES ({placeholders})",
                    vals,
                )
                existing[product_code] = cur.lastrowid
                inserted += 1

            if idx % PROGRESS_INTERVAL == 0 or idx == total_rows:
                progress = int(idx * 100 / total_rows) if total_rows else 100
                task.result_data = {
                    'progress': progress,
                    'inserted': inserted,
                    'updated': updated,
                    'skipped': skipped,
                    'total_rows': total_rows,
                }
                task.save(update_fields=['result_data'])

    return {
        'inserted': inserted,
        'updated': updated,
        'skipped': skipped,
        'total': inserted + updated + skipped,
    }
