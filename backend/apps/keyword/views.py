import io

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse

from . import services
from .models import KeywordTask


class KeywordProductUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)

        running = KeywordTask.objects.filter(
            task_type='ownerclan_upload', status__in=('pending', 'running')
        ).first()
        if running:
            return Response({
                'error': '이미 업로드 처리 중입니다.',
                'task_id': running.id,
            }, status=409)

        try:
            result = services.upload_excel_async(f)
            return Response(result, status=202)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=400)
        try:
            task = KeywordTask.objects.get(pk=int(task_id))
        except KeywordTask.DoesNotExist:
            return Response({'error': 'not found'}, status=404)
        return Response({
            'task_id': task.id,
            'status': task.status,
            'result_data': task.result_data,
        })


class KeywordProductCsvUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        try:
            result = services.upload_csv_status(f)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class KeywordSoldoutTxtUploadView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        try:
            result = services.upload_soldout_txt(f)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class KeywordProductListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None
        sort = request.query_params.get('sort') or None
        order = request.query_params.get('order') or 'asc'
        filter_col = request.query_params.get('filter_col') or None
        filter_vals_raw = request.query_params.get('filter_vals') or ''
        filter_vals = [v for v in filter_vals_raw.split('|') if v != ''] if filter_vals_raw else None
        codes_raw = request.query_params.get('codes') or ''
        codes = [c.strip() for c in codes_raw.split(',') if c.strip()] if codes_raw else None
        result = services.get_products(
            page, per_page,
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
            sort=sort,
            order=order,
            filter_col=filter_col,
            filter_vals=filter_vals,
            codes=codes,
        )
        return Response(result)


class KeywordProductDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        result = services.get_product_detail(pk)
        if not result:
            return Response({'error': '상품을 찾을 수 없습니다.'}, status=404)
        return Response(result)


class KeywordProductSyncView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        product_ids = request.data.get('product_ids')
        if product_ids and isinstance(product_ids, list):
            product_ids = [int(i) for i in product_ids]
        else:
            product_ids = None
        result = services.sync_products(product_ids)
        return Response(result)


class KeywordProductStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(services.get_stats())


class KeywordProductChangedFieldsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(services.get_changed_field_counts())


class KeywordProductExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None

        rows = services.get_products_for_export(
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )

        STATUS_LABELS = {1: '판매중', 2: '품절', 3: '단종'}

        wb = Workbook()
        ws = wb.active
        ws.title = '오너클랜 상품대장'

        headers = ['W코드', '상태', '동기화', '상품명', '원본상품명',
                    '마켓상품명', '원본마켓상품명', '오너클랜가', '원본오너클랜가',
                    '마켓가', '원본마켓가', '배송비', '원본배송비',
                    '반품비', '원본반품비', '카테고리', '제조사', '원산지']
        col_widths = [12, 8, 8, 35, 35, 35, 35, 12, 12, 12, 12, 8, 8, 8, 8, 20, 15, 10]

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='F0F0F0')
        changed_fill = PatternFill('solid', fgColor='FFF3E0')
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
        money_fmt = '#,##0'

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        for i, r in enumerate(rows, 2):
            is_changed = r.get('is_synced') == 0
            fill = changed_fill if is_changed else None

            def _cell(col, val, fmt=None):
                c = ws.cell(row=i, column=col, value=val)
                c.border = thin_border
                if fmt:
                    c.number_format = fmt
                if fill:
                    c.fill = fill
                return c

            _cell(1, r.get('product_code'))
            _cell(2, STATUS_LABELS.get(r.get('sale_status'), '?'))
            _cell(3, '변경됨' if is_changed else '일치')
            _cell(4, r.get('product_name'))
            _cell(5, r.get('orig_product_name'))
            _cell(6, r.get('market_product_name'))
            _cell(7, r.get('orig_market_product_name'))
            _cell(8, r.get('ownerclan_price', 0), money_fmt)
            _cell(9, r.get('orig_ownerclan_price', 0), money_fmt)
            _cell(10, r.get('market_price', 0), money_fmt)
            _cell(11, r.get('orig_market_price', 0), money_fmt)
            _cell(12, r.get('shipping_fee', 0), money_fmt)
            _cell(13, r.get('orig_shipping_fee', 0), money_fmt)
            _cell(14, r.get('return_fee', 0), money_fmt)
            _cell(15, r.get('orig_return_fee', 0), money_fmt)
            _cell(16, r.get('category_name'))
            _cell(17, r.get('manufacturer'))
            _cell(18, r.get('origin'))

        ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ownerclan_products.xlsx"'
        return response


class KeywordProductWCodesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None

        codes = services.get_w_codes(
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )
        return Response({'codes': codes, 'count': len(codes)})


class KeywordProductDeleteAllView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        confirm = request.data.get('confirm')
        if confirm != 'DELETE_ALL':
            return Response({'error': "확인 토큰 누락 (confirm='DELETE_ALL' 필요)"}, status=400)
        result = services.delete_all_products()
        return Response(result)


class KeywordProductDeleteByIdsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'error': '삭제할 id 리스트 필요'}, status=400)
        result = services.delete_products_by_ids(ids)
        return Response(result)


class KeywordProductDedupeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        result = services.dedupe_by_product_name()
        return Response(result)


class KeywordDistinctValuesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        column = request.query_params.get('column')
        if not column:
            return Response({'error': 'column 파라미터 필요'}, status=400)
        try:
            values = services.get_distinct_values(column)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        return Response({'column': column, 'values': values})
