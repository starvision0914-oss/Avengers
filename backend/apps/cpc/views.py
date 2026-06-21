from rest_framework import viewsets, views, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import models
from django.db.models import Sum, Count, Max
from django.db.models.functions import TruncDate
from .models import CPCDailyCost, CPCDeposit, CPCTransaction
from .serializers import CPCDailyCostSerializer, CPCDepositSerializer, CPCTransactionSerializer


class CPCDailyCostViewSet(viewsets.ModelViewSet):
    queryset = CPCDailyCost.objects.select_related('seller').all()
    serializer_class = CPCDailyCostSerializer
    filterset_fields = ['seller', 'date']
    search_fields = ['seller__seller_id', 'seller__seller_name']
    ordering_fields = ['date', 'total_cost']


class CPCDepositViewSet(viewsets.ModelViewSet):
    queryset = CPCDeposit.objects.select_related('seller').all()
    serializer_class = CPCDepositSerializer
    filterset_fields = ['seller', 'deposit_date']


class CPCTransactionViewSet(viewsets.ModelViewSet):
    queryset = CPCTransaction.objects.select_related('seller').all()
    serializer_class = CPCTransactionSerializer
    filterset_fields = ['seller', 'category']
    ordering_fields = ['transaction_time', 'amount']


class CPCSummaryView(views.APIView):
    def get(self, request):
        date = request.query_params.get('date')
        qs = CPCDailyCost.objects.select_related('seller')
        if date:
            qs = qs.filter(date=date)
        summary = qs.values('seller__seller_id', 'seller__seller_name').annotate(
            total_cpc=Sum('cpc_cost'),
            total_ai=Sum('ai_cost'),
            total_cost=Sum('total_cost'),
            total_clicks=Sum('clicks'),
        ).order_by('seller__seller_id')

        deposits = {}
        for d in CPCDeposit.objects.values('seller__seller_id').annotate(latest_balance=Sum('balance')):
            deposits[d['seller__seller_id']] = d['latest_balance']

        result = []
        for s in summary:
            sid = s['seller__seller_id']
            s['balance'] = deposits.get(sid, 0)
            result.append(s)
        return Response(result)


class CPCChartView(views.APIView):
    def get(self, request):
        date = request.query_params.get('date')
        seller_id = request.query_params.get('seller')
        qs = CPCDailyCost.objects.all()
        if date:
            qs = qs.filter(date=date)
        if seller_id:
            qs = qs.filter(seller_id=seller_id)
        data = qs.values('date').annotate(
            total_cpc=Sum('cpc_cost'),
            total_ai=Sum('ai_cost'),
            total_cost=Sum('total_cost'),
        ).order_by('date')
        return Response(list(data))


from .models import CrawlerAccount, CrawlerLog, GmarketDepositSnapshot, ElevenCostHistory, GmarketSellerGrade, ElevenSellerGrade
from .serializers import CrawlerAccountSerializer, CrawlerLogSerializer, GmarketDepositSnapshotSerializer, ElevenCostHistorySerializer, GmarketGradeSerializer, ElevenGradeSerializer
import threading

class CrawlerAccountViewSet(viewsets.ModelViewSet):
    queryset = CrawlerAccount.objects.all()
    serializer_class = CrawlerAccountSerializer
    filterset_fields = ['platform', 'is_active', 'crawling_status', 'is_focused']
    pagination_class = None

    @action(detail=False, methods=['post'], url_path='bulk-focus')
    def bulk_focus(self, request):
        ids = request.data.get('ids') or []
        focused = bool(request.data.get('focused'))
        if not isinstance(ids, list) or not ids:
            return Response({'error': 'ids 리스트 필요'}, status=400)
        updated = CrawlerAccount.objects.filter(id__in=[int(i) for i in ids]).update(is_focused=focused)
        return Response({'updated': updated, 'focused': focused})

class CrawlerLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CrawlerLogSerializer
    filterset_fields = ['platform', 'level', 'account_id']

    def get_queryset(self):
        return CrawlerLog.objects.all()[:200]

class GmarketSnapshotViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = GmarketDepositSnapshotSerializer
    filterset_fields = ['gmarket_id']
    pagination_class = None

    def get_queryset(self):
        qs = GmarketDepositSnapshot.objects.all()
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(collected_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(collected_at__date__lte=date_to)
        # 페이지네이션 없음 → 전량 직렬화 방지(1.6만행). 최신순 LIMIT(기본 200, ?limit= 조정)
        try:
            limit = min(int(self.request.query_params.get('limit', 200)), 2000)
        except (TypeError, ValueError):
            limit = 200
        return qs.order_by('-collected_at')[:limit]

class ElevenCostViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ElevenCostHistorySerializer
    filterset_fields = ['seller_id', 'transaction_type']
    pagination_class = None

    def get_queryset(self):
        qs = ElevenCostHistory.objects.all()
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(transaction_datetime__date__gte=date_from)
        if date_to:
            qs = qs.filter(transaction_datetime__date__lte=date_to)
        # 페이지네이션 없음 → 전량 직렬화 방지(20.8만행, ~30초). 최신순 LIMIT(기본 200, ?limit= 조정)
        try:
            limit = min(int(self.request.query_params.get('limit', 200)), 2000)
        except (TypeError, ValueError):
            limit = 200
        return qs.order_by('-transaction_datetime')[:limit]

class GmarketGradeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GmarketSellerGrade.objects.all()
    serializer_class = GmarketGradeSerializer
    filterset_fields = ['gmarket_id']

class ElevenGradeViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ElevenSellerGrade.objects.all()
    serializer_class = ElevenGradeSerializer
    filterset_fields = ['eleven_id']

class ProfitDashboardView(views.APIView):
    """전체 수익구조 대시보드 — 아이디별 AI상태, 매출, 광고비, 이익"""
    def get(self, request):
        from datetime import datetime, timedelta, date as dt_date
        import pytz
        kst = pytz.timezone('Asia/Seoul')
        today = datetime.now(kst).date()   # 서버 로컬 아닌 KST 기준(말일/새벽 어긋남 방지)
        month_start = today.replace(day=1)

        # 1) 계정별 최신 광고비 스냅샷
        latest_ids = GmarketDepositSnapshot.objects.values('gmarket_id').annotate(
            latest_id=Max('id')
        ).values_list('latest_id', flat=True)
        snaps = {s.gmarket_id: s for s in GmarketDepositSnapshot.objects.filter(id__in=latest_ids)}

        # 2) 이번 달 광고비 합계 — 거래내역(GmarketCostHistory) use_date 기준 CPC+AI매출업 합산.
        #    스냅샷 total_usage는 '당일 실시간 누적'이라 자정 리셋 + 크롤 시각 의존(8~19시만)으로
        #    월합산 시 실제의 ~1/8만 잡혀 부정확 → 거래원장으로 집계(멱등·정확).
        #    스냅샷은 today_*(실시간 당일사용액)·잔액 표시용으로만 사용.
        from .models import GmarketCostHistory
        month_cost_map = {}
        for r in (GmarketCostHistory.objects
                  .filter(use_date__gte=month_start, use_date__lte=today,
                          transaction_type__in=['CPC', 'AI매출업'])
                  .exclude(comment__icontains='판매예치금')   # 판매예치금 송금 등 비광고 차감 제외
                  .values('seller_id', 'transaction_type')
                  .annotate(amt=Sum('amount'))):
            gid = r['seller_id']
            m = month_cost_map.setdefault(gid, {'cpc': 0, 'ai': 0, 'total': 0})
            spend = abs(r['amt'] or 0)   # amount는 차감(음수) → 절대값이 광고비
            if r['transaction_type'] == 'CPC':
                m['cpc'] += spend
            else:  # AI매출업
                m['ai'] += spend
            m['total'] += spend

        # 3) AI 상태
        ai_map = {}
        for a in GmarketAiAdSummary.objects.all():
            key = a.gmarket_id
            if key not in ai_map:
                ai_map[key] = a.actual_status
            # 복수 그룹이면 하나라도 ON이면 ON
            if a.actual_status == 'ON':
                ai_map[key] = 'ON'

        # 4) 매출 데이터 (SalesRecord)
        try:
            from apps.sales.models import SalesRecord
            # '지마켓 매출'은 지마켓+옥션만 — platform 필터 없으면 11번가·스마트스토어·쿠팡까지 합산돼 부풀려짐
            month_sales_qs = SalesRecord.objects.filter(
                order_date__gte=month_start, order_date__lte=today, status='completed',
                platform__in=['gmarket', 'auction']
            )
            today_sales_qs = SalesRecord.objects.filter(
                order_date=today, status='completed', platform__in=['gmarket', 'auction'])

            month_sales_total = month_sales_qs.aggregate(
                revenue=Sum('total_price'), cost=Sum('commission'),
                profit=Sum('net_profit'), orders=Count('id')
            )
            today_sales_total = today_sales_qs.aggregate(
                revenue=Sum('total_price'), profit=Sum('net_profit'), orders=Count('id')
            )
        except Exception:
            month_sales_total = {'revenue': 0, 'cost': 0, 'profit': 0, 'orders': 0}
            today_sales_total = {'revenue': 0, 'profit': 0, 'orders': 0}

        # 5) 계정별 행 구성
        from apps.cpc.models import CrawlerAccount
        accounts = CrawlerAccount.objects.filter(platform='gmarket', is_active=True).order_by('display_order')

        sellers = []
        totals = {
            'month_cpc': 0, 'month_ai': 0, 'month_ad': 0,
            'today_cpc': 0, 'today_ai': 0, 'today_ad': 0,
            'balance': 0,
            'month_sales': month_sales_total.get('revenue') or 0,
            'today_sales': today_sales_total.get('revenue') or 0,
            'month_profit': month_sales_total.get('profit') or 0,
            'month_orders': month_sales_total.get('orders') or 0,
        }

        for acct in accounts:
            gid = acct.login_id
            snap = snaps.get(gid)
            mc = month_cost_map.get(gid, {'cpc': 0, 'ai': 0, 'total': 0})

            today_cpc = snap.gmarket_cpc if snap else 0
            today_ai = snap.ai_usage if snap else 0
            today_ad = snap.total_usage if snap else 0
            balance = snap.total_balance if snap else 0

            row = {
                'gmarket_id': gid,
                'seller_name': acct.seller_name or gid,
                'ai_status': ai_map.get(gid, '-'),
                'balance': balance,
                'today_cpc': today_cpc,
                'today_ai': today_ai,
                'today_ad': today_ad,
                'month_cpc': mc['cpc'],
                'month_ai': mc['ai'],
                'month_ad': mc['total'],
                'collected_at': snap.collected_at.isoformat() if snap and snap.collected_at else '',
            }
            sellers.append(row)

            totals['today_cpc'] += today_cpc
            totals['today_ai'] += today_ai
            totals['today_ad'] += today_ad
            totals['month_cpc'] += mc['cpc']
            totals['month_ai'] += mc['ai']
            totals['month_ad'] += mc['total']
            totals['balance'] += balance

        # 순이익 = 매출 - 광고비
        totals['net_profit'] = totals['month_sales'] - totals['month_ad']

        # === 11번가 데이터 === (광고비 수집 대상 = api_key 보유 계정만, /st11과 동일 기준)
        eleven_accounts = CrawlerAccount.objects.filter(
            platform='11st', is_active=True).exclude(api_key='').order_by('display_order')
        eleven_sellers = []
        eleven_totals = {'balance': 0, 'month_cost': 0, 'today_cost': 0,
                         'month_sales': 0, 'month_net_profit': 0,
                         'account_count': len(eleven_accounts)}

        # 11번가 이번 달 매출/상품순익 (정산받는금액, platform=11st, 셀러명 매칭)
        from apps.sales.models import SalesRecord
        e_sales_map = {}
        for r in SalesRecord.objects.filter(
            platform='11st', order_date__gte=month_start, order_date__lte=today
        ).values('seller__seller_name').annotate(rev=Sum('total_price'), prof=Sum('net_profit')):
            nm = r['seller__seller_name']
            if nm:
                e_sales_map[nm] = r

        # 11번가 광고비 (ElevenCostHistory) — __date 조회는 MySQL+USE_TZ에서 0건 반환되므로 KST datetime 범위 사용
        from django.db.models import Min
        _m_start = kst.localize(datetime.combine(month_start, datetime.min.time()))
        _d_start = kst.localize(datetime.combine(today, datetime.min.time()))
        _d_end = _d_start + timedelta(days=1)
        # 계정별 5쿼리 N+1(355쿼리) → 루프 전 일괄 집계(5쿼리)로 변경
        from apps.cpc.models import St11AdofficeCampaign
        e_ids = [a.login_id for a in eleven_accounts]
        _agg = lambda qs: {r['seller_id']: abs(r['t'] or 0) for r in qs}
        month_cost_by = _agg(ElevenCostHistory.objects.filter(
            seller_id__in=e_ids, transaction_type='CPC',
            transaction_datetime__gte=_m_start, transaction_datetime__lt=_d_end,
        ).values('seller_id').annotate(t=Sum('amount')))
        today_cost_by = _agg(ElevenCostHistory.objects.filter(
            seller_id__in=e_ids, transaction_type='CPC',
            transaction_datetime__gte=_d_start, transaction_datetime__lt=_d_end,
        ).values('seller_id').annotate(t=Sum('amount')))
        _bal_ids = list(ElevenCostHistory.objects.filter(seller_id__in=e_ids)
                        .values('seller_id').annotate(mx=Max('id')).values_list('mx', flat=True))
        balance_by = {r.seller_id: r.balance for r in ElevenCostHistory.objects.filter(id__in=_bal_ids)}
        _ai_ids = list(St11AdofficeCampaign.objects.filter(eleven_id__in=e_ids, is_ai=True)
                       .values('eleven_id').annotate(mx=Max('id')).values_list('mx', flat=True))
        ai_onoff_by = {c.eleven_id: c.onoff for c in St11AdofficeCampaign.objects.filter(id__in=_ai_ids)}
        _g_ids = list(ElevenSellerGrade.objects.filter(eleven_id__in=e_ids)
                      .values('eleven_id').annotate(mx=Max('id')).values_list('mx', flat=True))
        grade_by = {g.eleven_id: g for g in ElevenSellerGrade.objects.filter(id__in=_g_ids)}

        for acct in eleven_accounts:
            sid = acct.login_id
            month_cost = month_cost_by.get(sid, 0)
            today_cost = today_cost_by.get(sid, 0)
            balance = balance_by.get(sid, 0)

            # 이번 달 매출/순수익 (정산받는금액 - 구매가 = 상품순익, 순수익 = 상품순익 - 광고비)
            srow = e_sales_map.get(acct.seller_name or sid)
            s_sales = (srow['rev'] or 0) if srow else 0
            s_prof = (srow['prof'] or 0) if srow else 0
            s_net = s_prof - month_cost

            # AI 상태 (St11AdofficeCampaign): is_ai 캠페인 있고 onoff면 ON, 있으나 OFF면 OFF, 없으면 '-'
            ai_status = 'ON' if ai_onoff_by.get(sid) else ('OFF' if sid in ai_onoff_by else '-')
            grade_row = grade_by.get(sid)

            row = {
                'seller_id': sid,
                'seller_name': acct.seller_name or sid,
                'ai_status': ai_status,
                'balance': balance,
                'today_cost': today_cost,
                'month_cost': month_cost,
                'grade': grade_row.grade if grade_row else None,
                'grade_message': grade_row.grade_message if grade_row else '',
                'collected_at': acct.last_crawled_at.isoformat() if acct.last_crawled_at else '',
                'cost_type': acct.cost_type or 'sellerpoint',
                'sales': s_sales,
                'net_profit': s_net,
            }
            eleven_sellers.append(row)

            eleven_totals['balance'] += balance
            eleven_totals['month_cost'] += month_cost
            eleven_totals['today_cost'] += today_cost
            eleven_totals['month_sales'] += s_sales
            eleven_totals['month_net_profit'] += s_net

        totals['eleven_balance'] = eleven_totals['balance']
        totals['eleven_month_cost'] = eleven_totals['month_cost']
        totals['eleven_today_cost'] = eleven_totals['today_cost']
        totals['eleven_month_sales'] = eleven_totals['month_sales']
        totals['eleven_month_net_profit'] = eleven_totals['month_net_profit']
        totals['eleven_count'] = eleven_totals['account_count']

        # 전체 광고비 합계 (지마켓 + 11번가) — 참고용
        totals['total_ad_all'] = totals['month_ad'] + eleven_totals['month_cost']
        # 지마켓 순이익 = 지마켓 상품순익(정산-수수료-원가) - 지마켓 광고비 (11번가와 동일 기준)
        totals['net_profit'] = totals['month_profit'] - totals['month_ad']

        return Response({
            'date': today.isoformat(),
            'month': f'{today.year}-{today.month:02d}',
            'totals': totals,
            'sellers': sellers,
            'eleven_sellers': eleven_sellers,
            'eleven_totals': eleven_totals,
        })


class GmarketSummaryView(views.APIView):
    """지마켓 광고비 요약 - ai100 /api/cpc/summary/ 호환"""
    def get(self, request):
        from datetime import datetime, timedelta
        import pytz
        kst = pytz.timezone('Asia/Seoul')

        date_str = request.query_params.get('date')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        # Also accept start_date/end_date (ai100 compatibility)
        if not date_from:
            date_from = request.query_params.get('start_date')
        if not date_to:
            date_to = request.query_params.get('end_date')

        if date_from and date_to:
            start = kst.localize(datetime.strptime(date_from, '%Y-%m-%d'))
            end = kst.localize(datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        elif date_str:
            start = kst.localize(datetime.strptime(date_str, '%Y-%m-%d'))
            end = start + timedelta(days=1)
        else:
            from django.utils import timezone as tz
            today = tz.localdate()
            date_str = today.isoformat()
            start = kst.localize(datetime.combine(today, datetime.min.time()))
            end = start + timedelta(days=1)

        latest_ids = GmarketDepositSnapshot.objects.filter(
            collected_at__gte=start, collected_at__lt=end
        ).values('gmarket_id').annotate(
            latest_id=Max('id')
        ).values_list('latest_id', flat=True)

        sellers = []
        total_cpc = 0
        total_ai = 0
        total_usage = 0
        total_balance = 0

        # 광고상태, AI, 등급 미리 조회
        cpc_status_map = {}
        for s in GmarketCpcAdStatus.objects.all():
            cpc_status_map[s.gmarket_id] = {
                'cpc1_on': s.cpc1_on, 'cpc1_off': s.cpc1_off,
                'cpc2_on': s.cpc2_on, 'cpc2_off': s.cpc2_off,
            }
        ai_map = {}
        for a in GmarketAiAdSummary.objects.all():
            ai_map[a.gmarket_id] = {
                'actual_status': a.actual_status,
                'actual_reason': a.actual_reason,
                'button_status': a.button_status,
                'start_date': a.start_date,
            }
        grade_map = {}
        for g in GmarketSellerGrade.objects.order_by('-collected_at'):
            if g.gmarket_id not in grade_map:
                grade_map[g.gmarket_id] = {
                    'seller_grade': g.seller_grade,
                    'max_item_count': g.max_item_count,
                    'approval_status': g.approval_status,
                    'contact_expiry': g.contact_expiry,
                }

        for snap in GmarketDepositSnapshot.objects.filter(id__in=latest_ids).order_by('gmarket_id'):
            seller = {
                'seller_id': snap.gmarket_id,
                'seller_alias': snap.gmarket_id,
                'balance': snap.total_balance,
                'cpc_spend': snap.gmarket_cpc,
                'auction_cpc': snap.auction_cpc,
                'ai_spend': snap.ai_usage,
                'ad_total': snap.total_usage,
                'collected_at': snap.collected_at.isoformat() if snap.collected_at else '',
                'cpc_status': cpc_status_map.get(snap.gmarket_id),
                'ai_status': ai_map.get(snap.gmarket_id),
                'grade_info': grade_map.get(snap.gmarket_id),
            }
            sellers.append(seller)
            total_cpc += snap.gmarket_cpc
            total_ai += snap.ai_usage
            total_usage += snap.total_usage
            total_balance += snap.total_balance

        # Sales data integration
        from apps.cpc.models import CrawlerAccount
        acct_map = {a.login_id: a for a in CrawlerAccount.objects.filter(platform='gmarket', is_active=True)}

        total_sales = 0
        total_sales_count = 0
        total_profit = 0

        try:
            from apps.sales.models import SalesRecord
            sales_qs = SalesRecord.objects.filter(
                order_date__gte=start.date() if hasattr(start, 'date') else start,
                order_date__lte=(end - timedelta(days=1)).date() if hasattr(end, 'date') else end,
            )

            for seller in sellers:
                acct = acct_map.get(seller['seller_id'])
                if acct and acct.seller_name:
                    seller['seller_alias'] = acct.seller_name
                    s_sales = sales_qs.filter(seller_name=acct.seller_name).aggregate(
                        total=Sum('settlement_price'),
                        count=Count('id'),
                    )
                    seller['sales'] = s_sales['total'] or 0
                    seller['sales_count'] = s_sales['count'] or 0
                    seller['profit'] = seller['sales'] - seller['ad_total']
                else:
                    seller['sales'] = 0
                    seller['sales_count'] = 0
                    seller['profit'] = -seller['ad_total']

                seller['monthly_sales'] = 0
                seller['server_fee_date'] = acct.server_fee_date if acct and hasattr(acct, 'server_fee_date') else None
                seller['prime_spend'] = 0
                seller['cost'] = 0
                seller['cost_count'] = 0
                seller['last_tx'] = seller.get('collected_at')

                total_sales += seller['sales']
                total_sales_count += seller['sales_count']
                total_profit += seller['profit']
        except Exception:
            for seller in sellers:
                acct = acct_map.get(seller['seller_id'])
                if acct and acct.seller_name:
                    seller['seller_alias'] = acct.seller_name
                seller['sales'] = 0
                seller['sales_count'] = 0
                seller['profit'] = -seller['ad_total']
                seller['monthly_sales'] = 0
                seller['prime_spend'] = 0
                seller['cost'] = 0
                seller['cost_count'] = 0
                seller['last_tx'] = seller.get('collected_at')

        return Response({
            'date': date_str or (date_from + '~' + date_to if date_from else ''),
            'totals': {
                'cpc_spend': total_cpc,
                'ai_spend': total_ai,
                'ad_total': total_usage,
                'balance': total_balance,
                'sales': total_sales,
                'sales_count': total_sales_count,
                'profit': total_profit,
                'net_profit': total_sales - total_usage,
                'prime_spend': 0,
                'cost': 0,
                'cost_count': 0,
            },
            'sellers': sellers,
        })

class ElevenSummaryView(views.APIView):
    """11번가 광고비 요약 — 전체 활성 계정 + 최신 잔액 + 기간 CPC + 오피스 현황"""
    def get(self, request):
        from datetime import datetime, timedelta
        import pytz
        kst = pytz.timezone('Asia/Seoul')
        from .models import CrawlerAccount, ElevenSellerOfficeStat

        date_str = request.query_params.get('date')
        date_from = request.query_params.get('date_from') or request.query_params.get('start_date')
        date_to = request.query_params.get('date_to') or request.query_params.get('end_date')

        if date_from and date_to:
            start = kst.localize(datetime.strptime(date_from, '%Y-%m-%d'))
            end = kst.localize(datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        elif date_str:
            start = kst.localize(datetime.strptime(date_str, '%Y-%m-%d'))
            end = start + timedelta(days=1)
        else:
            from django.utils import timezone as tz
            today = tz.localdate()
            date_str = today.isoformat()
            start = kst.localize(datetime.combine(today, datetime.min.time()))
            end = start + timedelta(days=1)

        # 기간 내 거래 집계
        qs = ElevenCostHistory.objects.filter(
            transaction_datetime__gte=start, transaction_datetime__lt=end
        )
        stats_map = {}
        for s in qs.values('seller_id').annotate(
            cpc_total=Sum('amount', filter=models.Q(transaction_type='CPC')),
            charge_total=Sum('amount', filter=models.Q(transaction_type='CHARGE')),
            settle_total=Sum('amount', filter=models.Q(transaction_type='SETTLE')),
            server_fee_total=Sum('amount', filter=models.Q(transaction_type='OTHERS', raw_description__icontains='서버이용료')),
            reward_total=Sum('amount', filter=models.Q(transaction_type='REWARD')),
            spend_total=Sum('amount', filter=models.Q(amount__lt=0)),
            total_count=Count('id'),
        ):
            stats_map[s['seller_id']] = s

        # 매출/구매가/상품순익 (SalesRecord, 같은 기간, 셀러명 기준) — 순수익 계산용
        # 11번가 대시보드이므로 platform='11st' 매출만 합산 (지마켓/스마트스토어 등 제외)
        from apps.sales.models import SalesRecord
        sales_map = {}
        for r in SalesRecord.objects.filter(
            platform='11st',
            order_date__gte=start.date(), order_date__lt=end.date()
        ).values('seller__seller_name').annotate(
            rev=Sum('total_price'), prof=Sum('net_profit'), cnt=Count('id')
        ):
            nm = r['seller__seller_name']
            if nm:
                sales_map[nm] = r

        # 이전 크롤링 시점 CPC (증감 계산용) — CrawlerLog에서 직전 성공 시각 조회
        from .models import CrawlerLog
        prev_crawl_map = {}
        # .values로 message(TextField) 제외해 경량화 (인덱스 platform,level,-created_at 사용)
        for log in (CrawlerLog.objects.filter(platform='11st', level='success')
                    .order_by('-created_at').values('account_id', 'created_at')):
            sid = log['account_id']
            if sid not in prev_crawl_map:
                prev_crawl_map[sid] = []
            if len(prev_crawl_map[sid]) < 2:
                prev_crawl_map[sid].append(log['created_at'])

        # 표시 대상: 활성 11번가 계정 전체 (api 없는 대기계정도 로스터에 표시)
        accounts = list(
            CrawlerAccount.objects.filter(
                platform='11st', is_active=True
            ).order_by('display_order', 'login_id')
        )

        # 등급 정보
        grade_map = {}
        for g in ElevenSellerGrade.objects.order_by('-collected_at'):
            if g.eleven_id not in grade_map:
                grade_map[g.eleven_id] = {
                    'grade': g.grade,
                    'grade_message': g.grade_message,
                    'collected_at': g.collected_at.isoformat() if g.collected_at else None,
                }

        # 셀러오피스 현황 (계정별 최신 성공 데이터 — error 비어있는 것만)
        office_latest_ids = ElevenSellerOfficeStat.objects.filter(error='').values('account_id').annotate(
            latest=Max('id')
        ).values_list('latest', flat=True)
        office_map = {}
        for ofs in ElevenSellerOfficeStat.objects.filter(id__in=office_latest_ids).select_related('account'):
            office_map[ofs.account.login_id] = ofs

        # 계정별 N+1 제거: 최신잔액 + CPC증감을 루프 전 일괄 집계
        _acct_ids = [a.login_id for a in accounts]
        _bal_ids = list(ElevenCostHistory.objects.filter(seller_id__in=_acct_ids)
                        .values('seller_id').annotate(mx=Max('id')).values_list('mx', flat=True))
        balance_by = {r.seller_id: r.balance for r in ElevenCostHistory.objects.filter(id__in=_bal_ids)}
        # CPC 증감(delta) = 직전 성공크롤 시각 이후의 CPC 합 — seller별 경계가 달라 최근창 1쿼리로 묶음
        prev_bounds = {sid: t[1] for sid, t in prev_crawl_map.items() if len(t) >= 2}
        cpc_delta_by = {}
        if prev_bounds:
            _earliest = min(prev_bounds.values())
            for r in (ElevenCostHistory.objects.filter(
                    seller_id__in=list(prev_bounds.keys()), transaction_type='CPC',
                    transaction_datetime__gte=_earliest, transaction_datetime__lt=end)
                    .values('seller_id', 'transaction_datetime', 'amount')):
                b = prev_bounds.get(r['seller_id'])
                if b and r['transaction_datetime'] >= b:
                    cpc_delta_by[r['seller_id']] = cpc_delta_by.get(r['seller_id'], 0) + abs(r['amount'] or 0)

        sellers = []
        total_cpc = 0
        total_charge = 0
        total_balance = 0
        total_cash = 0
        total_point = 0
        total_products = 0
        total_limit = 0
        total_available = 0
        total_sales = 0
        total_cost = 0
        total_server_fee = 0
        total_reward = 0
        total_net_profit = 0
        last_collected_at = None

        for acct in accounts:
            sid = acct.login_id
            stat = stats_map.get(sid, {})
            cpc = abs(stat.get('cpc_total') or 0)
            charge = abs(stat.get('charge_total') or 0)
            settle = stat.get('settle_total') or 0
            server_fee = abs(stat.get('server_fee_total') or 0)   # 서버이용료(실비용)
            reward = stat.get('reward_total') or 0                # 프로모션 리워드(+)
            spend = abs(stat.get('spend_total') or 0)
            tx_count = stat.get('total_count') or 0
            is_cash = (acct.cost_type == 'sellercash')

            # 최신 잔액 / CPC 증감 — 루프 전 일괄 집계 결과 사용 (N+1 제거)
            balance = balance_by.get(sid, 0)
            cpc_delta = cpc_delta_by.get(sid, 0)

            seller = {
                'seller_id': sid,
                'seller_alias': acct.seller_name or sid,
                'cpc_spend': cpc,
                'cpc_delta': cpc_delta,
                'charge': charge if not is_cash else 0,
                'settle': settle if is_cash else 0,
                'ad_spend': spend,
                'balance': balance,
                'tx_count': tx_count,
                'ad_total': spend if is_cash else cpc,
                'cost_type': acct.cost_type or 'sellerpoint',
                'crawling_status': acct.crawling_status,
                'last_crawled_at': acct.last_crawled_at.isoformat() if acct.last_crawled_at else None,
                'fail_count': acct.fail_count,
                'no_api': not bool(acct.api_key),   # api 없는 대기/정지 계정 (대시보드 주황 표시용)
                'last_otp_at': acct.last_otp_at.isoformat() if acct.last_otp_at else None,
                # 실제 인증/로그인 신선도 — 쿠키워밍·크롤 로그인 성공 시 갱신(쿠키 롤포워드 포함).
                # OTP는 11번가가 요구할 때만 떠서 last_otp_at은 오래돼도 정상 → 인증상태는 이 값으로 판단.
                'cookie_saved_at': acct.cookie_saved_at.isoformat() if acct.cookie_saved_at else None,
            }

            # 매출/상품순익/순수익 (매출데이터 기준, 셀러명 매칭)
            # total_price=정산받는금액(매출), net_profit=정산받는금액-구매가(상품순익), 순수익=상품순익-광고비
            srow = sales_map.get(acct.seller_name or sid)
            sales_rev = (srow['rev'] or 0) if srow else 0
            prod_profit = (srow['prof'] or 0) if srow else 0
            buy_cost = sales_rev - prod_profit   # 구매가 = 매출 - 상품순익
            # 순수익 = 상품순익 - 광고비 - 서버이용료 (프로모션은 순수익 계산 제외, 표시만)
            net = prod_profit - cpc - server_fee
            seller['sales'] = sales_rev
            seller['cost'] = buy_cost
            seller['prod_profit'] = prod_profit
            seller['server_fee'] = server_fee
            seller['reward'] = reward
            seller['net_profit'] = net
            seller['sales_count'] = (srow['cnt'] or 0) if srow else 0
            total_sales += sales_rev
            total_cost += buy_cost
            total_server_fee += server_fee
            total_reward += reward
            total_net_profit += net

            # 등급
            gi = grade_map.get(sid)
            if gi:
                seller['grade'] = gi['grade']
                seller['grade_message'] = gi['grade_message']
                seller['grade_collected_at'] = gi['collected_at']

            # 오피스 현황
            ofs = office_map.get(sid)
            if ofs:
                seller['cash'] = ofs.cash
                seller['point'] = ofs.point
                seller['ad_balance'] = ofs.ad_balance
                seller['products'] = ofs.products
                seller['product_limit'] = ofs.product_limit
                seller['available'] = ofs.available
                seller['overdue'] = ofs.overdue
                seller['undelivered'] = ofs.undelivered
                seller['draft'] = ofs.draft
                seller['fulfillment'] = ofs.fulfillment
                seller['shipping'] = ofs.shipping
                seller['inquiry'] = ofs.inquiry
                seller['office_collected_at'] = ofs.collected_at.isoformat() if ofs.collected_at else None
                total_cash += ofs.cash
                total_point += ofs.point
                total_products += ofs.products
                total_limit += ofs.product_limit
                total_available += ofs.available
                if ofs.collected_at and (not last_collected_at or ofs.collected_at > last_collected_at):
                    last_collected_at = ofs.collected_at

            sellers.append(seller)
            total_cpc += cpc
            total_charge += charge
            total_balance += balance

            # last_collected_at from cost crawl
            if acct.last_crawled_at and (not last_collected_at or acct.last_crawled_at > last_collected_at):
                last_collected_at = acct.last_crawled_at

        # 매출은 있으나 대시보드(크롤계정 셀러명)에 매칭 안 된 쇼핑몰 — 누락 확인용
        matched_names = set()
        for acct in accounts:
            matched_names.add(acct.seller_name or acct.login_id)
        unmatched_shops = []
        unmatched_sales = 0
        unmatched_cost = 0
        unmatched_prof = 0
        unmatched_count = 0
        for nm, srow in sales_map.items():
            if nm and nm not in matched_names:
                rev = srow['rev'] or 0
                prof = srow['prof'] or 0
                cnt = srow['cnt'] or 0
                unmatched_sales += rev
                unmatched_prof += prof
                unmatched_cost += (rev - prof)
                unmatched_count += cnt
                unmatched_shops.append({'name': nm, 'sales': rev, 'cost': rev - prof,
                                        'net_profit': prof, 'count': cnt})
        unmatched_shops.sort(key=lambda x: -x['sales'])

        # 기타(미매칭)도 총합에 포함 — 매출/구매가/순수익(상품순익, 광고비 없음)
        total_sales += unmatched_sales
        total_cost += unmatched_cost
        total_net_profit += unmatched_prof

        return Response({
            'date': date_str or (date_from + '~' + date_to if date_from else ''),
            'totals': {
                'cpc_spend': total_cpc,
                'charge': total_charge,
                'balance': total_balance,
                'ad_total': total_cpc,
                'seller_count': len(sellers),
                'cash': total_cash,
                'point': total_point,
                'products': total_products,
                'product_limit': total_limit,
                'available': total_available,
                'sales': total_sales,
                'cost': total_cost,
                'server_fee': total_server_fee,
                'reward': total_reward,
                'net_profit': total_net_profit,
            },
            'sellers': sellers,
            'unmatched': {
                'sales': unmatched_sales,
                'cost': unmatched_cost,
                'net_profit': unmatched_prof,
                'count': unmatched_count,
                'shops': unmatched_shops,
            },
            'last_collected_at': last_collected_at.isoformat() if last_collected_at else None,
        })


class OverviewView(views.APIView):
    """통합 Overview — G마켓+11번가 광고비/잔액/계정현황 합산 + 경보.
    기간(period/date_from/date_to) 지원. 지마켓은 거래내역 기반 GmarketDashboardView를
    재사용해 /gmarket 대시보드와 합계 일치(스냅샷 마지막값만 잡던 기간합산 버그 제거)."""
    def get(self, request):
        from datetime import datetime, timedelta
        from django.http import QueryDict
        from django.utils import timezone as tz
        from .models import CrawlerAccount

        # ── 기간 해석: period 프리셋 > date_from/to > date > 기본(어제) ──
        today = tz.localdate()
        yesterday = today - timedelta(days=1)
        period = (request.query_params.get('period') or '').strip()
        df = request.query_params.get('date_from') or request.query_params.get('start_date')
        dt = request.query_params.get('date_to') or request.query_params.get('end_date')
        dsingle = request.query_params.get('date')

        def _pd(s):
            return datetime.strptime(s, '%Y-%m-%d').date()

        if df and dt:
            start_d, end_d = _pd(df), _pd(dt)
        elif dsingle:
            start_d = end_d = _pd(dsingle)
        elif period == 'today':
            start_d = end_d = today
        elif period in ('month', '30d'):
            end_d, start_d = today, today - timedelta(days=29)
        elif period == '7d':
            end_d, start_d = today, today - timedelta(days=6)
        elif period in ('mtd', 'thismonth'):
            end_d, start_d = today, today.replace(day=1)   # 이번 달 1일~오늘
        else:                                   # 기본 = 어제(완성된 최신일)
            start_d = end_d = yesterday
        if end_d < start_d:
            start_d, end_d = end_d, start_d

        # 하위 뷰가 동일 기간을 쓰도록 date_from/date_to 주입
        q = QueryDict(mutable=True)
        for k, v in request.query_params.items():
            q[k] = v
        q['date_from'] = start_d.isoformat()
        q['date_to'] = end_d.isoformat()
        q.pop('date', None)
        q.pop('period', None)
        request._request.GET = q

        # 지마켓 = 거래내역 기반(기간합산 정확) / 11번가 = 기간 거래 집계
        g = GmarketDashboardView().get(request).data
        e = ElevenSummaryView().get(request).data
        gt = g.get('totals', {}) or {}
        et = e.get('totals', {}) or {}

        def acct_stats(platform):
            qs = CrawlerAccount.objects.filter(platform=platform, is_active=True)
            total = qs.count()
            failed = qs.filter(crawling_status__in=['실패', '차단됨']).count()
            return total, total - failed, failed

        g_total, g_normal, g_failed = acct_stats('gmarket')
        e_total, e_normal, e_failed = acct_stats('11st')

        g_ad = gt.get('ad_spend', 0) or 0                                  # G마켓 = CPC+AI+서버(+옥션)
        e_ad = et.get('cpc_spend', 0) or 0                                 # 11번가 = CPC
        g_bal = gt.get('balance', 0) or 0                                  # G마켓 예치금
        e_bal = et.get('point', 0) or 0                                    # 11번가 = 셀러포인트만 (캐시는 내 돈 아님 → 제외)
        # 순익 의미 통일: profit=상품순익(광고 전), net_after_ad=순수익(광고 후)
        g_sales = gt.get('revenue', 0) or 0                                # G마켓 매출(정산받는금액)
        g_profit = gt.get('profit', 0) or 0                                # G마켓 상품순익(매출-원가)
        g_net = gt.get('net_after_ad', 0) or 0                             # G마켓 순수익(상품순익-광고비)
        e_sales = et.get('sales', 0) or 0                                  # 11번가 매출
        e_net = et.get('net_profit', 0) or 0                               # 11번가 순수익(이미 광고+서버 차감)
        # 11번가 상품순익(광고 전) = 순수익 + 광고비 + 서버이용료 → 지마켓과 동일 기준으로 환산
        e_profit = e_net + (et.get('cpc_spend', 0) or 0) + (et.get('server_fee', 0) or 0)

        markets = [
            {'key': 'gmarket', 'label': 'G마켓', 'color': '#6cc24a',
             'ad_cost': g_ad, 'cpc': gt.get('cpc_spend', 0) or 0, 'ai': gt.get('ai_spend', 0) or 0,
             'balance': g_bal, 'accounts': g_total, 'normal': g_normal, 'failed': g_failed,
             'sales': g_sales, 'profit': g_profit, 'net_after_ad': g_net,
             'orders': gt.get('orders', 0) or 0,
             'last_collected': None},
            {'key': '11st', 'label': '11번가', 'color': '#ff5a2e',
             'ad_cost': e_ad, 'cpc': e_ad, 'ai': 0,
             'balance': e_bal, 'cash': et.get('cash', 0) or 0, 'point': et.get('point', 0) or 0,
             'accounts': e_total, 'normal': e_normal, 'failed': e_failed,
             'sales': e_sales, 'profit': e_profit, 'net_after_ad': e_net,
             'last_collected': e.get('last_collected_at')},
        ]

        # 경보: 실패/차단 계정 (전 플랫폼)
        failed_accounts = [
            {'platform': a.platform, 'login_id': a.login_id,
             'seller_name': a.seller_name, 'status': a.crawling_status}
            for a in CrawlerAccount.objects.filter(
                is_active=True, crawling_status__in=['실패', '차단됨']
            ).order_by('platform', 'login_id')
        ]
        # 11번가 잔액부족 (캐시+포인트 < 5만)
        low_balance, zero_ad = [], []
        for s in (e.get('sellers', []) or []):
            tot = (s.get('cash', 0) or 0) + (s.get('point', 0) or 0)
            if 0 < tot < 50000:
                low_balance.append({'platform': '11st', 'seller': s.get('seller_alias'), 'balance': tot})
            if (s.get('cpc_spend', 0) or 0) == 0:
                zero_ad.append({'platform': '11st', 'seller': s.get('seller_alias')})

        totals = {
            'ad_cost': g_ad + e_ad,
            'balance': g_bal + e_bal,
            'accounts': g_total + e_total,
            'normal': g_normal + e_normal,
            'failed': g_failed + e_failed,
            'sales': g_sales + e_sales,
            'profit': g_profit + e_profit,                 # 상품순익(광고 전, 양 마켓 동일기준)
            'net_after_ad': g_net + e_net,                 # 순수익(광고+서버 차감 후)
        }

        return Response({
            'date_from': start_d.isoformat(),
            'date_to': end_d.isoformat(),
            'date': end_d.isoformat(),
            'totals': totals,
            'markets': markets,
            'alerts': {
                'failed_accounts': failed_accounts,
                'low_balance': low_balance,
                'zero_ad': zero_ad,
            },
            'last_collected': e.get('last_collected_at'),
        })


class AdDetailView(views.APIView):
    """광고비 상세 내역 - CPC/AI 클릭 시 모달용"""
    def get(self, request):
        from datetime import datetime, timedelta
        import pytz
        kst = pytz.timezone('Asia/Seoul')

        seller_id = request.query_params.get('seller_id')
        platform = request.query_params.get('platform', 'gmarket')
        date_str = request.query_params.get('date')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        category = request.query_params.get('category')

        if not seller_id:
            return Response({'error': 'seller_id 필요'}, status=400)

        if date_from and date_to:
            start = kst.localize(datetime.strptime(date_from, '%Y-%m-%d'))
            end = kst.localize(datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        elif date_str:
            start = kst.localize(datetime.strptime(date_str, '%Y-%m-%d'))
            end = start + timedelta(days=1)
        else:
            from django.utils import timezone as tz
            today = tz.localdate()
            start = kst.localize(datetime.combine(today, datetime.min.time()))
            end = start + timedelta(days=1)

        if platform == '11st':
            qs = ElevenCostHistory.objects.filter(
                seller_id=seller_id,
                transaction_datetime__gte=start,
                transaction_datetime__lt=end,
            )
            kind = request.query_params.get('kind')
            if kind == 'cpc':
                qs = qs.filter(transaction_type='CPC')
            elif kind == 'settle':   # 충전/정산/프로모션
                qs = qs.filter(transaction_type__in=['CHARGE', 'SETTLE', 'REWARD'])
            elif kind == 'server_fee':
                qs = qs.filter(transaction_type='OTHERS', raw_description__icontains='서버이용료')
            elif category:
                qs = qs.filter(transaction_type=category)

            rows = []
            for r in qs.order_by('-transaction_datetime'):
                rows.append({
                    'time': r.transaction_datetime.astimezone(kst).strftime('%m/%d %H:%M'),
                    'category': r.transaction_type,
                    'description': r.raw_description,
                    'amount': r.amount,
                })
        else:
            # 지마켓 - 스냅샷 이력
            qs = GmarketDepositSnapshot.objects.filter(
                gmarket_id=seller_id,
                collected_at__gte=start,
                collected_at__lt=end,
            )
            rows = []
            for s in qs.order_by('-collected_at'):
                if category == 'CPC':
                    rows.append({'time': s.collected_at.astimezone(kst).strftime('%H:%M:%S'), 'category': 'CPC', 'description': 'CPC 광고비', 'amount': -s.gmarket_cpc})
                elif category == 'AI':
                    rows.append({'time': s.collected_at.astimezone(kst).strftime('%H:%M:%S'), 'category': 'AI', 'description': 'AI 광고비', 'amount': -s.ai_usage})
                else:
                    if s.gmarket_cpc:
                        rows.append({'time': s.collected_at.astimezone(kst).strftime('%H:%M:%S'), 'category': 'CPC', 'description': 'CPC 광고비', 'amount': -s.gmarket_cpc})
                    if s.ai_usage:
                        rows.append({'time': s.collected_at.astimezone(kst).strftime('%H:%M:%S'), 'category': 'AI', 'description': 'AI 광고비', 'amount': -s.ai_usage})

        # 요약
        summary = {}
        for r in rows:
            cat = r['category']
            if cat not in summary:
                summary[cat] = {'count': 0, 'total': 0}
            summary[cat]['count'] += 1
            summary[cat]['total'] += r['amount']

        return Response({'rows': rows, 'summary': summary})

from .models import GmarketAiAdSummary, GmarketAiAdHistory, St11AdofficeCampaign
from .serializers import GmarketAiSummarySerializer, GmarketAiHistorySerializer, St11CampaignSerializer

from .models import GmarketCpcAdStatus, Cpc2Schedule, Cpc2History, AiSchedule, TelegramConfig, TelegramRecipient, SellerGroup
from .serializers import CpcAdStatusSerializer, Cpc2ScheduleSerializer, Cpc2HistorySerializer, AiScheduleSerializer, TelegramConfigSerializer, TelegramRecipientSerializer, SellerGroupSerializer

class CpcAdStatusViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GmarketCpcAdStatus.objects.all()
    serializer_class = CpcAdStatusSerializer

def _sched_dow_csv(weekdays):
    """모델 요일(1=월..7=일) → crontab 요일(0=일,1=월..6=토). 빈값=매일('*')."""
    if not weekdays:
        return '*'
    cron = sorted({(0 if int(w) == 7 else int(w)) for w in weekdays})
    return ','.join(str(x) for x in cron)


def _write_schedule_cron(tag, items):
    """tag가 붙은 기존 크론 라인을 모두 지우고 items로 교체.
    items: [(TimeField, dow_csv, script_path), ...] (time None이면 스킵).
    크론 수정 전 백업(메모리: sed # 사고 방지)."""
    import subprocess, os
    from datetime import datetime
    try:
        res = subprocess.run(['crontab', '-l'], capture_output=True, text=True)
        lines = res.stdout.split('\n') if res.stdout else []
    except Exception:
        lines = []
    # 백업
    try:
        bdir = os.path.expanduser('~/cron_backups'); os.makedirs(bdir, exist_ok=True)
        with open(os.path.join(bdir, f'crontab_{datetime.now():%Y%m%d_%H%M%S}.txt'), 'w') as f:
            f.write('\n'.join(lines))
    except Exception:
        pass
    # 해당 tag 라인 제거(활성 라인만; 주석 #로 시작하는 건 유지)
    lines = [l for l in lines if not (tag in l and not l.lstrip().startswith('#'))]
    for t, dow, script in items:
        if t is None:
            continue
        lines.append(f'{t.minute} {t.hour} * * {dow} {script} # {tag}')
    cron_text = '\n'.join([l for l in lines if l.strip()]) + '\n'
    subprocess.run(['crontab', '-'], input=cron_text, text=True)


def _regenerate_ad_crons():
    """간편(CPC2)+AI 스케줄을 읽어 광고 ON/OFF 크론을 통째로 재생성.
    AI와 간편의 (시각·요일)이 동일하면 → 통합 1회 로그인 크론(cron_ad_combined),
    다르면 → 각자 별도 크론. 한 시각에 둘이 겹쳐 한쪽이 스킵되는 문제·이중 로그인 방지.
    태그 'AD_SCHEDULE'(구 CPC2_AD_SCHEDULE/AI_AD_SCHEDULE도 substring 매칭으로 정리됨)."""
    from apps.cpc.models import Cpc2Schedule, AiSchedule
    base = '/home/rejoice888/Avengers/backend/scripts'
    cpc2 = Cpc2Schedule.objects.first()
    ai = AiSchedule.objects.filter(platform='gmarket').first()
    cpc2_on = bool(cpc2 and cpc2.selected_accounts)
    ai_on = bool(ai and ai.selected_accounts)

    def same(t1, t2, d1, d2):
        return t1 and t2 and t1 == t2 and sorted(d1 or []) == sorted(d2 or [])

    items = []  # (TimeField, dow_csv, script)
    # --- ON ---
    if cpc2_on and ai_on and same(cpc2.on_time, ai.on_time, cpc2.weekdays, ai.weekdays):
        items.append((cpc2.on_time, _sched_dow_csv(cpc2.weekdays), f'{base}/cron_ad_combined_on.sh'))
    else:
        if cpc2_on:
            items.append((cpc2.on_time, _sched_dow_csv(cpc2.weekdays), f'{base}/cron_cpc2_on.sh'))
        if ai_on:
            items.append((ai.on_time, _sched_dow_csv(ai.weekdays), f'{base}/cron_ai_on.sh'))
    # --- OFF ---
    if cpc2_on and ai_on and same(cpc2.off_time, ai.off_time, cpc2.off_weekdays, ai.off_weekdays):
        items.append((cpc2.off_time, _sched_dow_csv(cpc2.off_weekdays), f'{base}/cron_ad_combined_off.sh'))
    else:
        if cpc2_on:
            items.append((cpc2.off_time, _sched_dow_csv(cpc2.off_weekdays), f'{base}/cron_cpc2_off.sh'))
        if ai_on:
            items.append((ai.off_time, _sched_dow_csv(ai.off_weekdays), f'{base}/cron_ai_off.sh'))

    _write_schedule_cron('AD_SCHEDULE', items)


class Cpc2ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Cpc2Schedule.objects.all()
    serializer_class = Cpc2ScheduleSerializer

    def perform_create(self, serializer):
        serializer.save(); _regenerate_ad_crons()

    def perform_update(self, serializer):
        serializer.save(); _regenerate_ad_crons()


class Cpc2HistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Cpc2History.objects.all()[:100]
    serializer_class = Cpc2HistorySerializer

class AiScheduleViewSet(viewsets.ModelViewSet):
    queryset = AiSchedule.objects.all()
    serializer_class = AiScheduleSerializer

    def perform_create(self, serializer):
        serializer.save(); _regenerate_ad_crons()

    def perform_update(self, serializer):
        serializer.save(); _regenerate_ad_crons()

class TelegramConfigViewSet(viewsets.ModelViewSet):
    queryset = TelegramConfig.objects.all()
    serializer_class = TelegramConfigSerializer

class TelegramRecipientViewSet(viewsets.ModelViewSet):
    queryset = TelegramRecipient.objects.all()
    serializer_class = TelegramRecipientSerializer

class SellerGroupViewSet(viewsets.ModelViewSet):
    queryset = SellerGroup.objects.all()
    serializer_class = SellerGroupSerializer

class TelegramSendView(views.APIView):
    def post(self, request):
        import requests as req
        config = TelegramConfig.objects.first()
        if not config or not config.bot_token:
            return Response({'error': '텔레그램 봇 토큰이 설정되지 않았습니다.'}, status=400)

        message = request.data.get('message', '')
        recipients = TelegramRecipient.objects.filter(is_active=True)
        sent = 0
        for r in recipients:
            try:
                req.post(f'https://api.telegram.org/bot{config.bot_token}/sendMessage',
                    json={'chat_id': r.chat_id, 'text': message, 'parse_mode': 'HTML'}, timeout=10)
                sent += 1
            except:
                pass
        return Response({'sent': sent})

class GmarketControlStopView(views.APIView):
    """지마켓 광고제어(간편/AI/통합) 강제 중지 — 중지플래그 설정.
    실행 루프가 다음 계정 전에 확인해 중단(현재 계정은 최대 40초 내 마무리). 동시 11번가 크롤엔 영향 없음."""
    def post(self, request):
        from apps.cpc import eleven_block_guard as guard
        guard.request_control_stop('gmarket')
        return Response({'status': 'stopping',
                         'message': '강제중지 요청 — 현재 계정 처리(최대 40초) 후 중단됩니다.'})


class GmarketControlStatusView(views.APIView):
    """지마켓 광고제어 예약·실행 현황 — 상단 배너용(중복/겹침 확인)."""
    def get(self, request):
        import os, subprocess
        from apps.cpc.models import Cpc2Schedule, AiSchedule
        WD = {1: '월', 2: '화', 3: '수', 4: '목', 5: '금', 6: '토', 7: '일'}

        def days(w):
            return ''.join(WD[int(x)] for x in (w or [])) if w else '매일'

        # 실행 판정 = 실제 제어 프로세스 존재 여부(락 pid 재사용 오판 방지).
        # ps로 실제 명령 프로세스만 카운트(pgrep 자기매칭/ps 자신 제외).
        pats = ('crawl_gmarket_cpc2', 'run_ai_schedule', 'crawl_gmarket_ad_combined', 'gmarket_ai_control')
        procs = []
        try:
            out = subprocess.run(['ps', '-eo', 'pid,args'], capture_output=True, text=True).stdout
            for line in out.splitlines():
                if any(p in line for p in pats) and 'ps -eo' not in line and 'pgrep' not in line:
                    procs.append(line.strip())
        except Exception:
            pass
        proc_count = len(procs)

        # 광고제어 실행중 마커 — 대시보드 버튼은 스레드라 ps에 안 잡히므로 이 마커가 정확한 진행상태.
        from apps.cpc import eleven_block_guard as guard
        busy = guard.adcontrol_busy_info('gmarket')

        lockf = '/tmp/avengers_crawl_chrome_gmarket.lock'
        running = None
        if busy:
            since = busy['since'][11:19] if len(busy.get('since') or '') >= 19 else ''
            running = {'name': busy['name'], 'since': since}
        elif proc_count > 0:
            name, since = '실행중', ''
            try:
                raw = (open(lockf).read() or '').strip().split('|')
                if len(raw) > 1 and raw[1]:
                    name = raw[1]
                if len(raw) > 2 and len(raw[2]) >= 19:
                    since = raw[2][11:19]
            except Exception:
                pass
            running = {'name': name, 'since': since}
        else:
            # 프로세스 없는데 락만 남음 = 스테일(죽은/재사용 pid) → 자동 정리
            try:
                if os.path.exists(lockf):
                    os.remove(lockf)
            except Exception:
                pass

        c = Cpc2Schedule.objects.first()
        a = AiSchedule.objects.filter(platform='gmarket').first()
        cpc2 = None
        if c:
            cpc2 = {'on_time': str(c.on_time or '')[:5], 'off_time': str(c.off_time or '')[:5],
                    'on_days': days(c.weekdays), 'off_days': days(c.off_weekdays),
                    'accounts': len(c.selected_accounts or []), 'include_cpc1': c.include_cpc1}
        ai = None
        if a:
            ai = {'on_time': str(a.on_time or '')[:5], 'off_time': str(a.off_time or '')[:5],
                  'on_days': days(a.weekdays), 'off_days': days(a.off_weekdays),
                  'accounts': len(a.selected_accounts or [])}
        return Response({'running': running, 'proc_count': proc_count, 'cpc2': cpc2, 'ai': ai})


class Cpc2ControlView(views.APIView):
    def post(self, request):
        import threading as th
        from apps.cpc import eleven_block_guard as guard
        busy = guard.adcontrol_busy_info('gmarket')
        if busy:
            return Response({'status': 'busy',
                             'message': f'이미 광고제어 실행 중({busy["name"]}) — 끝난 뒤 다시 시도하세요.'},
                            status=409)
        action = request.data.get('action', 'on')
        accounts = request.data.get('accounts')
        source = request.data.get('source', 'manual')
        include_cpc1 = bool(request.data.get('include_cpc1', False))
        def run():
            from crawlers.gmarket_cpc2_control_crawler import run_control
            run_control(action, source, account_filter=accounts, include_cpc1=include_cpc1)
        th.Thread(target=run, daemon=True).start()
        return Response({'status': 'started', 'action': action, 'include_cpc1': include_cpc1})

class AiControlView(views.APIView):
    def post(self, request):
        import threading as th
        from apps.cpc import eleven_block_guard as guard
        busy = guard.adcontrol_busy_info('gmarket')
        if busy:
            return Response({'status': 'busy',
                             'message': f'이미 광고제어 실행 중({busy["name"]}) — 끝난 뒤 다시 시도하세요.'},
                            status=409)
        action = request.data.get('action', 'on')
        accounts = request.data.get('accounts')
        source = request.data.get('source', 'manual')
        def run():
            from crawlers.gmarket_ai_control_crawler import run_control
            run_control(action, source, account_filter=accounts)
        th.Thread(target=run, daemon=True).start()
        return Response({'status': 'started', 'action': action})


class St11AdStrategyCampaignsView(views.APIView):
    """캠페인 이름 목록.
    GET ?eid= : DB(St11AdofficeCampaign)에 저장된 이름 즉시 반환(최대100).
    POST {eid}: 광고센터에서 실시간 조회(백그라운드) 시작 → run_id 반환, 로그(status=CAMP) 폴링."""
    def get(self, request):
        from apps.cpc.models import St11AdofficeCampaign
        eid = request.query_params.get('eid', '')
        qs = St11AdofficeCampaign.objects.all()
        if eid:
            qs = qs.filter(eleven_id=eid)
        names = [n for n in qs.values_list('campaign_name', flat=True).distinct()[:100] if n]
        return Response({'eid': eid, 'campaigns': names})

    def post(self, request):
        import threading as th, time as _t
        eid = request.data.get('eid', '')
        if not eid:
            return Response({'error': '계정(eid)을 지정하세요.'}, status=400)
        run_id = _t.strftime('%Y%m%d%H%M%S')

        def run():
            from crawlers.eleven_ad_strategy import list_campaigns
            list_campaigns(eid, run_id=run_id)
        th.Thread(target=run, daemon=True).start()
        return Response({'status': 'started', 'run_id': run_id})


class St11AdStrategyAccountsView(views.APIView):
    """11번가 계정 목록을 '1등급>2>3>4>광고이력>나머지' 순으로 정렬해 반환(계정 선택 표시순)."""
    def get(self, request):
        from apps.cpc.models import CrawlerAccount, ElevenSellerGrade, St11ProductDaily
        from crawlers.eleven_ad_strategy import order_accounts
        eids = list(CrawlerAccount.objects.filter(platform='11st').values_list('login_id', flat=True))
        ordered = order_accounts(eids)
        grade_map = {}
        for r in ElevenSellerGrade.objects.order_by('eleven_id', '-collected_at').values('eleven_id', 'grade'):
            grade_map.setdefault(r['eleven_id'], r['grade'])
        adhist = set(St11ProductDaily.objects.values_list('eleven_id', flat=True).distinct())
        def bucket(e):
            g = grade_map.get(e)
            if g in (1, 2, 3, 4): return f'{g}등급'
            return '광고이력' if e in adhist else '나머지'
        return Response({'accounts': [
            {'login_id': e, 'grade': grade_map.get(e), 'bucket': bucket(e)} for e in ordered]})


class St11AdStrategyControlView(views.APIView):
    """11번가 광고그룹 노출 스케줄 전략 적용 실행(백그라운드).
    body: accounts[], campaigns[], on_start, on_end, weekdays[], execute(bool)"""
    def post(self, request):
        import threading as th, time as _t
        d = request.data
        accounts = d.get('accounts') or []
        campaigns = d.get('campaigns') or []
        on_start = int(d.get('on_start', 8))
        on_end = int(d.get('on_end', 16))
        weekdays = [int(w) for w in (d.get('weekdays') or [1, 2, 3, 4, 5])]
        execute = bool(d.get('execute', False))
        if not accounts or not campaigns:
            return Response({'error': '계정과 캠페인을 선택하세요.'}, status=400)
        run_id = _t.strftime('%Y%m%d%H%M%S')

        def run():
            from crawlers.eleven_ad_strategy import run_strategy
            run_strategy(accounts, campaigns, on_start=on_start, on_end=on_end,
                         weekdays=weekdays, execute=execute, run_id=run_id)
        th.Thread(target=run, daemon=True).start()
        return Response({'status': 'started', 'run_id': run_id,
                         'mode': '실제적용' if execute else '드라이런'})


class St11AdStrategyScheduleView(views.APIView):
    """11번가 전략설정 저장(예약). GET=조회, PUT/POST=저장(싱글톤 id=1).
    body: name, accounts[], campaigns[], on_start, on_end, weekdays[], enabled"""
    def _serialize(self, s):
        return {
            'id': s.id, 'name': s.name, 'accounts': s.accounts or [],
            'campaigns': s.campaigns or [], 'on_start': s.on_start, 'on_end': s.on_end,
            'weekdays': s.weekdays or [], 'enabled': s.enabled,
            'last_applied_at': s.last_applied_at.strftime('%Y-%m-%d %H:%M') if s.last_applied_at else None,
            'updated_at': s.updated_at.strftime('%Y-%m-%d %H:%M'),
        }

    def get(self, request):
        from apps.cpc.models import St11AdStrategySchedule
        s = St11AdStrategySchedule.objects.order_by('id').first()
        return Response({'schedule': self._serialize(s) if s else None})

    def post(self, request):
        return self.put(request)

    def put(self, request):
        from apps.cpc.models import St11AdStrategySchedule
        d = request.data
        s = St11AdStrategySchedule.objects.order_by('id').first() or St11AdStrategySchedule()
        s.name = d.get('name', s.name or '기본 전략')
        s.accounts = d.get('accounts', s.accounts) or []
        s.campaigns = d.get('campaigns', s.campaigns) or []
        s.on_start = int(d.get('on_start', s.on_start))
        s.on_end = int(d.get('on_end', s.on_end))
        s.weekdays = [int(w) for w in (d.get('weekdays', s.weekdays) or [])]
        if 'enabled' in d:
            s.enabled = bool(d.get('enabled'))
        s.save()
        return Response({'schedule': self._serialize(s)})


class St11AdStrategyRunsView(views.APIView):
    """전략 실행 내역 목록(최근 20건) — 미리보기/실제적용 진행상황 표시용."""
    def get(self, request):
        from apps.cpc.models import St11AdStrategyLog
        from django.utils import timezone as tz
        run_ids = list(St11AdStrategyLog.objects.values_list('run_id', flat=True)
                       .distinct().order_by('-run_id')[:20])
        out = []
        for rid in run_ids:
            rows = list(St11AdStrategyLog.objects.filter(run_id=rid).order_by('id'))
            if not rows:
                continue
            accts = sorted({r.eleven_id for r in rows if r.eleven_id})
            camps = sorted({r.campaign_name for r in rows if r.campaign_name})
            applied = sum(1 for r in rows if r.status == 'APPLIED')
            skip = sum(1 for r in rows if r.status == 'SKIP')
            err = sum(1 for r in rows if r.status == 'ERROR')
            mode = '드라이런'
            for r in rows:
                if r.status == 'START':
                    mode = '실제적용' if '실제적용' in r.detail else '드라이런'
                    break
            running = rows[-1].status != 'DONE'
            out.append({
                'run_id': rid, 'accounts': accts, 'campaigns': camps,
                'mode': mode, 'applied': applied, 'skip': skip, 'error': err,
                'running': running,
                'started': tz.localtime(rows[0].created_at).strftime('%m-%d %H:%M:%S'),
                'last': tz.localtime(rows[-1].created_at).strftime('%H:%M:%S'),
            })
        return Response({'runs': out})


class St11AdStrategyLogView(views.APIView):
    """전략 실행 로그 조회(폴링용). run_id 주면 그 실행만, 없으면 최근 200줄.
    running = 가장 최근 START 이후 DONE 이 아직 없으면 True."""
    def get(self, request):
        from apps.cpc.models import St11AdStrategyLog
        from django.utils import timezone as tz
        run_id = request.query_params.get('run_id', '')
        qs = St11AdStrategyLog.objects.all()
        if run_id:
            qs = qs.filter(run_id=run_id)
        rows = list(qs.order_by('-id')[:200])
        rows.reverse()
        data = [{'id': r.id, 'run_id': r.run_id, 'eid': r.eleven_id,
                 'campaign': r.campaign_name, 'group': r.group_name,
                 'status': r.status, 'detail': r.detail,
                 'at': tz.localtime(r.created_at).strftime('%H:%M:%S')} for r in rows]
        running = bool(rows) and rows[-1].status != 'DONE'
        return Response({'logs': data, 'running': running})


class GmarketAiViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GmarketAiAdSummary.objects.all()
    serializer_class = GmarketAiSummarySerializer
    filterset_fields = ['gmarket_id', 'actual_status']

class GmarketAiHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = GmarketAiHistorySerializer
    filterset_fields = ['gmarket_id']
    def get_queryset(self):
        return GmarketAiAdHistory.objects.all()[:200]

class St11CampaignViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = St11AdofficeCampaign.objects.all()
    serializer_class = St11CampaignSerializer
    filterset_fields = ['eleven_id', 'is_ai']

CRAWL_LOCKFILE = '/tmp/avengers_crawl_chrome.lock'


def _crawl_lock_busy():
    """크롬 크롤러 락이 살아있는 프로세스에 잡혀 있으면 (pid, True), 아니면 (None, False).
    cron 스크립트와 동일한 락 파일을 공유 → 수동/자동 크롤이 서로 겹치지 않음."""
    import os
    if not os.path.exists(CRAWL_LOCKFILE):
        return None, False
    try:
        # 락 형식은 'pid|name|time' (eleven_block_guard.acquire_global_lock) — 첫 필드가 pid.
        # 통째 int 변환하면 ValueError로 멀쩡한 락을 stale로 오인·삭제했었음(2026-06-11 수정).
        raw = (open(CRAWL_LOCKFILE).read() or '').strip()
        pid = int((raw.split('|')[0] or '0').strip() or 0)
    except Exception:
        pid = 0
    if pid > 0:
        try:
            os.kill(pid, 0)        # 살아있음 → 실행 중
            return pid, True
        except OSError:
            pass
    # 죽은 PID / 손상 → stale 락 정리
    try:
        os.remove(CRAWL_LOCKFILE)
    except Exception:
        pass
    return None, False


class CrawlTriggerView(views.APIView):
    def post(self, request):
        import os
        platform = request.data.get('platform', 'gmarket')
        crawl_type = request.data.get('type', 'cost')  # cost, grade, or ai
        accounts_filter = request.data.get('accounts')

        # ── 중복 실행 방지: 이미 다른 크롤러(수동/자동)가 돌고 있으면 거부 ──
        busy_pid, busy = _crawl_lock_busy()
        if busy:
            return Response(
                {'status': 'busy',
                 'error': f'이미 다른 크롤러가 실행 중입니다 (PID={busy_pid}). 끝난 뒤 다시 시도하세요.'},
                status=409)

        # 락 획득 (백엔드 프로세스 PID 기록 → cron 스크립트도 kill -0 로 인식해 스킵)
        try:
            with open(CRAWL_LOCKFILE, 'w') as f:
                f.write(str(os.getpid()))
        except Exception:
            pass

        def run():
            try:
                if crawl_type == 'product':
                    # 11번가 등록상품(대량엑셀) 재크롤 — 선택 계정만, 강제 재수집
                    from crawlers.eleven_product_crawler import run_all_accounts as _prun
                    _prun(account_filter=accounts_filter, only_no_api_key=False, force=True)
                    return
                if crawl_type == 'cpc_status':
                    from crawlers.gmarket_cpc_status_crawler import run_all_accounts
                elif crawl_type == 'ai':
                    if platform == 'gmarket':
                        from crawlers.gmarket_ai_crawler import run_all_accounts
                    else:
                        from crawlers.eleven_ai_crawler import run_all_accounts
                elif crawl_type == 'grade':
                    if platform == 'gmarket':
                        from crawlers.gmarket_grade_crawler import run_all_accounts
                    else:
                        from crawlers.eleven_grade_crawler import run_all_accounts
                else:
                    if platform == 'gmarket':
                        from crawlers.gmarket_crawler import run_all_accounts
                    else:
                        from crawlers.eleven_crawler import run_all_accounts
                run_all_accounts(account_filter=accounts_filter)
            finally:
                # 작업 끝나면 락 해제
                try:
                    os.remove(CRAWL_LOCKFILE)
                except Exception:
                    pass

        threading.Thread(target=run, daemon=True).start()
        return Response({'status': 'started', 'platform': platform, 'type': crawl_type})


class St11ProductDailyCrawlView(views.APIView):
    """상품/키워드 ROAS 기간지정 수동 크롤 (일별). 강제중지는 St11CrawlStopView 공유."""
    def post(self, request):
        import os, re, subprocess
        busy_pid, busy = _crawl_lock_busy()
        if busy:
            return Response({'status': 'busy', 'error': f'이미 크롤러 실행 중입니다 (PID={busy_pid}).'}, status=409)
        df = (request.data.get('date_from') or '').strip()
        dt = (request.data.get('date_to') or '').strip()
        accts = request.data.get('accounts') or []
        datere = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        if df and not datere.match(df):
            return Response({'error': '시작일 형식 오류'}, status=400)
        if dt and not datere.match(dt):
            return Response({'error': '종료일 형식 오류'}, status=400)
        accts = [a for a in accts if re.match(r'^[A-Za-z0-9_]+$', str(a))]
        args = 'manage.py crawl_11st_product_daily'
        if df:
            args += f' --from {df}'
        if dt:
            args += f' --to {dt}'
        if accts:
            args += ' --accounts ' + ' '.join(accts)
        script = (f'echo $$ > {CRAWL_LOCKFILE}; trap "rm -f {CRAWL_LOCKFILE}" EXIT; '
                  f'cd /home/rejoice888/Avengers/backend && /usr/bin/python3 {args} '
                  f'>> /tmp/cron_11st_product_daily.log 2>&1')
        try:
            subprocess.Popen(['bash', '-c', script], start_new_session=True,
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            return Response({'status': 'error', 'error': str(e)}, status=500)
        return Response({'status': 'started', 'date_from': df or '2026-01-01', 'date_to': dt or 'yesterday'})


class ElevenLossDeleteView(views.APIView):
    """적자상품 자동 판매중지·삭제 실행 (셀러오피스).
    안전: 셀러오피스 삭제 플로우 검증 전이므로 기본 dry-run(1상품) 으로만 실행.
    실삭제는 ?real=1 + 검증 완료 후 활성화."""
    def post(self, request):
        import re, subprocess
        d = request.data
        df = (d.get('date_from') or '2026-01-01').strip()
        dt = (d.get('date_to') or '').strip()
        datere = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        if not datere.match(df) or (dt and not datere.match(dt)):
            return Response({'error': '날짜 형식 오류'}, status=400)
        rmax = float(d.get('roas_max', 100)); cmin = float(d.get('cost_min', 2000)); kmin = float(d.get('clicks_min', 10))
        real = str(d.get('real', '')).lower() in ('1', 'true', 'yes')
        # 실삭제 활성화됨. 프론트에서 "검증(dry-run) → 확인 → 실삭제" 다단계 확인을 거침.
        # 안전: 실삭제도 limit(소량)을 받아 첫 실행은 소수만 삭제해 실제 작동을 검증할 수 있음.
        VERIFIED = True
        if real and not VERIFIED:
            return Response({'status': 'blocked',
                             'message': '⚠️ 실삭제는 아직 비활성화 상태입니다. 1상품 dry-run으로 검증한 뒤 활성화됩니다.'}, status=400)
        # ★ 상품번호 지정 삭제(나의상품 선택삭제) — date/적자 산출 없이 그 상품만
        pnos = [re.sub(r'\D', '', str(p)) for p in (d.get('product_nos') or [])]
        pnos = [p for p in pnos if p]
        if pnos:
            eid = str(d.get('eid') or '').strip()
            if not eid or not re.match(r'^[A-Za-z0-9_]+$', eid):
                return Response({'error': '상품지정 삭제는 eid(계정) 필수'}, status=400)
            a = f'manage.py delete_loss_products --eid {eid} --product-nos ' + ' '.join(pnos[:300])
            if real:
                a += ' --real'
            sc = (f'cd /home/rejoice888/Avengers/backend && /usr/bin/python3 {a} >> /tmp/delete_loss.log 2>&1')
            try:
                subprocess.Popen(['bash', '-c', sc], start_new_session=True,
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception as e:
                return Response({'status': 'error', 'error': str(e)}, status=500)
            return Response({'status': 'started',
                             'message': f'🗑 11번가 지정상품 {len(pnos)}개 {"실삭제" if real else "검증(dry-run)"} 시작 — 텔레그램/로그로 확인하세요.'})
        try:
            lim = int(d.get('limit')) if str(d.get('limit') or '').strip() else None
        except (ValueError, TypeError):
            lim = None
        # dry-run(real 아님)=1상품 검증 / real+limit=소량 실삭제 / real+limit없음=전체 실삭제
        limit = 1 if not real else lim
        args = (f'manage.py delete_loss_products --from {df}'
                + (f' --to {dt}' if dt else '')
                + f' --roas-max {rmax} --cost-min {cmin} --clicks-min {kmin}'
                + (f' --limit {limit}' if limit else '')
                + (' --real' if real else ''))
        script = (f'cd /home/rejoice888/Avengers/backend && /usr/bin/python3 {args} '
                  f'>> /tmp/delete_loss.log 2>&1')
        try:
            subprocess.Popen(['bash', '-c', script], start_new_session=True,
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            return Response({'status': 'error', 'error': str(e)}, status=500)
        if real:
            scope = f'소량 {limit}개(테스트)' if limit else '전체'
            msg = f'🗑 실삭제 시작 — {scope} 상품을 셀러오피스에서 판매중지+삭제합니다. 진행상황은 텔레그램/로그(/tmp/delete_loss.log)로 확인하세요.'
        else:
            msg = '🔎 검증(dry-run) 실행 — 1개 상품으로 셀러오피스 접속·검색·셀렉터를 확인합니다(삭제 안 함). 결과는 텔레그램/로그로 확인하세요.'
        return Response({'status': 'started', 'message': msg})


class ElevenLossMarkDeletedView(views.APIView):
    """적자상품 삭제완료 표시 — 11번가에서 삭제한 상품을 기록(비고 '삭제완료' 파란색).
    body: {product_nos:[...], eleven_id?} 또는 {all:true}(현재 적자 전체 삭제완료 처리)."""
    def post(self, request):
        from apps.cpc.models import St11LossDeleted
        d = request.data
        eid = d.get('eleven_id') or ''
        pnos = d.get('product_nos') or []
        items = d.get('items') or []   # [{eleven_id, product_no}, ...] — 화면에 보이는 목록 처리
        marked = 0
        if items:
            for it in items:
                e = it.get('eleven_id') or ''; p = str(it.get('product_no') or '')
                if not e or not p:
                    continue
                St11LossDeleted.objects.get_or_create(
                    eleven_id=e, product_no=p, defaults={'seller_code': it.get('seller_code') or ''})
                marked += 1
        elif d.get('all'):
            # 현재 적자 전체를 삭제완료 처리 (판매금지는 별도처리 → 제외)
            from datetime import date, timedelta
            d0 = date(2026, 1, 1); d1 = timezone.localdate() - timedelta(days=1)
            for e in ([eid] if eid else _active_eids()):
                rows = _eleven_product_rows(e, d0, d1, None, 100, 2000, 10)
                for r in rows:
                    if r.get('status') == '판매금지':
                        continue
                    St11LossDeleted.objects.get_or_create(
                        eleven_id=e, product_no=str(r['product_no']),
                        defaults={'seller_code': r['seller_code'] or ''})
                    marked += 1
        else:
            for p in pnos:
                St11LossDeleted.objects.get_or_create(eleven_id=eid, product_no=str(p))
                marked += 1
        return Response({'status': 'ok', 'marked': marked, 'message': f'{marked}개 삭제완료 처리됨 (비고 표시)'})


class St11CostCrawlView(views.APIView):
    """11번가 광고비 수동 크롤 — 매시간 cron과 동일한 스크립트를 별도 프로세스로 실행.
    별도 프로세스(세션 분리)라 강제중지(St11CrawlStopView)로 깔끔히 종료 가능."""
    def post(self, request):
        import os, subprocess
        busy_pid, busy = _crawl_lock_busy()
        if busy:
            return Response(
                {'status': 'busy', 'error': f'이미 크롤러가 실행 중입니다 (PID={busy_pid}).'},
                status=409)
        # 수동 실행은 '오늘 스킵' 설정을 무시하고 바로 실행
        try:
            os.remove('/tmp/avengers_11st_cost_skip')
        except Exception:
            pass
        script = '/home/rejoice888/Avengers/backend/scripts/cron_11st_cost.sh'
        try:
            subprocess.Popen(
                ['bash', script], start_new_session=True,
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            return Response({'status': 'error', 'error': str(e)}, status=500)
        return Response({'status': 'started'})


class St11CrawlStopView(views.APIView):
    """실행 중인 크롤을 강제 중지 — 락 PID의 프로세스그룹 종료 + 크롬 정리."""
    def post(self, request):
        import os, signal, time, subprocess
        pid, busy = _crawl_lock_busy()
        if not busy:
            return Response({'status': 'idle', 'message': '실행 중인 크롤이 없습니다.'})
        # 백엔드 자신(스레드 모드 크롤)이면 죽이면 안 됨 → 크롬만 종료해 중단 유도
        if pid == os.getpid():
            subprocess.call(['pkill', '-f', 'org.chromium.Chromium.scoped_dir'])
            subprocess.call(['pkill', '-f', 'chromedriver'])
            return Response({'status': 'stopped', 'message': '진행 중 크롤의 크롬을 종료했습니다.'})
        try:
            os.killpg(os.getpgid(pid), signal.SIGTERM)
            time.sleep(2)
            try:
                os.kill(pid, 0)
                os.killpg(os.getpgid(pid), signal.SIGKILL)
            except OSError:
                pass
        except Exception:
            try:
                os.kill(pid, signal.SIGKILL)
            except Exception:
                pass
        subprocess.call(['pkill', '-f', 'org.chromium.Chromium.scoped_dir'])
        subprocess.call(['pkill', '-f', 'chromedriver'])
        try:
            os.remove(CRAWL_LOCKFILE)
        except Exception:
            pass
        return Response({'status': 'stopped', 'pid': pid})


class St11CrawlStatusView(views.APIView):
    """크롤 실행 여부 + 광고비 최종 수집 종료시간 조회 (버튼/현황 표시용)."""
    def get(self, request):
        from .models import ElevenCostHistory
        from django.db.models import Max
        pid, busy = _crawl_lock_busy()
        last = ElevenCostHistory.objects.aggregate(m=Max('created_at'))['m']
        return Response({'running': busy, 'pid': pid, 'last_collected_at': last})


class ElevenRoasView(views.APIView):
    """11번가 계정별 ROAS(광고비 대비 매출). 매출·광고비를 login_id 기준으로 정확히 조인.
    period=month(기본)|today|range(date_from,date_to). 목표 ROAS 500%."""
    def get(self, request):
        from apps.sales.models import SalesRecord
        from django.db.models import Sum
        import pytz
        from datetime import datetime as _dt, timedelta, date as _date
        kst = pytz.timezone('Asia/Seoul')
        now = timezone.localtime()
        period = request.query_params.get('period', 'month')
        if period == 'today':
            d0 = d1 = now.date()
        elif period == 'range':
            try:
                d0 = _dt.strptime(request.query_params.get('date_from'), '%Y-%m-%d').date()
                d1 = _dt.strptime(request.query_params.get('date_to'), '%Y-%m-%d').date()
            except Exception:
                d0 = _date(now.year, now.month, 1); d1 = now.date()
        else:
            d0 = _date(now.year, now.month, 1); d1 = now.date()
        ms = kst.localize(_dt.combine(d0, _dt.min.time()))
        me = kst.localize(_dt.combine(d1, _dt.min.time())) + timedelta(days=1)

        # 매출/순익 — login_id(SellerAccount.seller_id) 기준 집계
        sales_map = {}
        for r in (SalesRecord.objects
                  .filter(platform='11st', order_date__gte=d0, order_date__lte=d1)
                  .values('seller__seller_id')
                  .annotate(s=Sum('total_price'), p=Sum('net_profit'))):
            sid = r['seller__seller_id']
            if sid:
                sales_map[sid] = (r['s'] or 0, r['p'] or 0)
        # 광고비(CPC) — login_id 기준 집계
        cost_map = {}
        for r in (ElevenCostHistory.objects
                  .filter(transaction_type='CPC', transaction_datetime__gte=ms, transaction_datetime__lt=me)
                  .values('seller_id').annotate(c=Sum('amount'))):
            cost_map[r['seller_id']] = abs(r['c'] or 0)

        # 광고전환매출(광고센터/adoffice) — St11ProductDaily.conv_amount, login_id(eleven_id) 기준 집계
        from apps.cpc.models import St11ProductDaily
        conv_map = {}
        for r in (St11ProductDaily.objects.filter(stat_date__gte=d0, stat_date__lte=d1)
                  .values('eleven_id').annotate(cv=Sum('conv_amount'))):
            conv_map[r['eleven_id']] = r['cv'] or 0

        # 셀러 등급 — 계정별 최신 등급
        from apps.cpc.models import ElevenSellerGrade
        grade_map = {}
        for g in (ElevenSellerGrade.objects.filter(grade__isnull=False)
                  .order_by('collected_at').values('eleven_id', 'grade')):
            grade_map[g['eleven_id']] = g['grade']

        rows = []
        tot_sales = tot_cost = tot_profit = tot_conv = 0
        for a in CrawlerAccount.objects.filter(platform='11st', is_active=True):
            lid = a.login_id
            sales, profit = sales_map.get(lid, (0, 0))
            cost = cost_map.get(lid, 0)
            conv = conv_map.get(lid, 0)
            if sales == 0 and cost == 0 and conv == 0:
                continue
            roas = round(sales / cost * 100) if cost else None
            rows.append({
                'login_id': lid, 'name': a.seller_name or lid,
                'grade': grade_map.get(lid),
                'sales': sales, 'profit': profit, 'cost': cost,
                'roas': roas, 'net_after_ad': profit - cost,
                'conv_amount': conv,                                      # 광고전환매출(광고센터)
                'conv_roas': round(conv / cost * 100) if cost else None,  # 광고전환ROAS
            })
            tot_sales += sales; tot_cost += cost; tot_profit += profit; tot_conv += conv
        # 낮은 ROAS(적자) 먼저 — 광고비 있는데 ROAS 없는(매출0) 건 최상단
        rows.sort(key=lambda x: (x['roas'] is not None, x['roas'] if x['roas'] is not None else -1))
        return Response({
            'period': period, 'date_from': str(d0), 'date_to': str(d1),
            'target_roas': 500,
            'totals': {
                'sales': tot_sales, 'cost': tot_cost, 'profit': tot_profit,
                'roas': round(tot_sales / tot_cost * 100) if tot_cost else None,
                'net_after_ad': tot_profit - tot_cost,
                'conv_amount': tot_conv,
                'conv_roas': round(tot_conv / tot_cost * 100) if tot_cost else None,
                'count': len(rows),
                'above_target': sum(1 for x in rows if x['roas'] is not None and x['roas'] >= 500),
                'below_target': sum(1 for x in rows if x['cost'] > 0 and (x['roas'] is None or x['roas'] < 500)),
            },
            'rows': rows,
        })


def _roas_period(request):
    from datetime import datetime as _dt, timedelta
    today = timezone.localdate()
    def _pd(s, d):
        try:
            return _dt.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return d
    d_to = _pd(request.query_params.get('date_to'), today - timedelta(days=1))
    d_from = _pd(request.query_params.get('date_from'), d_to - timedelta(days=29))
    def _f(v):
        try:
            return float(v)
        except Exception:
            return None
    return d_from, d_to, _f


def _active_eids():
    # 영구정지 계정은 AD OFFICE 접속 불가 → 광고비/ROAS 판단(적자 산출)에서 제외
    from apps.cpc.eleven_block_guard import exclude_perma_banned
    qs = exclude_perma_banned(CrawlerAccount.objects.filter(platform='11st', is_active=True))
    return list(qs.values_list('login_id', flat=True))


def _bare_seller_code(c):
    """판매자코드에서 WDM_/auto_ 접두어 제거 (매출자료는 접두어 없는 W코드로 저장됨).
    LCE_ 등 다른 접두어는 매출자료에도 있으므로 건드리지 않음."""
    if not c:
        return c
    cl = c.lower()
    for pre in ('wdm_', 'auto_'):
        if cl.startswith(pre):
            return c[len(pre):]
    return c


def _norm_pname(s):
    """상품명 정규화(옵션/구분기호 제거) — 중복등록 판정용."""
    import re as _re
    s = (s or '').lower()
    s = _re.sub(r'☞.*$', '', s)
    s = _re.sub(r'[\s/,\-_()\[\]]', '', s)
    return s


def _mark_loss_duplicates(rows, d_from, d_to):
    """매출 0인 적자행 중, '같은 상품명이 다른 판매자코드로 팔리는' 중복등록을 표시.
    status='중복등록', dup=True 로 마킹(삭제 제외용). 결정적(데이터 동일하면 결과 동일)."""
    from apps.sales.models import SalesRecord
    from django.db.models import Sum
    idx = {}
    for s in (SalesRecord.objects.filter(platform='11st', order_date__gte=d_from, order_date__lte=d_to)
              .exclude(product_code='').values('product_code', 'product_name').annotate(a=Sum('total_price'))):
        k = _norm_pname(s['product_name'])
        if len(k) < 8:
            continue
        e = idx.setdefault(k, {'codes': set(), 'amt': 0})
        e['codes'].add(_bare_seller_code(s['product_code']))
        e['amt'] += s['a'] or 0
    for r in rows:
        if r.get('sales', 0) > 0:
            continue
        k = _norm_pname(r.get('product_name'))
        if len(k) < 8:
            continue
        hit = idx.get(k)
        if hit and abs(hit['amt']) > 0 and (hit['codes'] - {_bare_seller_code(r.get('seller_code') or '')}):
            r['status'] = '중복등록'
            r['dup'] = True
    return rows


def _eleven_product_rows(eid, d_from, d_to, rmin, rmax, cmin, kmin):
    """단일 계정 상품 ROAS 행 (광고비=adoffice, 매출=SalesRecord). 필터 적용."""
    from apps.cpc.models import St11ProductDaily, ElevenMyProduct, St11LossDeleted
    from apps.sales.models import SalesRecord
    from django.db.models import Sum, Count
    import re as _re
    ad = list(St11ProductDaily.objects.filter(eleven_id=eid, stat_date__gte=d_from, stat_date__lte=d_to)
              .values('product_no').annotate(cost=Sum('cost'), clicks=Sum('clicks'), impressions=Sum('impressions'),
                                             conversions=Sum('conversions'), conv_amount=Sum('conv_amount')))
    if not ad:
        return []
    # 상품번호의 '(삭제됨)' 등 괄호표기를 떼어 순수 숫자로 합산(삭제 전/후 같은 상품 병합).
    # 삭제여부는 비고로 보존 → 상품번호 검색 가능 + 매핑/매출도 정상 연결.
    agg = {}
    for x in ad:
        rawno = str(x['product_no'])
        clean = _re.sub(r'\s*\(.*?\)\s*', '', rawno).strip()
        a = agg.setdefault(clean, {'cost': 0, 'clicks': 0, 'impressions': 0,
                                   'conversions': 0, 'conv_amount': 0, 'deleted': False})
        a['cost'] += x['cost'] or 0
        a['clicks'] += x['clicks'] or 0
        a['impressions'] += x['impressions'] or 0
        a['conversions'] += x['conversions'] or 0
        a['conv_amount'] += x['conv_amount'] or 0
        if '삭제' in rawno:
            a['deleted'] = True
    pnos = [int(k) for k in agg if k.isdigit()]
    # 상품번호→판매자코드 다리는 전역 조회(11번가 상품번호는 전역 고유). 광고계정의 내상품DB에
    # 없어도 다른 계정에 등록돼 있으면 복구 → '상품번호 인식 못함' 방지.
    code_map = {}
    if pnos:
        for p in (ElevenMyProduct.objects.filter(product_no__in=pnos).exclude(seller_product_code='')
                  .values('product_no', 'seller_product_code', 'product_name', 'status_type')):
            code_map.setdefault(str(p['product_no']),
                                (p['seller_product_code'] or '', p['product_name'] or '', p['status_type'] or ''))
    # 비고(판매상태)는 '상품코드(product_no)'만으로 매칭 — 판매자코드 유무와 무관.
    # → crawl_11st_products가 eleven_my_product.status_type를 갱신하면 비고에 즉시 반영됨.
    status_by_pno = {}
    if pnos:
        for p in (ElevenMyProduct.objects.filter(product_no__in=pnos)
                  .values('product_no', 'status_type')):
            status_by_pno.setdefault(str(p['product_no']), p['status_type'] or '')
    # 카탈로그에서 삭제된 상품은 영구보존고(ProductCodeArchive)에서 판매자코드 보충 → 빈칸 자동채움
    miss = [str(p) for p in pnos if str(p) not in code_map]
    if miss:
        from apps.cpc.models import ProductCodeArchive
        for p in (ProductCodeArchive.objects.filter(platform='11st', product_no__in=miss)
                  .exclude(seller_code='').values('product_no', 'seller_code', 'product_name')):
            code_map.setdefault(str(p['product_no']),
                                (p['seller_code'] or '', p['product_name'] or '', '삭제(코드보존)'))
    # 매출은 판매자코드 기준 전역 매칭: 같은 상품(판매자코드)이 광고계정과 다른 계정에
    # 매출로 기록되는 경우가 많아(예: 광고=tmxk26, 매출=tmxk24), 계정한정 매칭 시
    # 매출 0 → 적자 오분류가 발생. 판매자코드는 사실상 상품과 1:1(상품명 2개이상 1%, 옵션차이뿐).
    # 매출 매칭 코드 후보: WDM_/auto_ 접두어는 매출자료엔 없으므로 벗긴 코드도 포함.
    codes = set()
    for v in code_map.values():
        c = v[0]
        if c:
            codes.add(c)
            codes.add(_bare_seller_code(c))
    sales_by_code = {}
    sales_cnt_by_code = {}
    if codes:
        for s in (SalesRecord.objects.filter(platform='11st', product_code__in=list(codes),
                                             order_date__gte=d_from, order_date__lte=d_to)
                  .values('product_code').annotate(s=Sum('total_price'), oc=Count('id'))):
            sales_by_code[s['product_code']] = s['s'] or 0
            sales_cnt_by_code[s['product_code']] = s['oc'] or 0
    deleted_done = set(St11LossDeleted.objects.filter(eleven_id=eid).values_list('product_no', flat=True))
    rows = []
    for pno, a in agg.items():
        cost = a['cost']; clk = a['clicks']
        sc, nm, st = code_map.get(pno, ('', '', ''))
        mapped = bool(sc)   # 판매자코드 다리 유무 (없으면 매출 산출 불가)
        # 비고(상태): 매일 크롤한 '실제 등록상태'(내상품DB, st)를 최우선 → 믿을 수 있는 자료.
        # st가 실판매상태(판매중/판매중지/품절/판매금지/재고부족)면 그대로 사용.
        # '삭제완료'(우리 삭제기록)는 카탈로그에 실제로 없을 때만 표시(과거엔 우선시해 오표시됐음).
        real = status_by_pno.get(pno, '')   # 상품코드(product_no)로 매칭된 실제 현재 등록상태
        if real:
            status = real
        elif a['deleted']:
            status = '삭제됨'
        elif pno in deleted_done:
            status = '삭제완료'
        elif st == '삭제(코드보존)':
            status = '삭제(코드보존)'
        elif not mapped:
            status = '미등록'
        else:
            status = '판매중'
        raw = sum(sales_by_code.get(x, 0) for x in {sc, _bare_seller_code(sc)}) if sc else 0
        real_orders = sum(sales_cnt_by_code.get(x, 0) for x in {sc, _bare_seller_code(sc)}) if sc else 0
        sales = abs(raw)
        roas = round(sales / cost * 100, 2) if cost else 0
        # 광고전환매출(adoffice 광고센터 집계: 직접+간접 전환) → 광고전환 ROAS
        conv_amount = abs(a['conv_amount'])
        conv_roas = round(conv_amount / cost * 100, 2) if cost else 0
        # ROAS 필터(적자판단 등)는 매출을 알 수 있는(매핑된) 상품에만 적용.
        # 매핑 없으면 매출 불명 → 적자(ROAS≤100)로 단정하지 않음(오분류 방지).
        if (rmin is not None or rmax is not None) and not mapped:
            continue
        if rmin is not None and roas < rmin:
            continue
        if rmax is not None and roas > rmax:
            continue
        if cmin is not None and cost < cmin:
            continue
        if kmin is not None and clk < kmin:
            continue
        rows.append({'eleven_id': eid, 'product_no': pno, 'seller_code': sc, 'product_name': nm,
                     'mapped': mapped, 'deleted': a['deleted'], 'status': status, 'cost': cost, 'sales': sales,
                     'roas_pct': roas, 'conversions': a['conversions'], 'conv_amount': conv_amount,
                     'conv_roas_pct': conv_roas, 'clicks': clk, 'impressions': a['impressions'],
                     'real_orders': real_orders,
                     'avg_click_cost': round(cost / clk) if clk else 0})
    return rows


def _eleven_keyword_rows(eid, d_from, d_to, rmin, rmax):
    """단일 계정 키워드 ROAS 행 (adoffice 광고전환매출 기준)."""
    from apps.cpc.models import St11KeywordDaily, ElevenMyProduct
    from django.db.models import Sum
    import re as _re
    raw = St11KeywordDaily.objects.filter(eleven_id=eid, stat_date__gte=d_from, stat_date__lte=d_to) \
        .values('product_no', 'keyword').annotate(
            cost=Sum('cost'), conv_amount=Sum('conv_amount'),
            clicks=Sum('clicks'), conversions=Sum('conversions'))
    # 상품번호 '(삭제됨)' 표기 제거 후 (순수번호,키워드)로 병합, 삭제여부는 비고로 보존
    agg = {}
    for x in raw:
        rawno = str(x['product_no'])
        clean = _re.sub(r'\s*\(.*?\)\s*', '', rawno).strip()
        d = agg.setdefault((clean, x['keyword']),
                           {'cost': 0, 'conv_amount': 0, 'clicks': 0, 'conversions': 0, 'deleted': False})
        d['cost'] += x['cost'] or 0
        d['conv_amount'] += x['conv_amount'] or 0
        d['clicks'] += x['clicks'] or 0
        d['conversions'] += x['conversions'] or 0
        if '삭제' in rawno:
            d['deleted'] = True
    if not agg:
        return []
    pnos = [int(k[0]) for k in agg if k[0].isdigit()]
    smap = {}
    if pnos:
        for p in (ElevenMyProduct.objects.filter(product_no__in=pnos).exclude(seller_product_code='')
                  .values('product_no', 'seller_product_code')):
            smap.setdefault(str(p['product_no']), p['seller_product_code'] or '')
    rows = []
    for (pno, kw), d in agg.items():
        c = d['cost']; a = d['conv_amount']
        roas = round(a / c * 100, 2) if c else 0
        if rmin is not None and roas < rmin:
            continue
        if rmax is not None and roas > rmax:
            continue
        rows.append({'eleven_id': eid, 'product_no': pno, 'seller_code': smap.get(pno, ''),
                     'keyword': kw, 'deleted': d['deleted'], 'cost': c, 'conv_amount': a,
                     'roas_pct': roas, 'clicks': d['clicks'], 'conversions': d['conversions']})
    return rows


class ElevenProductRoasView(views.APIView):
    """상품코드별 ROAS — 광고비=adoffice, 매출=직접입력. eleven_id 없으면 전체 계정.
    ?eleven_id&date_from&date_to&roas_min&roas_max&cost_min&clicks_min&export=1"""
    def get(self, request):
        from apps.cpc.models import St11ProductDaily
        from apps.sales.models import SalesRecord
        from django.db.models import Sum, Max
        eid = request.query_params.get('eleven_id')
        d_from, d_to, _f = _roas_period(request)
        rmin = _f(request.query_params.get('roas_min')); rmax = _f(request.query_params.get('roas_max'))
        cmin = _f(request.query_params.get('cost_min')); kmin = _f(request.query_params.get('clicks_min'))
        eids = [eid] if eid else _active_eids()
        rows = []
        for e in eids:
            rows += _eleven_product_rows(e, d_from, d_to, rmin, rmax, cmin, kmin)
        # 정책: 중복이어도 그 상품코드로 광고비 효율 없으면(매출 적으면) 삭제 대상 → 중복 제외 안 함
        rows.sort(key=lambda x: (-x['cost'], str(x['product_no'])))

        if request.query_params.get('export'):
            import csv as _csv, re as _re2
            from django.http import HttpResponse
            resp = HttpResponse(content_type='text/csv; charset=utf-8')
            w = _csv.writer(resp)

            # 키워드별 펼침(롱포맷): 한 상품에 키워드 N개면 N행, 상품정보는 각 행 반복.
            # 키워드는 (계정,상품번호,키워드) 단위라 중복 자동 제거. 키워드 없는 상품은 키워드 빈칸 1행.
            if request.query_params.get('expand_kw'):
                from apps.cpc.models import St11KeywordDaily
                resp['Content-Disposition'] = f'attachment; filename="roas_kw_{eid or "all"}_{d_from}_{d_to}.csv"'
                resp.write('﻿')
                pairs = {(r['eleven_id'], str(r['product_no'])) for r in rows}
                kw_by = {}
                for k in (St11KeywordDaily.objects.filter(stat_date__gte=d_from, stat_date__lte=d_to)
                          .values('eleven_id', 'product_no', 'keyword').annotate(c=Sum('cost'))):
                    pnok = _re2.sub(r'\s*\(.*?\)\s*', '', str(k['product_no'])).strip()
                    key = (k['eleven_id'], pnok)
                    if key in pairs and k['keyword']:
                        kw_by.setdefault(key, []).append((k['keyword'], k['c'] or 0))
                for key in kw_by:
                    kw_by[key].sort(key=lambda t: -t[1])   # 광고비순
                w.writerow(['계정', '상품번호', '판매자코드', '평균단가', '광고비', '키워드', '클릭',
                            '구매수(광고센터)', '구매금액(광고센터)', 'ROAS(광고센터)',
                            '실구매건수(참고)', '실매출(참고)', '실ROAS(참고)', '비고'])
                for x in rows:
                    head = [x['eleven_id'], x['product_no'], x['seller_code'],
                            x.get('avg_click_cost', 0), x['cost']]
                    tail = [x['clicks'], x.get('conversions', 0), x.get('conv_amount', 0),
                            x.get('conv_roas_pct', 0), x.get('real_orders', 0), x['sales'], x['roas_pct'],
                            x.get('status') or ('삭제됨' if x.get('deleted') else '')]
                    kws = kw_by.get((x['eleven_id'], str(x['product_no'])), [])
                    if kws:
                        for kw, _c in kws:
                            w.writerow(head + [kw] + tail)
                    else:
                        w.writerow(head + [''] + tail)
                return resp

            resp['Content-Disposition'] = f'attachment; filename="roas_{eid or "all"}_{d_from}_{d_to}.csv"'
            resp.write('﻿')
            w.writerow(['아이디', '상품번호', '판매자코드', '상품명', '광고비',
                        '광고전환매출', '광고전환ROAS(%)', '전환수', '실매출', '실매출ROAS(%)',
                        '클릭', '노출', '비고'])
            for x in rows:
                w.writerow([x['eleven_id'], x['product_no'], x['seller_code'], x['product_name'], x['cost'],
                            x.get('conv_amount', 0), x.get('conv_roas_pct', 0), x.get('conversions', 0),
                            x['sales'], x['roas_pct'], x['clicks'], x['impressions'],
                            x.get('status') or ('삭제됨' if x.get('deleted') else '')])
            return resp

        if eid:
            tot_cost = St11ProductDaily.objects.filter(eleven_id=eid, stat_date__gte=d_from, stat_date__lte=d_to).aggregate(s=Sum('cost'))['s'] or 0
            tot_sales = SalesRecord.objects.filter(platform='11st', seller__seller_id=eid, order_date__gte=d_from, order_date__lte=d_to).aggregate(s=Sum('total_price'))['s'] or 0
            collected = St11ProductDaily.objects.filter(eleven_id=eid, stat_date__gte=d_from, stat_date__lte=d_to).aggregate(m=Max('collected_at'))['m']
        else:
            tot_cost = sum(r['cost'] for r in rows); tot_sales = sum(r['sales'] for r in rows); collected = None
        tot_conv = sum(r.get('conv_amount', 0) for r in rows)
        return Response({
            'eleven_id': eid, 'date_from': str(d_from), 'date_to': str(d_to),
            'count': len(rows), 'collected_at': collected,
            'totals': {'cost': tot_cost, 'sales': abs(tot_sales),
                       'roas': round(abs(tot_sales) / tot_cost * 100, 1) if tot_cost else None,
                       'conv_amount': tot_conv,
                       'conv_roas': round(tot_conv / tot_cost * 100, 1) if tot_cost else None},
            # 특정 계정 적자모달은 전부 반환(상품번호 복사·삭제 누락 방지), 전체계정은 5000 상한.
            'rows': rows if eid else rows[:5000],
        })


class ElevenKeywordRoasView(views.APIView):
    """키워드별 ROAS(adoffice). eleven_id 없으면 전체 계정. ?date_from&date_to&roas_min&roas_max&export"""
    def get(self, request):
        eid = request.query_params.get('eleven_id')
        d_from, d_to, _f = _roas_period(request)
        rmin = _f(request.query_params.get('roas_min')); rmax = _f(request.query_params.get('roas_max'))
        eids = [eid] if eid else _active_eids()
        rows = []
        for e in eids:
            rows += _eleven_keyword_rows(e, d_from, d_to, rmin, rmax)
        rows.sort(key=lambda x: -x['cost'])

        if request.query_params.get('export'):
            import csv as _csv
            from django.http import HttpResponse
            resp = HttpResponse(content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = f'attachment; filename="keyword_roas_{eid or "all"}_{d_from}_{d_to}.csv"'
            resp.write('﻿')
            w = _csv.writer(resp)
            w.writerow(['아이디', '상품번호', '판매자코드', '키워드', '광고비', '전환매출', 'ROAS(%)', '클릭', '전환', '비고'])
            for x in rows:
                w.writerow([x['eleven_id'], x['product_no'], x['seller_code'], x['keyword'],
                            x['cost'], x['conv_amount'], x['roas_pct'], x['clicks'], x['conversions'],
                            '삭제됨' if x.get('deleted') else ''])
            return resp

        return Response({
            'eleven_id': eid, 'date_from': str(d_from), 'date_to': str(d_to),
            'count': len(rows), 'collected_at': None, 'rows': rows[:1000],
        })


class SellerAutoLoginView(views.APIView):
    """셀러 아이디 클릭 시 11번가 셀러오피스 자동 로그인 페이지 반환"""
    def get(self, request, seller_id):
        from django.http import HttpResponse
        acct = CrawlerAccount.objects.filter(login_id=seller_id, platform='11st').first()
        if not acct:
            return HttpResponse('계정을 찾을 수 없습니다', status=404)
        html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{acct.seller_name or seller_id} 로그인</title>
<style>
body{{display:flex;align-items:center;justify-content:center;height:100vh;font-family:-apple-system,sans-serif;background:#f5f5f5}}
.box{{background:#fff;padding:40px 50px;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.1);text-align:center}}
.name{{font-size:22px;font-weight:700;margin-bottom:8px}}
.id{{font-size:16px;color:#666;margin-bottom:20px}}
.msg{{font-size:15px;color:#e67700}}
</style></head>
<body><div class="box">
<div class="name">{acct.seller_name or seller_id}</div>
<div class="id">({seller_id})</div>
<div class="msg">셀러오피스 로그인 중...</div>
<form id="f" method="POST" action="https://login.11st.co.kr/auth/front/selleroffice/login.tmall">
<input type="hidden" name="loginName" value="{acct.login_id}">
<input type="hidden" name="passWord" value="{acct.password_enc}">
</form>
<script>setTimeout(function(){{document.getElementById("f").submit()}},500)</script>
</div></body></html>'''
        return HttpResponse(html, content_type='text/html')


from .models import CronSchedule

class CronScheduleViewSet(viewsets.ModelViewSet):
    queryset = CronSchedule.objects.all()
    serializer_class = None  # inline below

    def get_serializer_class(self):
        from rest_framework import serializers as sz
        class S(sz.ModelSerializer):
            class Meta:
                model = CronSchedule
                fields = '__all__'
        return S

class CronApplyView(views.APIView):
    """CronSchedule 테이블 기반으로 실제 crontab 적용"""
    def post(self, request):
        import subprocess
        schedules = CronSchedule.objects.all()
        lines = ['# Avengers 자동 수집 스케줄 (UI에서 관리)\n']
        for s in schedules:
            prefix = '' if s.is_active else '#'
            lines.append(f'{prefix}{s.cron_expr} {s.command}  # {s.display_name}\n')

        cron_content = ''.join(lines)
        proc = subprocess.run(['crontab', '-'], input=cron_content, capture_output=True, text=True)
        if proc.returncode != 0:
            return Response({'error': proc.stderr}, status=500)
        return Response({'applied': len(schedules), 'content': cron_content})

class AccountUnblockView(views.APIView):
    """차단된 계정 해제"""
    def post(self, request):
        from .models import CrawlerAccount
        account_id = request.data.get('id')
        if not account_id:
            return Response({'error': 'id 필요'}, status=400)
        try:
            acct = CrawlerAccount.objects.get(id=account_id)
            acct.crawling_status = '정상'
            acct.fail_count = 0
            acct.save()
            return Response({'unblocked': acct.login_id})
        except CrawlerAccount.DoesNotExist:
            return Response({'error': '계정 없음'}, status=404)


class BlockedAccountsView(views.APIView):
    def get(self, request):
        from .models import CrawlerAccount
        blocked = CrawlerAccount.objects.filter(
            platform='gmarket', crawling_status='차단됨'
        ).values('login_id', 'seller_name')
        return Response({
            'blocked': [
                {'seller_id': b['login_id'], 'seller_alias': b['seller_name'] or b['login_id']}
                for b in blocked
            ]
        })

    def post(self, request):
        from .models import CrawlerAccount
        seller_id = request.data.get('seller_id')
        if not seller_id:
            return Response({'error': 'seller_id 필요'}, status=400)
        try:
            acct = CrawlerAccount.objects.get(login_id=seller_id, platform='gmarket')
            acct.crawling_status = '정상'
            acct.fail_count = 0
            acct.save()
            return Response({'unblocked': seller_id})
        except CrawlerAccount.DoesNotExist:
            return Response({'error': '계정 없음'}, status=404)


class GmarketTimeseriesView(views.APIView):
    def get(self, request):
        from datetime import datetime, timedelta
        import pytz
        kst = pytz.timezone('Asia/Seoul')

        date_str = request.query_params.get('date')
        ids_str = request.query_params.get('ids', '')

        if not date_str:
            from django.utils import timezone as tz
            date_str = tz.localdate().isoformat()

        start = kst.localize(datetime.strptime(date_str, '%Y-%m-%d'))
        end = start + timedelta(days=1)

        ids = [i.strip() for i in ids_str.split(',') if i.strip()]

        qs = GmarketDepositSnapshot.objects.filter(
            collected_at__gte=start, collected_at__lt=end
        )
        if ids:
            qs = qs.filter(gmarket_id__in=ids)

        data = []
        for s in qs.order_by('collected_at'):
            data.append({
                'id': s.gmarket_id,
                'ts': s.collected_at.astimezone(kst).strftime('%H:%M'),
                'cpc': s.gmarket_cpc,
                'ai': s.ai_usage,
                'prime': 0,
            })

        sales = []
        if ids:
            try:
                from apps.sales.models import SalesRecord
                from apps.cpc.models import CrawlerAccount
                acct = CrawlerAccount.objects.filter(login_id__in=ids, platform='gmarket').first()
                if acct and acct.seller_name:
                    sales_qs = SalesRecord.objects.filter(
                        order_date=date_str,
                        seller_name=acct.seller_name,
                    ).values('created_at').annotate(
                        total=Sum('settlement_price')
                    ).order_by('created_at')
                    for sr in sales_qs:
                        sales.append({
                            'id': ids[0],
                            'ts': sr['created_at'].astimezone(kst).strftime('%H:%M') if hasattr(sr['created_at'], 'astimezone') else '',
                            'sales': sr['total'] or 0,
                        })
            except Exception:
                pass

        return Response({
            'date': date_str,
            'data': data,
            'sales': sales,
        })

from .models import ReceivedSmsMessage, SmsOutbox, SmsDeviceHeartbeat, SmsMessageImage
from datetime import datetime
from django.utils import timezone

def _publish_sms_event(sms_id):
    try:
        import redis as redis_client
        import json
        r = redis_client.Redis(host='localhost', port=6379, db=0)
        r.publish('sms:new', json.dumps({'last_id': sms_id}))
    except Exception:
        pass


def _forward_sms_to_telegram(sms):
    """수신 SMS를 텔레그램 봇으로 전송 (TelegramConfig 활성 시) — 모든 신규 문자 전달"""
    try:
        from .models import TelegramConfig, TelegramRecipient
        import requests as req
        config = TelegramConfig.objects.first()
        if not config or not config.bot_token:
            return
        if config.mode == 'off':
            return
        recipients = TelegramRecipient.objects.filter(is_active=True)
        if not recipients.exists():
            return

        sender = sms.checkphone_number or '알 수 없음'
        receiver = sms.csphone_number or '내 번호'
        message_text = sms.message or ''
        msg_type = getattr(sms, 'msg_type', 'SMS')

        text = (
            f'<b>📩 새 문자 수신 ({msg_type})</b>\n'
            f'━━━━━━━━━━━━━━━\n'
            f'📱 <b>받는 번호</b>: <code>{receiver}</code>\n'
            f'☎️ <b>보낸 번호</b>: <code>{sender}</code>\n'
            f'🕒 <b>시각</b>: {sms.received_at.strftime("%m/%d %H:%M:%S") if sms.received_at else "-"}\n'
            f'━━━━━━━━━━━━━━━\n'
            f'<pre>{message_text[:1000]}</pre>'
        )

        for r in recipients:
            try:
                req.post(
                    f'https://api.telegram.org/bot{config.bot_token}/sendMessage',
                    json={'chat_id': r.chat_id, 'text': text, 'parse_mode': 'HTML', 'disable_web_page_preview': True},
                    timeout=5,
                )
            except Exception:
                pass
    except Exception:
        pass

def _ms_to_datetime(ms):
    try:
        return datetime.fromtimestamp(int(ms) / 1000.0, tz=timezone.get_current_timezone())
    except Exception:
        return None

# 실제 문자가 아닌 노이즈 알림(페어링 기기 안내 등) — 수신해도 저장·전달 안 함
SMS_NOISE_PHRASES = (
    '페어링한 기기에 메시지가 표시됩니다',
)


class SmsReceiveView(views.APIView):
    """SMS 수신 API - smsApp(신규) + 기존 외부 연동 모두 지원"""
    permission_classes = []  # 인증 불필요 (외부 연동)

    def post(self, request):
        # 신규 포맷 (smsApp): csphone_number / checkphone_number / message / receive_time
        # 구포맷: phone / message / csphone
        d = request.data
        csphone = d.get('csphone_number') or d.get('csphone', '')
        check = d.get('checkphone_number') or d.get('phone', '')
        message = d.get('message', '')
        receive_time_ms = d.get('receive_time')

        if not message:
            return Response({'error': 'message 필요'}, status=400)

        # 페어링 기기 알림 등 노이즈(실제 문자 아님) 무시 — 저장·텔레그램 전달 안 함
        if any(s in message for s in SMS_NOISE_PHRASES):
            return Response({'id': None, 'received': False, 'ignored': 'noise'}, status=200)

        msg_bytes = len(message.encode('utf-8'))
        msg_type = 'LMS' if msg_bytes > 80 else 'SMS'

        # 중복 차단: 같은 수신번호+동일 내용이 최근 90초 내 이미 들어왔으면 스킵(저장·텔레그램 모두).
        # 한 문자가 SMS/LMS/MMS·알림 등 여러 채널로 중복 수신돼 텔레그램이 2~3번 가던 문제 방지.
        from datetime import timedelta as _td
        from django.utils import timezone as _tz
        dup = (ReceivedSmsMessage.objects
               .filter(checkphone_number=check, message=message,
                       received_at__gte=_tz.now() - _td(seconds=90))
               .order_by('-id').first())
        if dup:
            return Response({'id': dup.id, 'received': True, 'duplicate': True}, status=200)

        sms = ReceivedSmsMessage.objects.create(
            csphone_number=csphone,
            checkphone_number=check,
            message=message,
            msg_type=msg_type,
            receive_time=_ms_to_datetime(receive_time_ms) if receive_time_ms else None,
        )

        _publish_sms_event(sms.id)
        _forward_sms_to_telegram(sms)
        return Response({'id': sms.id, 'received': True}, status=201)

    def get(self, request):
        """최근 SMS 조회"""
        msgs = ReceivedSmsMessage.objects.all()[:50]
        data = [{'id': m.id, 'phone': m.checkphone_number, 'message': m.message,
                 'received_at': m.received_at.isoformat()} for m in msgs]
        return Response(data)

class SmsOtpTestView(views.APIView):
    """SMS OTP 테스트 - Redis에서 OTP 수신 대기"""
    def post(self, request):
        import redis as redis_client
        import json, re, time

        timeout = int(request.data.get('timeout', 30))
        r = redis_client.Redis(host='localhost', port=6379, db=0, decode_responses=True)
        ps = r.pubsub()
        ps.subscribe('sms:new')

        start = time.time()
        while time.time() - start < timeout:
            msg = ps.get_message(timeout=1)
            if msg and msg['type'] == 'message':
                try:
                    payload = json.loads(msg['data'])
                    last_id = payload.get('last_id')
                    if last_id:
                        sms = ReceivedSmsMessage.objects.filter(id=last_id).first()
                        if sms:
                            match = re.search(r'\[(\d{6})\]', sms.message)
                            if match:
                                ps.unsubscribe()
                                return Response({'code': match.group(1), 'message': sms.message})
                except Exception:
                    pass
        ps.unsubscribe()
        return Response({'code': None, 'timeout': True})

from .models import SmsPhoneSetting

class SmsPhoneSettingView(views.APIView):
    permission_classes = []

    def get(self, request):
        phones = SmsPhoneSetting.objects.all()
        data = [{'id': p.id, 'phone_number': p.phone_number, 'name': p.name, 'is_active': p.is_active} for p in phones]
        return Response(data)

    def post(self, request):
        phone = request.data.get('phone_number', '')
        name = request.data.get('name', '')
        if not phone:
            return Response({'error': '전화번호 필요'}, status=400)
        obj, created = SmsPhoneSetting.objects.get_or_create(
            phone_number=phone, defaults={'name': name}
        )
        if not created:
            obj.name = name
            obj.save()
        return Response({'id': obj.id, 'created': created})

    def delete(self, request):
        pid = request.data.get('id')
        SmsPhoneSetting.objects.filter(id=pid).delete()
        return Response({'deleted': True})

class SmsLatestView(views.APIView):
    """최근 SMS 목록 (위젯용, 인증 불필요)"""
    permission_classes = []

    def get(self, request):
        limit = int(request.query_params.get('limit', 30))
        since_id = request.query_params.get('since_id')
        qs = ReceivedSmsMessage.objects.all()
        if since_id:
            qs = qs.filter(id__gt=since_id)
        msgs = qs[:limit]
        data = [{
            'id': m.id, 'csphone': m.csphone_number, 'phone': m.checkphone_number,
            'message': m.message, 'received_at': m.received_at.isoformat()
        } for m in msgs]
        return Response(data)


# ==========================
# smsApp 게이트웨이 신규 API
# ==========================

class MmsReceiveView(views.APIView):
    """MMS 수신 API - smsApp이 multipart/form-data로 이미지 포함 전송"""
    permission_classes = []

    def post(self, request):
        import os, uuid
        from django.conf import settings as dj_settings

        d = request.data
        csphone = d.get('csphone_number', '')
        check = d.get('checkphone_number', '')
        message = d.get('message', '')
        receive_time_ms = d.get('receive_time')

        # 페어링 기기 알림 등 노이즈 무시(이미지 없는 텍스트 한정 — 이미지 MMS는 고유 콘텐츠라 유지)
        if message and not request.FILES.getlist('images') and any(s in message for s in SMS_NOISE_PHRASES):
            return Response({'id': None, 'received': False, 'ignored': 'noise'}, status=200)

        # 중복 차단: 내용 동일+이미지 없는 텍스트가 최근 90초내 이미 들어왔으면 스킵(SMS/LMS와 중복 발송 방지).
        # 이미지가 있는 MMS는 고유 콘텐츠라 절대 스킵하지 않음.
        if message and not request.FILES.getlist('images'):
            from datetime import timedelta as _td
            from django.utils import timezone as _tz
            dup = (ReceivedSmsMessage.objects
                   .filter(checkphone_number=check, message=message,
                           received_at__gte=_tz.now() - _td(seconds=90))
                   .order_by('-id').first())
            if dup:
                return Response({'id': dup.id, 'received': True, 'duplicate': True}, status=200)

        sms = ReceivedSmsMessage.objects.create(
            csphone_number=csphone,
            checkphone_number=check,
            message=message or '',
            msg_type='MMS',
            receive_time=_ms_to_datetime(receive_time_ms) if receive_time_ms else None,
        )

        # 이미지 저장: media/sms/received/YYYY-MM-DD/{uuid}.{ext}
        files = request.FILES.getlist('images')
        date_dir = timezone.localdate().strftime('%Y-%m-%d')
        save_root = os.path.join(dj_settings.MEDIA_ROOT, 'sms', 'received', date_dir)
        os.makedirs(save_root, exist_ok=True)

        for f in files:
            ext = (f.name.rsplit('.', 1)[-1] if '.' in f.name else 'jpg').lower()
            if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                ext = 'jpg'
            fname = f"{uuid.uuid4().hex}.{ext}"
            fpath = os.path.join(save_root, fname)
            with open(fpath, 'wb') as out:
                for chunk in f.chunks():
                    out.write(chunk)
            rel = os.path.relpath(fpath, dj_settings.MEDIA_ROOT)
            SmsMessageImage.objects.create(
                message=sms,
                filename=f.name,
                filepath=rel,
                content_type=f.content_type or '',
                size=f.size or 0,
            )

        _publish_sms_event(sms.id)
        _forward_sms_to_telegram(sms)
        return Response({'id': sms.id, 'received': True, 'images': len(files)}, status=201)


class SmsOutboxView(views.APIView):
    """발송 대기열 조회 (smsApp이 5초마다 폴링) - JSON 배열 직접 반환"""
    permission_classes = []

    def get(self, request):
        items = SmsOutbox.objects.filter(status='pending').order_by('created_at')[:50]
        data = [{
            'id': o.id,
            'phone_number': o.phone_number,
            'message': o.message,
            'sender_phone': o.sender_phone,
            'template_id': o.template_id,
            'status': o.status,
            'error_message': o.error_message or None,
            'created_at': o.created_at.isoformat() if o.created_at else None,
            'sent_at': o.sent_at.isoformat() if o.sent_at else None,
        } for o in items]
        return Response(data)

    def post(self, request):
        """관리자/내부에서 발송 요청 추가"""
        d = request.data
        phone = (d.get('phone_number') or '').strip()
        message = (d.get('message') or '').strip()
        if not phone or not message:
            return Response({'error': 'phone_number, message 필수'}, status=400)
        obj = SmsOutbox.objects.create(
            phone_number=phone,
            message=message,
            sender_phone=d.get('sender_phone', '') or '',
            template_id=d.get('template_id'),
        )
        return Response({'id': obj.id, 'status': obj.status}, status=201)


class SmsOutboxResultView(views.APIView):
    """발송 결과 보고 (smsApp → 서버)"""
    permission_classes = []

    def post(self, request, pk):
        try:
            obj = SmsOutbox.objects.get(pk=pk)
        except SmsOutbox.DoesNotExist:
            return Response({'error': 'not found'}, status=404)

        d = request.data
        status_val = d.get('status', 'sent')
        obj.status = status_val if status_val in ('sent', 'failed') else 'failed'
        obj.error_message = (d.get('error_message') or '')[:500]
        sent_at = d.get('sent_at')
        if sent_at:
            try:
                obj.sent_at = datetime.fromisoformat(sent_at)
            except Exception:
                obj.sent_at = timezone.now()
        else:
            obj.sent_at = timezone.now()
        obj.save()
        return Response({'id': obj.id, 'status': obj.status})


class SmsHeartbeatView(views.APIView):
    """smsApp 디바이스 heartbeat (30초 주기)"""
    permission_classes = []

    def post(self, request):
        d = request.data
        phone = (d.get('phone_number') or '').strip()
        version = d.get('app_version') or ''
        if not phone:
            return Response({'ok': False, 'connected': False, 'error': 'phone_number 필수'}, status=400)

        SmsDeviceHeartbeat.objects.update_or_create(
            phone_number=phone,
            defaults={'app_version': version},
        )
        # smsApp 핸드폰을 자동으로 SmsPhoneSetting에도 등록
        SmsPhoneSetting.objects.get_or_create(
            phone_number=phone,
            defaults={'name': 'smsApp', 'is_active': True},
        )
        return Response({
            'ok': True,
            'connected': True,
            'server_time': timezone.now().isoformat(),
        })


class SmsChangeNumberView(views.APIView):
    """smsApp 폰번호 변경 (고스트 레코드 방지)"""
    permission_classes = []

    def post(self, request):
        d = request.data
        old_phone = (d.get('old_phone') or '').strip()
        new_phone = (d.get('new_phone') or '').strip()
        if not new_phone:
            return Response({'ok': False, 'message': 'new_phone 필수'}, status=400)

        action = 'noop'

        if old_phone and old_phone != new_phone:
            old_hb = SmsDeviceHeartbeat.objects.filter(phone_number=old_phone).first()
            new_hb = SmsDeviceHeartbeat.objects.filter(phone_number=new_phone).first()
            if old_hb and not new_hb:
                old_hb.phone_number = new_phone
                old_hb.save()
                action = 'renamed'
            elif old_hb and new_hb:
                old_hb.delete()
                action = 'merged_to_new'
            else:
                SmsDeviceHeartbeat.objects.update_or_create(
                    phone_number=new_phone, defaults={}
                )
                action = 'created'

            # SmsPhoneSetting도 같이 정리
            old_ps = SmsPhoneSetting.objects.filter(phone_number=old_phone).first()
            new_ps = SmsPhoneSetting.objects.filter(phone_number=new_phone).first()
            if old_ps and not new_ps:
                old_ps.phone_number = new_phone
                old_ps.save()
            elif old_ps and new_ps:
                old_ps.delete()
            elif not new_ps:
                SmsPhoneSetting.objects.create(phone_number=new_phone, name='smsApp')
        else:
            SmsDeviceHeartbeat.objects.update_or_create(phone_number=new_phone, defaults={})
            SmsPhoneSetting.objects.get_or_create(
                phone_number=new_phone,
                defaults={'name': 'smsApp', 'is_active': True},
            )
            action = 'activated'

        return Response({'ok': True, 'action': action, 'message': '변경 완료'})


class CrawlerAccountExcelUploadView(views.APIView):
    """크롤러 계정 엑셀 일괄 업로드"""
    permission_classes = []

    def post(self, request):
        import openpyxl
        f = request.FILES.get('file')
        if not f:
            return Response({'error': '파일 필요'}, status=400)

        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            return Response({'error': f'파일 읽기 실패: {e}'}, status=400)

        if len(rows) < 2:
            return Response({'error': '데이터 행 없음 (헤더 + 최소 1행 필요)'}, status=400)

        headers = [str(h or '').strip().lower() for h in rows[0]]
        created, updated, errors = 0, 0, []

        for i, row in enumerate(rows[1:], start=2):
            d = dict(zip(headers, row))
            platform = str(d.get('platform') or d.get('플랫폼') or '').strip()
            login_id = str(d.get('login_id') or d.get('아이디') or '').strip()
            password = str(d.get('password') or d.get('비밀번호') or '').strip()

            if not platform or not login_id:
                errors.append(f'{i}행: platform/login_id 누락')
                continue

            obj, is_new = CrawlerAccount.objects.get_or_create(
                platform=platform, login_id=login_id,
                defaults={
                    'password_enc': password,
                    'seller_name': str(d.get('seller_name') or d.get('셀러명') or '').strip(),
                    'cost_type': str(d.get('cost_type') or d.get('타입') or 'sellerpoint').strip(),
                    'is_active': True,
                }
            )
            if is_new:
                created += 1
            else:
                if password:
                    obj.password_enc = password
                sn = str(d.get('seller_name') or d.get('셀러명') or '').strip()
                if sn:
                    obj.seller_name = sn
                ct = str(d.get('cost_type') or d.get('타입') or '').strip()
                if ct:
                    obj.cost_type = ct
                obj.save()
                updated += 1

        return Response({'created': created, 'updated': updated, 'errors': errors})


class CrawlerAccountExcelSampleView(views.APIView):
    """크롤러 계정 엑셀 샘플 다운로드"""
    permission_classes = []

    def get(self, request):
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '크롤러 계정'
        headers = ['platform', 'login_id', 'password', 'seller_name', 'cost_type']
        ws.append(headers)
        ws.append(['gmarket', 'example_id', 'password123', '셀러이름', 'sellerpoint'])
        ws.append(['11st', 'example_11st', 'pass456', '11번가셀러', 'sellercash'])

        # 컬럼 너비
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=crawler_accounts_sample.xlsx'
        wb.save(response)
        return response


class CrawlerAccountStatsView(views.APIView):
    """크롤러 계정 셋팅 현황 (대시보드용)"""
    permission_classes = []

    def get(self, request):
        from django.db.models import Q, Count
        from .models import CrawlerAccount
        result = []
        for platform, label in [('gmarket', '지마켓'), ('11st', '11번가')]:
            qs = CrawlerAccount.objects.filter(platform=platform)
            total = qs.count()
            active = qs.filter(is_active=True).count()
            with_pw = qs.exclude(Q(password_enc='') | Q(password_enc__isnull=True)).count()
            no_pw = total - with_pw
            blocked = qs.filter(fail_count__gte=30).count()
            warn = qs.filter(fail_count__gte=10, fail_count__lt=30).count()
            with_cookie = qs.exclude(Q(cookie_data='') | Q(cookie_data__isnull=True)).count()
            ready_pct = round(100.0 * with_pw / total, 1) if total else 0
            result.append({
                'platform': platform,
                'label': label,
                'total': total,
                'active': active,
                'inactive': total - active,
                'with_password': with_pw,
                'without_password': no_pw,
                'with_cookie': with_cookie,
                'blocked': blocked,
                'warn': warn,
                'ready_pct': ready_pct,
                'all_set': (no_pw == 0 and total > 0),
            })
        return Response(result)


class SmsOutboxHistoryView(views.APIView):
    """SMS 발송 전체 이력 (대시보드용)"""
    permission_classes = []

    def get(self, request):
        limit = int(request.query_params.get('limit', 100))
        status_filter = request.query_params.get('status')
        qs = SmsOutbox.objects.all()
        if status_filter:
            qs = qs.filter(status=status_filter)
        items = qs[:limit]
        data = [{
            'id': o.id,
            'phone_number': o.phone_number,
            'message': o.message,
            'sender_phone': o.sender_phone,
            'status': o.status,
            'error_message': o.error_message or None,
            'created_at': o.created_at.isoformat() if o.created_at else None,
            'sent_at': o.sent_at.isoformat() if o.sent_at else None,
        } for o in items]
        return Response(data)


class SmsDeviceListView(views.APIView):
    """SMS 디바이스(smsApp 설치된 폰) 목록 + 온라인 상태"""
    permission_classes = []

    def get(self, request):
        from datetime import timedelta
        threshold = timezone.now() - timedelta(seconds=90)
        devices = SmsDeviceHeartbeat.objects.all()
        data = []
        for d in devices:
            data.append({
                'id': d.id,
                'phone_number': d.phone_number,
                'app_version': d.app_version,
                'last_seen_at': d.last_seen_at.isoformat() if d.last_seen_at else None,
                'is_online': d.last_seen_at >= threshold if d.last_seen_at else False,
            })
        return Response(data)


class SmsSettingsListView(views.APIView):
    """smsApp /api/settings/ - 등록된 폰 설정 목록 (위젯의 SmsPhoneSetting 기반)"""
    permission_classes = []

    def get(self, request):
        phones = SmsPhoneSetting.objects.filter(is_active=True)
        data = []
        for idx, p in enumerate(phones, start=1):
            data.append({
                'id': p.id,
                'csphone_number': p.phone_number,
                'checkphone_number': '',
                'alias': p.name or p.phone_number,
                'is_save_to_db': 1,
                'is_notify_pc': 1,
                'is_notify_telegram': 0,
                'created_at': p.created_at.isoformat() if p.created_at else '',
                'is_admin': 1,
                'payment_date': '',
            })
        return Response(data)


from rest_framework.permissions import IsAuthenticated
from . import eleven_my_product_service as _emp_svc


class ElevenMyProductSyncView(views.APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        account_id = request.data.get('account_id')
        try:
            if account_id:
                result = _emp_svc.sync_products_for_account(int(account_id))
            else:
                result = _emp_svc.sync_focused_accounts()
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ElevenMyProductListView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        account_id = request.query_params.get('account_id')
        status_q = request.query_params.get('status') or None
        search = request.query_params.get('search') or None
        focused_only = request.query_params.get('focused_only') in ('1', 'true', 'True')
        sort = request.query_params.get('sort') or None
        order = request.query_params.get('order') or 'asc'
        needs_check = request.query_params.get('needs_check') in ('1', 'true', 'True')

        # 전체 다운로드(export=1): 현재 필터에 맞는 '전체' 상품을 CSV로 스트리밍 (페이지 무관, 선택 무관)
        if request.query_params.get('export'):
            import csv as _csv
            from django.http import StreamingHttpResponse
            from django.db.models import Q
            from apps.cpc.models import ElevenMyProduct
            qs = ElevenMyProduct.objects.select_related('account')
            if account_id:
                qs = qs.filter(account_id=int(account_id))
            if status_q:
                qs = qs.filter(status_type=status_q)
            if search:
                qs = qs.filter(Q(product_name__icontains=search) | Q(seller_product_code__icontains=search)
                               | Q(account__login_id__icontains=search) | Q(account__seller_name__icontains=search))
            if focused_only:
                qs = qs.filter(account__is_focused=True)
            qs = qs.order_by('account_id', 'product_no')

            class _Echo:
                def write(self, value):
                    return value

            header = ['셀러', '로그인ID', '상품번호', '판매자상품코드', '상품명', '판매가', '재고', '판매상태', '카테고리']

            def _rows():
                w = _csv.writer(_Echo())
                yield '﻿' + w.writerow(header)
                for p in qs.iterator(chunk_size=2000):
                    yield w.writerow([
                        p.account.seller_name, p.account.login_id, p.product_no,
                        p.seller_product_code, p.product_name, p.sale_price,
                        p.stock_quantity, p.status_type, p.category_id,
                    ])

            resp = StreamingHttpResponse(_rows(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = 'attachment; filename="eleven_my_products.csv"'
            return resp

        result = _emp_svc.get_my_products(
            account_id=int(account_id) if account_id else None,
            page=page, per_page=per_page,
            status=status_q, search=search,
            focused_only=focused_only,
            sort=sort, order=order, needs_check=needs_check,
        )
        return Response(result)


# ── 지마켓/옥션(ESM) 나의 상품 — 11번가와 완전 분리된 독립 API ──
_GMKT_SORT = {
    'product_no': 'product_no', 'product_name': 'product_name', 'sale_price': 'sale_price',
    'stock_quantity': 'stock_quantity', 'status_type': 'status_type', 'market': 'market',
    'seller_product_code': 'seller_product_code', 'synced_at': 'synced_at',
    'login_id': 'account__login_id', 'seller_name': 'account__seller_name',
}


class GmarketMyProductListView(views.APIView):
    """지마켓/옥션 나의 상품 목록 — 쇼핑몰(market)·아이디(account)·상태·검색 필터."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import math
        from django.db.models import Q
        from apps.cpc.models import GmarketMyProduct
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        account_id = request.query_params.get('account_id')
        login_id = request.query_params.get('login_id')
        market = request.query_params.get('market') or None       # gmarket | auction
        status_q = request.query_params.get('status') or None
        search = request.query_params.get('search') or None
        sort = request.query_params.get('sort') or 'product_no'
        order = request.query_params.get('order') or 'asc'

        qs = GmarketMyProduct.objects.select_related('account')
        if account_id:
            qs = qs.filter(account_id=int(account_id))
        if login_id:
            qs = qs.filter(account__login_id=login_id)
        if market:
            qs = qs.filter(market=market)
        if status_q:
            qs = qs.filter(status_type=status_q)
        if search:
            qs = qs.filter(Q(product_name__icontains=search) | Q(product_no__icontains=search)
                           | Q(seller_product_code__icontains=search) | Q(account__login_id__icontains=search))
        # 중복제외: 같은 (계정, 판매자코드)는 1개만(가장 빠른 id) — 같은 상품의 다중 상품번호/마켓 중복 제거.
        # 기존 `id__in=<keep_ids 서브쿼리>`는 48만행 semi-join이 매 페이지 재실행돼 ~353초였음.
        # 제거 대상(loser=그룹 내 min 초과분)은 ~3.9만개뿐 → loser id만 캐시하고 exclude.
        dedup_on = bool(request.query_params.get('dedup'))
        if dedup_on:
            from django.db.models import Min
            from django.core.cache import cache as _cache
            sig = f"gmkt_dedup_losers:{account_id}:{login_id}:{market}:{status_q}:{search}"
            loser_ids = _cache.get(sig)
            if loser_ids is None:
                base = qs.exclude(seller_product_code='')
                keep = set(base.values('account_id', 'seller_product_code')
                               .annotate(mid=Min('id')).values_list('mid', flat=True))
                loser_ids = [i for i in base.values_list('id', flat=True) if i not in keep]
                _cache.set(sig, loser_ids, 180)
            if loser_ids:
                qs = qs.exclude(id__in=loser_ids)
        sf = _GMKT_SORT.get(sort, 'product_no')
        qs = qs.order_by('-' + sf if order == 'desc' else sf)

        if request.query_params.get('export'):
            import csv as _csv
            from django.http import StreamingHttpResponse

            class _Echo:
                def write(self, v):
                    return v
            header = ['쇼핑몰', '로그인ID', '상품번호', '상품상태', '판매자관리코드', '상품명', '판매가', '재고', '카테고리']
            mk = {'gmarket': '지마켓', 'auction': '옥션'}

            def _rows():
                w = _csv.writer(_Echo())
                yield '﻿' + w.writerow(header)
                for p in qs.iterator(chunk_size=2000):
                    yield w.writerow([mk.get(p.market, p.market), p.account.login_id, p.product_no,
                                      p.status_type, p.seller_product_code, p.product_name,
                                      p.sale_price, p.stock_quantity, p.category_code])
            resp = StreamingHttpResponse(_rows(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = 'attachment; filename="gmarket_my_products.csv"'
            return resp

        # COUNT(*)는 48만행이라 페이지 이동마다 ~2초 → 필터별 캐시(동기화 때만 변함)
        from django.core.cache import cache as _cache
        cnt_key = f"gmkt_my_count:{account_id}:{login_id}:{market}:{status_q}:{search}:{int(dedup_on)}"
        total = _cache.get(cnt_key)
        if total is None:
            total = qs.count()
            _cache.set(cnt_key, total, 180)
        start = (page - 1) * per_page
        items = [{
            'id': p.id, 'login_id': p.account.login_id, 'seller_name': p.account.seller_name,
            'market': p.market, 'product_no': p.product_no, 'product_name': p.product_name,
            'sale_price': p.sale_price, 'stock_quantity': p.stock_quantity,
            'status_type': p.status_type, 'seller_product_code': p.seller_product_code,
            'category_code': p.category_code,
            'synced_at': p.synced_at.isoformat() if p.synced_at else None,
        } for p in qs[start:start + per_page]]
        return Response({'items': items, 'total': total, 'page': page, 'per_page': per_page,
                         'total_pages': math.ceil(total / per_page) if per_page else 1})


class GmarketCrawlStatusView(views.APIView):
    """지마켓 상품별광고비(ad_report) 수집 상태 — 오늘 갱신/미갱신 계정 + 최근 에러(원인)."""
    def get(self, request):
        from apps.cpc.models import GmarketProductAdCost as G, CrawlerAccount, CrawlerLog
        from django.db.models import Max
        from datetime import timedelta
        import subprocess
        today = timezone.localdate()
        kst = timezone.get_current_timezone()
        masters = [a.login_id for a in CrawlerAccount.objects.filter(platform='gmarket', is_active=True)
                   if not (a.gmarket_origin_id and a.gmarket_origin_id != a.login_id)]
        last = {r['login_id']: r['m'] for r in
                G.objects.filter(login_id__in=masters).values('login_id').annotate(m=Max('collected_at'))}
        done, failed = [], []
        for m in masters:
            lm = last.get(m)
            if lm and lm.astimezone(kst).date() == today:
                done.append(m)
            else:
                failed.append({'login_id': m,
                               'last': lm.astimezone(kst).strftime('%m-%d %H:%M') if lm else '수집기록 없음'})
        # 최근 12h 에러(원인 요약)
        since = timezone.now() - timedelta(hours=12)
        errs = [{'account': l.account_id, 'msg': (l.message or '')[:90],
                 'at': l.created_at.astimezone(kst).strftime('%H:%M')}
                for l in CrawlerLog.objects.filter(platform='gmarket', level='error', created_at__gte=since)
                .order_by('-created_at')[:15]]
        running = False
        try:
            running = 'crawl_gmarket_ad_report' in subprocess.run(
                ['ps', '-eo', 'args'], capture_output=True, text=True).stdout
        except Exception:
            pass
        return Response({'total': len(masters), 'done': len(done),
                         'failed': failed, 'errors': errs, 'running': running})


class GmarketRecrawlView(views.APIView):
    """실패(미갱신) 계정만 상품별광고비 재크롤 (백그라운드). accounts=login_id 리스트."""
    def post(self, request):
        import threading as th
        accounts = request.data.get('accounts') or []
        with_keywords = bool(request.data.get('with_keywords', False))
        if not accounts:
            return Response({'error': '재크롤할 계정이 없습니다.'}, status=400)

        def run():
            from crawlers.gmarket_ad_report_crawler import run as adrun
            adrun(login_ids=accounts, with_keywords=with_keywords)
        th.Thread(target=run, daemon=True).start()
        return Response({'status': 'started', 'count': len(accounts), 'accounts': accounts})


class GmarketDashboardView(views.APIView):
    """지마켓 대시보드 요약 — 계정별 잔액/광고비/상품수 (기간 광고비 집계)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import re as _re
        from datetime import datetime, timedelta
        from django.db.models import Count, Sum, Max
        from apps.cpc.models import (CrawlerAccount, GmarketDepositSnapshot,
                                     GmarketCostHistory, GmarketMyProduct)
        df = request.query_params.get('date_from')
        dt = request.query_params.get('date_to')
        d1 = datetime.strptime(dt, '%Y-%m-%d').date() if _re.match(r'^\d{4}-\d{2}-\d{2}$', dt or '') else timezone.localdate()
        d0 = datetime.strptime(df, '%Y-%m-%d').date() if _re.match(r'^\d{4}-\d{2}-\d{2}$', df or '') else (d1 - timedelta(days=30))

        market = request.query_params.get('market') or 'gmarket'
        if market not in ('gmarket', 'auction'):
            market = 'gmarket'
        accts = list(CrawlerAccount.objects.filter(platform='gmarket', is_active=True)
                     .order_by('display_order', 'login_id'))
        # 옥션 뷰: 옥션에 없는 공유ESM 중복 서브아이디만 제외(나머지 계정은 유지).
        # 이 계정들은 옥션 상품이 복제돼 있을 뿐 옥션 거래·매출이 없음. 지마켓 계정 레코드는 보존(삭제 아님).
        if market == 'auction':
            # 옥션 = 지마켓 계정과 동일하되, 공유ESM 중복 서브아이디 5개만 제외하고
            # 1번(rejoice666) 자리를 rejoice7942로 교체(옥션 매출이 rejoice7942로 기록됨).
            _NO_AUCTION = {'rejoice223', 'rejoice224', 'rejoice235', 'rejoice236', 'starvisi'}
            _REPLACE = {'rejoice666': 'rejoice7942'}
            _ord = {a.login_id: a.display_order for a in accts}
            accts = [a for a in accts
                     if a.login_id not in _NO_AUCTION and a.login_id not in _REPLACE]
            for _old, _new in _REPLACE.items():
                _ex = CrawlerAccount.objects.filter(login_id=_new).first()
                if _ex and _ex.login_id not in {a.login_id for a in accts}:
                    _ex.display_order = _ord.get(_old, _ex.display_order)
                    accts.append(_ex)
            accts.sort(key=lambda a: (a.display_order, a.login_id))
        # 계정별 최신 잔액 스냅샷
        bal = {}
        for s in (GmarketDepositSnapshot.objects.values('gmarket_id')
                  .annotate(m=Max('collected_at'))):
            last = (GmarketDepositSnapshot.objects.filter(gmarket_id=s['gmarket_id'], collected_at=s['m'])
                    .values('total_balance', 'total_usage', 'gmarket_cpc', 'ai_usage',
                            'auction_cpc', 'collected_at').first())
            if last:
                bal[s['gmarket_id']] = last
        # 계정별 기간 CPC/AI — 광고센터 스냅샷은 '당일 소진액'(매일 리셋)이므로
        # 기간 내 각 날짜의 '마지막 스냅샷'을 골라 일별로 합산해야 기간 총액이 됨.
        from collections import defaultdict
        import pytz as _pytz
        _kst = _pytz.timezone('Asia/Seoul')
        _start = _kst.localize(datetime.combine(d0, datetime.min.time()))
        _end = _kst.localize(datetime.combine(d1, datetime.min.time()) + timedelta(days=1))
        _daily = {}  # (gid, KST일자) -> (collected_at, cpc, ai, auction)  그날 마지막값
        for s in (GmarketDepositSnapshot.objects
                  .filter(collected_at__gte=_start, collected_at__lt=_end)
                  .values('gmarket_id', 'collected_at', 'gmarket_cpc', 'ai_usage', 'auction_cpc')):
            ca = s['collected_at']
            key = (s['gmarket_id'], ca.astimezone(_kst).date())
            prev = _daily.get(key)
            if prev is None or ca > prev[0]:
                _daily[key] = (ca, s['gmarket_cpc'] or 0, s['ai_usage'] or 0, s['auction_cpc'] or 0)
        # 거래내역(GmarketCostHistory)은 1~2일 지연 기록 → 최신 반영일 이후(미반영일)는
        # 광고센터 당일소진액 스냅샷(계정별 그날 마지막값)으로 실시간 보충. 반영일까지는 거래내역만 써 중복 방지.
        # ★ 차단선은 반드시 '계정별'로 — 전역이면 한 계정이 최신 거래내역을 받는 순간
        #   아직 못 받은 다른 계정의 미수집일까지 보충이 꺼져 광고비가 갑자기 0으로 떨어짐(부분수집/백필 중 사고).
        _def_cut = d0 - timedelta(days=1)
        _cut_g_by = defaultdict(lambda: _def_cut)
        for r in (GmarketCostHistory.objects.filter(market='gmarket', use_date__lte=d1)
                  .values('seller_id').annotate(m=Max('use_date'))):
            _cut_g_by[r['seller_id']] = r['m']
        _cut_a_by = defaultdict(lambda: _def_cut)
        for r in (GmarketCostHistory.objects.filter(market='auction', use_date__lte=d1)
                  .values('seller_id').annotate(m=Max('use_date'))):
            _cut_a_by[r['seller_id']] = r['m']
        snap = defaultdict(lambda: {'cpc': 0, 'ai': 0, 'auction': 0, 'days': 0})
        for (_gid, _d), (_ca, _cpc, _ai, _au) in _daily.items():
            v = snap[_gid]
            if _d > _cut_g_by[_gid]:
                v['cpc'] += _cpc; v['ai'] += _ai; v['days'] += 1
            if _d > _cut_a_by[_gid]:
                v['auction'] += _au
        # 광고비 — 판매예치금 거래내역(GmarketCostHistory, market 태그)이 정확한 출처.
        # (광고센터 스냅샷은 '당일 소진액'이라 기간 누적이 비어 0으로 나왔음 → 거래내역으로 대체)
        # 지마켓(market='gmarket') CPC/AI/서버 + 옥션(market='auction') 별도 집계.
        # market 파라미터로 지마켓/옥션 전환(위에서 이미 해석). 'auction'키 = 반대(他) 마켓 합계.
        other = 'auction' if market == 'gmarket' else 'gmarket'
        TYPE_KEY = {'CPC': 'cpc', 'AI매출업': 'ai', '서버비용': 'server'}
        cost = defaultdict(lambda: {'cpc': 0, 'ai': 0, 'server': 0, 'auction': 0, 'cnt': 0})
        for r in (GmarketCostHistory.objects
                  .filter(market=market, transaction_type__in=list(TYPE_KEY.keys()),
                          use_date__gte=d0, use_date__lte=d1)
                  .exclude(comment__icontains='판매예치금')   # 판매예치금 송금 등 비광고 차감 제외
                  .values('seller_id', 'transaction_type').annotate(spend=Sum('amount'), cnt=Count('id'))):
            c = cost[r['seller_id']]
            c[TYPE_KEY[r['transaction_type']]] += abs(r['spend'] or 0)
            c['cnt'] += r['cnt']
        for r in (GmarketCostHistory.objects
                  .filter(market=other, transaction_type__in=list(TYPE_KEY.keys()),
                          use_date__gte=d0, use_date__lte=d1)
                  .exclude(comment__icontains='판매예치금')   # 판매예치금 송금 등 비광고 차감 제외
                  .values('seller_id').annotate(spend=Sum('amount'), cnt=Count('id'))):
            c = cost[r['seller_id']]
            c['auction'] += abs(r['spend'] or 0)   # 他 마켓 광고비 합계
            c['cnt'] += r['cnt']
        # 계정별 상품수
        prod = {r['account__login_id']: r['n'] for r in (
            GmarketMyProduct.objects.values('account__login_id').annotate(n=Count('id')))}
        prod_mkt = {(r['account__login_id'], r['market']): r['n'] for r in (
            GmarketMyProduct.objects.values('account__login_id', 'market').annotate(n=Count('id')))}

        # 계정별 매출/순수익 — SalesRecord(지마켓+옥션 플랫폼, 셀러 login_id, 기간)
        # 공유ESM 서브아이디(활성계정 아님)의 매출은 상호(shop_name)로 부모 활성계정에 합산.
        # (서브아이디↔login_id 불일치로 매출이 어느 행에도 안 더해져 ~37% 누락되던 것 복구)
        def _norm_shop(s):
            s = (s or '').strip()
            for _suf in (' 지마켓', ' 옥션', '지마켓', '옥션'):
                if s.endswith(_suf):
                    s = s[:-len(_suf)].strip()
            return s
        sales = {}
        try:
            from apps.sales.models import SalesRecord
            active_lids = {a.login_id for a in accts}
            # 상호 → 활성 login_id 인덱스 (셀러명 + 활성계정 대표 매출 shop_name)
            name2lid = {}
            for a in accts:
                if a.seller_name:
                    name2lid.setdefault(_norm_shop(a.seller_name), a.login_id)
            for r in (SalesRecord.objects.filter(platform__in=['gmarket', 'auction'],
                                                 seller__seller_id__in=active_lids)
                      .exclude(shop_name='').values('seller__seller_id', 'shop_name')
                      .annotate(n=Count('id')).order_by('-n')):
                name2lid.setdefault(_norm_shop(r['shop_name']), r['seller__seller_id'])
            # 고아 seller_id의 대표 상호 → 부모 login_id
            orphan_rep = {}
            for r in (SalesRecord.objects.filter(platform=market, order_date__gte=d0, order_date__lte=d1)
                      .exclude(seller__seller_id__in=active_lids)
                      .values('seller__seller_id', 'shop_name')
                      .annotate(n=Count('id')).order_by('-n')):
                sid = r['seller__seller_id']
                if sid and sid not in orphan_rep:
                    orphan_rep[sid] = name2lid.get(_norm_shop(r['shop_name']))
            for r in (SalesRecord.objects
                      .filter(platform=market, order_date__gte=d0, order_date__lte=d1)
                      .values('seller__seller_id')
                      .annotate(revenue=Sum('total_price'), profit=Sum('net_profit'), orders=Count('id'))):
                sid = r['seller__seller_id']
                if not sid:
                    continue
                lid = sid if sid in active_lids else orphan_rep.get(sid)
                if not lid:
                    continue
                d = sales.setdefault(lid, {'revenue': 0, 'profit': 0, 'orders': 0})
                d['revenue'] += r['revenue'] or 0
                d['profit'] += r['profit'] or 0
                d['orders'] += r['orders']
        except Exception:
            pass

        # 쇼핑몰명(매출자료 shop_name) — 계정별 대표값(최다빈도), 마켓접미(' 지마켓'/' 옥션') 제거. ID옆 표시용.
        shop_map = {}
        try:
            from apps.sales.models import SalesRecord as _SR2
            for r in (_SR2.objects.filter(platform__in=['gmarket', 'auction'])
                      .exclude(shop_name='').values('seller__seller_id', 'shop_name')
                      .annotate(n=Count('id')).order_by('-n')):
                sid = r['seller__seller_id']
                if sid and sid not in shop_map:
                    nm = (r['shop_name'] or '').strip()
                    for _suf in (' 지마켓', ' 옥션', '지마켓', '옥션'):
                        if nm.endswith(_suf):
                            nm = nm[:-len(_suf)].strip()
                    shop_map[sid] = nm
        except Exception:
            pass

        rows = []
        tot = {'ad_spend': 0, 'cpc_spend': 0, 'ai_spend': 0, 'server_spend': 0,
               'auction_spend': 0, 'balance': 0, 'product_count': 0,
               'revenue': 0, 'profit': 0, 'net_after_ad': 0, 'orders': 0}
        for a in accts:
            lid = a.login_id
            b = bal.get(lid) or {}
            c = cost.get(lid) or {'cpc': 0, 'ai': 0, 'server': 0, 'auction': 0, 'cnt': 0}
            sp = snap.get(lid) or {'cpc': 0, 'ai': 0, 'auction': 0, 'days': 0}
            # CPC/AI/서버 = 판매예치금 거래내역(반영일까지) + 미반영일(오늘 등) 광고센터 스냅샷 실시간 보충
            if market == 'gmarket':
                cpc = c['cpc'] + sp['cpc']; ai = c['ai'] + sp['ai']
                auction = c['auction'] + sp['auction']        # 他마켓(옥션) 실시간 보충
            else:
                cpc = c['cpc'] + sp['auction']; ai = c['ai']  # 옥션 CPC 실시간 보충
                auction = c['auction'] + sp['cpc'] + sp['ai']  # 他마켓(지마켓) 실시간 보충
            server = c['server']
            spend = cpc + ai + server   # 광고비합계 = 현재 마켓(지마켓/옥션 토글)만 — 마켓별 완전 분리
            pc = prod_mkt.get((lid, market), 0)   # 현재 마켓 상품수만 (지마켓/옥션 분리)
            sl = sales.get(lid) or {'revenue': 0, 'profit': 0, 'orders': 0}
            revenue = sl['revenue']; profit = sl['profit']
            net_after_ad = profit - spend   # 실질순이익 = 순수익(매출-원가) - 광고비
            rows.append({
                'no': a.display_order, 'login_id': lid, 'seller_name': a.seller_name,
                'shop_name': shop_map.get(lid, ''),
                'balance': b.get('total_balance') or 0,
                'ad_spend': spend, 'cpc_spend': cpc, 'ai_spend': ai, 'server_spend': server,
                'auction_spend': auction,
                'ad_count': c['cnt'], 'product_count': pc,
                'gmarket_products': prod_mkt.get((lid, 'gmarket'), 0),
                'auction_products': prod_mkt.get((lid, 'auction'), 0),
                'revenue': revenue, 'profit': profit, 'net_after_ad': net_after_ad,
                'orders': sl['orders'],
                'margin': round(net_after_ad * 100.0 / revenue, 1) if revenue else 0,  # 순수익(광고비 차감 후) 마진
                'roas': round(revenue / spend, 1) if spend else 0,
                'collected_at': b.get('collected_at').isoformat() if b.get('collected_at') else None,
            })
            tot['ad_spend'] += spend; tot['cpc_spend'] += cpc; tot['ai_spend'] += ai
            tot['server_spend'] += server; tot['auction_spend'] += auction
            # 잔액 합계: 공유ESM 서브계정은 마스터와 같은 지갑(잔액) → 중복합산 방지 위해 서브는 제외(마스터만 1회)
            _is_sub = bool(a.gmarket_origin_id and a.gmarket_origin_id != a.login_id)
            if not _is_sub:
                tot['balance'] += (b.get('total_balance') or 0)
            tot['product_count'] += pc
            tot['revenue'] += revenue; tot['profit'] += profit
            tot['net_after_ad'] += net_after_ad; tot['orders'] += sl['orders']
        tot['account_count'] = len(accts)
        return Response({'market': market, 'other_market': other,
                         'date_from': str(d0), 'date_to': str(d1), 'totals': tot, 'rows': rows})


class GmarketCostDetailView(views.APIView):
    """지마켓 계정의 기간 광고비(거래내역) 상세 — 대시보드 모달용."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count
        from apps.cpc.models import GmarketCostHistory
        sid = request.query_params.get('seller_id')
        df = request.query_params.get('date_from')
        dt = request.query_params.get('date_to')
        qs = GmarketCostHistory.objects.all()
        if sid:
            qs = qs.filter(seller_id=sid)
        if df:
            qs = qs.filter(use_date__gte=df)
        if dt:
            qs = qs.filter(use_date__lte=dt)
        ad_types = ['CPC', 'AI매출업', '서버비용']
        ad = qs.filter(transaction_type__in=ad_types)
        agg = ad.aggregate(spend=Sum('amount'), cnt=Count('id'))
        by_type = {x['transaction_type']: abs(x['s'] or 0) for x in
                   ad.values('transaction_type').annotate(s=Sum('amount'))}
        rows = [{
            'use_date': str(r.use_date), 'use_type': r.use_type,
            'transaction_type': r.transaction_type, 'comment': r.comment,
            'amount': r.amount,
        } for r in qs.order_by('-use_date', 'seq')[:2000]]
        return Response({
            'seller_id': sid, 'date_from': df, 'date_to': dt,
            'ad_spend': abs(agg['spend'] or 0), 'ad_count': agg['cnt'] or 0,
            'cpc_spend': by_type.get('CPC', 0), 'ai_spend': by_type.get('AI매출업', 0),
            'server_spend': by_type.get('서버비용', 0),
            'total_rows': qs.count(), 'rows': rows,
        })


class GmarketAdDailyView(views.APIView):
    """지마켓 계정의 기간 광고비 '시간대별' 내역 — 판매예치금 거래원장(GmarketCostHistory)의
    거래시각(traded_at)을 '시(時)' 단위로 묶어 그 시간대에 실제 나간 CPC/AI/서버 광고비를 표시.
    합계(cpc_spend/ai_spend)는 대시보드 셀과 동일. 스냅샷이 아니라 실제 거래내역 기반이라 정확."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import re as _re
        from datetime import datetime, timedelta
        import pytz
        from django.db.models import Sum
        from apps.cpc.models import GmarketCostHistory
        kst = pytz.timezone('Asia/Seoul')
        sid = request.query_params.get('seller_id')
        df = request.query_params.get('date_from')
        dt = request.query_params.get('date_to')
        d1 = datetime.strptime(dt, '%Y-%m-%d').date() if _re.match(r'^\d{4}-\d{2}-\d{2}$', dt or '') else timezone.localdate()
        d0 = datetime.strptime(df, '%Y-%m-%d').date() if _re.match(r'^\d{4}-\d{2}-\d{2}$', df or '') else (d1 - timedelta(days=30))
        TYPE_KEY = {'CPC': 'cpc', 'AI매출업': 'ai', '서버비용': 'server'}
        qs = (GmarketCostHistory.objects
              .filter(market='gmarket', use_date__gte=d0, use_date__lte=d1,
                      transaction_type__in=list(TYPE_KEY.keys()))
              .exclude(comment__icontains='판매예치금'))   # 판매예치금 송금 등 비광고 차감 제외
        if sid:
            qs = qs.filter(seller_id=sid)
        # 거래시각(traded_at)을 KST '시' 단위로 버킷팅 → 그 시간대 실제 발생 광고비
        buckets = {}   # (date, hour) -> dict
        for r in qs.values('traded_at', 'use_date', 'transaction_type', 'amount'):
            ta = r['traded_at']
            if ta is not None:
                lt = ta.astimezone(kst)
                key = (lt.strftime('%Y-%m-%d'), lt.hour)
                hh = '%02d:00~%02d:59' % (lt.hour, lt.hour)
                dlabel = lt.strftime('%Y-%m-%d')
            else:
                key = (str(r['use_date']), -1)
                hh = '시각미상'
                dlabel = str(r['use_date'])
            b = buckets.setdefault(key, {'cpc': 0, 'ai': 0, 'server': 0, 'cnt': 0, 'hh': hh, 'date': dlabel})
            b[TYPE_KEY[r['transaction_type']]] += abs(r['amount'] or 0)
            b['cnt'] += 1
        rows = []
        for (d, h), b in sorted(buckets.items(), key=lambda x: (x[0][0], x[0][1]), reverse=True):
            tot = b['cpc'] + b['ai'] + b['server']
            rows.append({
                'datetime': f"{b['date']} {b['hh']}", 'date': b['date'], 'time': b['hh'],
                'cpc': b['cpc'], 'ai': b['ai'], 'auction': 0, 'server': b['server'],
                'total': tot, 'count': b['cnt'],
                # 시간 버킷 자체가 그 시간대 실제 발생액
                'd_cpc': b['cpc'], 'd_ai': b['ai'], 'd_total': tot,
            })
        bt = {x['transaction_type']: abs(x['s'] or 0)
              for x in qs.values('transaction_type').annotate(s=Sum('amount'))}
        return Response({
            'seller_id': sid, 'date_from': str(d0), 'date_to': str(d1),
            'cpc_spend': bt.get('CPC', 0), 'ai_spend': bt.get('AI매출업', 0),
            'ad_spend': bt.get('CPC', 0) + bt.get('AI매출업', 0) + bt.get('서버비용', 0),
            'points': len(rows), 'rows': rows,
        })


class GmarketAdGroupView(views.APIView):
    """지마켓 CPC 광고그룹별 성과(노출/클릭/광고비) — GmarketAdGroupPerf.
    params: date(stat_date, 기본 최신일), account(선택), ad_type(normal/smart 선택), min_cost(기본 0)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import re as _re
        from datetime import datetime
        from django.db.models import Sum, Count
        from apps.cpc.models import GmarketAdGroupPerf
        qs = GmarketAdGroupPerf.objects.all()
        acct = request.query_params.get('account')
        ad_type = request.query_params.get('ad_type')
        date = request.query_params.get('date')
        try:
            min_cost = int(request.query_params.get('min_cost') or 0)
        except ValueError:
            min_cost = 0
        # 기준일: 지정 없으면 최신 stat_date
        if _re.match(r'^\d{4}-\d{2}-\d{2}$', date or ''):
            stat_date = datetime.strptime(date, '%Y-%m-%d').date()
        else:
            latest = qs.order_by('-stat_date').values_list('stat_date', flat=True).first()
            stat_date = latest
        if stat_date is None:
            return Response({'date': None, 'totals': {}, 'rows': [], 'accounts': []})
        qs = qs.filter(stat_date=stat_date)
        if acct:
            qs = qs.filter(gmarket_id=acct)
        if ad_type in ('normal', 'smart'):
            qs = qs.filter(ad_type=ad_type)
        if min_cost > 0:
            qs = qs.filter(total_cost__gte=min_cost)
        tot = qs.aggregate(groups=Count('id'), impressions=Sum('impressions'),
                           clicks=Sum('clicks'), cost=Sum('total_cost'))
        rows = [{
            'gmarket_id': r.gmarket_id, 'ad_type': r.ad_type, 'ad_group_name': r.ad_group_name,
            'status': r.status, 'ad_on': r.ad_on, 'ad_off': r.ad_off,
            'impressions': r.impressions, 'clicks': r.clicks, 'ctr': float(r.ctr),
            'avg_click_cost': r.avg_click_cost, 'total_cost': r.total_cost,
            'product_count': r.product_count, 'daily_budget': r.daily_budget,
        } for r in qs.order_by('-total_cost', '-clicks')[:3000]]
        # 계정 목록(드롭다운용) — 해당 일자 수집된 계정
        accounts = list(GmarketAdGroupPerf.objects.filter(stat_date=stat_date)
                        .values_list('gmarket_id', flat=True).distinct())
        return Response({
            'date': str(stat_date),
            'totals': {'groups': tot['groups'] or 0, 'impressions': tot['impressions'] or 0,
                       'clicks': tot['clicks'] or 0, 'cost': tot['cost'] or 0},
            'rows': rows, 'accounts': sorted(accounts),
        })


class GmarketMyAccountsView(views.APIView):
    """지마켓 계정 목록(아이디 선택용) + 계정별 상품수."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Count
        from apps.cpc.models import CrawlerAccount, GmarketMyProduct
        cnt = {r['account']: r['n'] for r in GmarketMyProduct.objects.values('account').annotate(n=Count('id'))}
        accts = CrawlerAccount.objects.filter(platform='gmarket', is_active=True).order_by('display_order', 'login_id')
        # 공유ESM 서브 제외 — 마스터(대표)만. 간편광고는 마스터 ESM 단위로 제어됨(서브는 대표에 묶임).
        accts = [a for a in accts if not (a.gmarket_origin_id and a.gmarket_origin_id != a.login_id)]
        data = [{'account_id': a.id, 'login_id': a.login_id, 'seller_name': a.seller_name,
                 'product_count': cnt.get(a.id, 0)} for a in accts]
        return Response({'accounts': data})


class ElevenMyProductDetailView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        result = _emp_svc.get_my_product_detail(pk)
        if not result:
            return Response({'error': '상품을 찾을 수 없습니다.'}, status=404)
        return Response(result)


class ElevenMyProductAccountSummaryView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        all_accounts = request.query_params.get('all') in ('1', 'true', 'True')
        return Response(_emp_svc.get_account_summary(all_accounts=all_accounts))


class ElevenMyProductDuplicateView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .eleven_my_product_service import find_duplicates
        mode = request.query_params.get('mode', 'strict')
        if mode not in ('strict', 'loose', 'image'):
            mode = 'strict'
        return Response(find_duplicates(mode=mode))


class ElevenIntegratedSyncView(views.APIView):
    permission_classes = [IsAuthenticated]
    COOLDOWN_NAME = 'integrated_sync'
    COOLDOWN_MINUTES = 5

    def post(self, request):
        from .eleven_my_product_service import trigger_integrated_sync
        from . import eleven_block_guard as guard

        # 1) 글로벌 차단 락 확인
        blocked, remaining, until = guard.is_blocked()
        if blocked:
            return Response({
                'error': '11번가 글로벌 차단 모드 활성화 중',
                'block_remaining_seconds': remaining,
                'block_until': until.isoformat() if until else None,
            }, status=429)

        # 2) 엔드포인트 쿨다운 (5분 내 연타 방지)
        force = bool(request.data.get('force'))
        cd = guard.cooldown_remaining(self.COOLDOWN_NAME, minutes=self.COOLDOWN_MINUTES)
        if cd > 0 and not force:
            return Response({
                'error': f'쿨다운 중 — {cd}초 후 다시 시도하세요 (force=true로 강제 실행 가능)',
                'cooldown_remaining_seconds': cd,
            }, status=429)

        tasks = request.data.get('tasks') or []
        account_id = request.data.get('account_id')
        try:
            result = trigger_integrated_sync(tasks=tasks, account_id=account_id)
            guard.mark_cooldown(self.COOLDOWN_NAME)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class ElevenSyncStatusView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .eleven_my_product_service import get_sync_status
        from . import eleven_block_guard as guard
        result = get_sync_status()
        # 차단 상태도 함께 반환 (UI 배너용)
        blocked, remaining, until = guard.is_blocked()
        result['block'] = {
            'active': blocked,
            'remaining_seconds': remaining,
            'until': until.isoformat() if until else None,
        }
        cd = guard.cooldown_remaining(ElevenIntegratedSyncView.COOLDOWN_NAME,
                                       minutes=ElevenIntegratedSyncView.COOLDOWN_MINUTES)
        result['integrated_sync_cooldown_seconds'] = cd
        return Response(result)


class ElevenBlockClearView(views.APIView):
    """관리자용 — 차단 락 수동 해제 (잘못 걸렸을 때 빨리 풀려고)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from . import eleven_block_guard as guard
        cleared = guard.clear_block()
        return Response({'cleared': cleared})


class TaxVatSummaryView(views.APIView):
    """11번가 부가세(VAT) 종합 — 계정별·월별 과세매출 + 수집 진행률"""
    def get(self, request):
        from apps.cpc.models import TaxVatMonthly, CrawlerAccount
        year = int(request.query_params.get('year') or 2026)
        qs = TaxVatMonthly.objects.filter(platform='11st', year=year)
        # 셀러명 앞 3글자로 그룹 합산 (예: 스타코1/2/3 → 스타코) + 그룹별 대표 아이디(login_id)
        groups = {}
        monthly_totals = {}
        for r in qs.values('login_id', 'seller_name', 'month').annotate(taxable=Sum('taxable_sales')):
            key = (r['seller_name'] or '?')[:3]
            g = groups.setdefault(key, {'group': key, 'months': {}, 'total': 0, '_members': {}})
            g['months'][str(r['month'])] = g['months'].get(str(r['month']), 0) + r['taxable']
            g['total'] += r['taxable']
            g['_members'][r['login_id']] = g['_members'].get(r['login_id'], 0) + (r['taxable'] or 0)
            monthly_totals[str(r['month'])] = monthly_totals.get(str(r['month']), 0) + r['taxable']
        for g in groups.values():
            members = g.pop('_members')
            # 대표 아이디 = 그룹 내 과세매출 최대 계정
            g['rep_login_id'] = max(members.items(), key=lambda x: x[1])[0] if members else ''
            g['member_count'] = len(members)
        accounts = sorted(groups.values(), key=lambda x: -x['total'])
        target = CrawlerAccount.objects.filter(platform='11st', is_active=True).exclude(api_key='').count()
        collected = qs.values('login_id').distinct().count()
        last = qs.order_by('-collected_at').first()
        grand = sum(monthly_totals.values())
        return Response({
            'year': year,
            'progress': {
                'collected': collected, 'target': target,
                'last_collected_at': last.collected_at.isoformat() if last else None,
            },
            'accounts': accounts,
            'monthly_totals': monthly_totals,
            'grand_total': grand,
            'vat_payable': round(grand / 11),
        })


class ElevenGradeLatestView(views.APIView):
    """계정별 최신 등급 (최신이 null이면 직전 실제 등급으로 보정) — 등급현황 모달용. 페이지네이션 없음."""
    def get(self, request):
        from .models import ElevenSellerGrade
        latest = {}
        for g in ElevenSellerGrade.objects.order_by('-collected_at'):
            cur = latest.get(g.eleven_id)
            if cur is None:
                latest[g.eleven_id] = {
                    'eleven_id': g.eleven_id,
                    'seller_name': g.seller_name or g.eleven_id,
                    'grade': g.grade,
                    'grade_message': g.grade_message or '',
                    'required_sales': g.required_sales,
                    'collected_at': g.collected_at.isoformat() if g.collected_at else '',
                }
            elif cur['grade'] is None and g.grade is not None:
                cur['grade'] = g.grade
                if not cur['grade_message']:
                    cur['grade_message'] = g.grade_message or ''
        rows = sorted(latest.values(), key=lambda x: -(x['grade'] or 0))
        return Response({'results': rows, 'count': len(rows)})


class AllMallProfitView(views.APIView):
    """전체 쇼핑몰 순수익 분석 — 플랫폼별 매출/원가/순익/광고비/순수익.
    ?month=YYYY-MM (미지정 시 이번 달). 순익=매출-원가(SalesRecord.net_profit),
    순수익=순익-광고비. 광고비(전계정 통일, 거래내역 기준): gmarket/auction=GmarketCostHistory,
    11st=ElevenCostHistory(CPC 차감), 그 외 플랫폼은 광고비 데이터 없음(0)."""
    def get(self, request):
        from datetime import date as dt_date, datetime as dt_dt, timedelta as dt_td
        import pytz as _pytz
        from apps.sales.models import SalesRecord
        from apps.cpc.models import GmarketCostHistory, ElevenCostHistory

        mp = request.query_params.get('month')  # 'YYYY-MM'
        today = dt_date.today()
        if mp:
            y, m = int(mp[:4]), int(mp[5:7])
        else:
            y, m = today.year, today.month
        ms = dt_date(y, m, 1)
        me = dt_date(y + (m // 12), (m % 12) + 1, 1) - __import__('datetime').timedelta(days=1)

        # 1) 매출/원가/순익 (플랫폼별)
        base = {}
        for r in (SalesRecord.objects.filter(order_date__gte=ms, order_date__lte=me)
                  .values('platform')
                  .annotate(rev=Sum('total_price'), cost=Sum('cost'),
                            prof=Sum('net_profit'), comm=Sum('commission'), n=Count('id'))):
            p = r['platform'] or '기타'
            base[p] = {'revenue': r['rev'] or 0, 'cost': r['cost'] or 0,
                       'gross_profit': r['prof'] or 0, 'commission': r['comm'] or 0,
                       'orders': r['n']}

        # 2) 광고비 (플랫폼별)
        ad = {}
        for r in (GmarketCostHistory.objects
                  .filter(use_date__gte=ms, use_date__lte=me,
                          transaction_type__in=['CPC', 'AI매출업'])
                  .exclude(comment__icontains='판매예치금')   # 판매예치금 송금 등 비광고 차감 제외
                  .values('market').annotate(a=Sum('amount'))):
            mk = r['market'] or 'gmarket'
            ad[mk] = ad.get(mk, 0) + abs(r['a'] or 0)
        # 11번가도 거래내역(ElevenCostHistory) 기준으로 통일 — 전계정 CPC 차감 합계(지마켓과 동일 방식).
        _kst = _pytz.timezone('Asia/Seoul')
        _s = _kst.localize(dt_dt.combine(ms, dt_dt.min.time()))
        _e = _kst.localize(dt_dt.combine(me, dt_dt.min.time()) + dt_td(days=1))
        st11_ad = abs(ElevenCostHistory.objects.filter(
            transaction_datetime__gte=_s, transaction_datetime__lt=_e,
            transaction_type='CPC', amount__lt=0).aggregate(s=Sum('amount'))['s'] or 0)
        ad_map = {'gmarket': ad.get('gmarket', 0), 'auction': ad.get('auction', 0), '11st': st11_ad}

        # 3) 플랫폼 행 구성
        LABELS = {'gmarket': '지마켓', 'auction': '옥션', '11st': '11번가',
                  'smartstore': '스마트스토어', 'coupang': '쿠팡',
                  'lotteon': '롯데온', '18.롯데온': '롯데온(기타)'}
        rows = []
        tot = {'revenue': 0, 'cost': 0, 'gross_profit': 0, 'ad_cost': 0,
               'net_profit': 0, 'orders': 0, 'commission': 0}
        for p, b in base.items():
            ad_cost = ad_map.get(p, 0)
            net = b['gross_profit'] - ad_cost
            rev = b['revenue']
            rows.append({
                'platform': p, 'label': LABELS.get(p, p),
                'revenue': rev, 'cost': b['cost'], 'commission': b['commission'],
                'gross_profit': b['gross_profit'], 'ad_cost': ad_cost,
                'net_profit': net, 'orders': b['orders'],
                'net_margin': round(net / rev * 100, 1) if rev else 0,
                'ad_ratio': round(ad_cost / rev * 100, 1) if rev else 0,
                'has_ad_data': p in ad_map,
            })
            for k in ('revenue', 'cost', 'gross_profit', 'commission', 'orders'):
                tot[k] += b[k]
            tot['ad_cost'] += ad_cost
            tot['net_profit'] += net
        rows.sort(key=lambda x: -x['revenue'])

        rev = tot['revenue']
        tot['net_margin'] = round(tot['net_profit'] / rev * 100, 1) if rev else 0
        tot['gross_margin'] = round(tot['gross_profit'] / rev * 100, 1) if rev else 0
        return Response({'month': f'{y}-{m:02d}', 'date_from': str(ms), 'date_to': str(me),
                         'rows': rows, 'totals': tot})


class ElevenAdKilllistView(views.APIView):
    """11번가 광고 킬-리스트 — 매출 0인데 광고비 쓰는 상품(광고 끌 대상).
    ?month=YYYY-MM&min_cost=2000&export=1. 고광고비 후보만 매출매칭(효율)."""
    def get(self, request):
        from datetime import date as dt_date
        import datetime as _dt
        import re as _re
        from apps.cpc.models import St11ProductDaily, ElevenMyProduct
        from apps.sales.models import SalesRecord

        mp = request.query_params.get('month')
        today = dt_date.today()
        if mp:
            y, m = int(mp[:4]), int(mp[5:7])
        else:
            y, m = today.year, today.month
        ms = dt_date(y, m, 1)
        me = dt_date(y + (m // 12), (m % 12) + 1, 1) - _dt.timedelta(days=1)
        try:
            min_cost = max(0, int(request.query_params.get('min_cost') or 2000))
        except ValueError:
            min_cost = 2000

        # 1) 월 광고비 ≥ min_cost 상품 후보 (eleven_id, 정제 product_no)
        cand = {}
        for r in (St11ProductDaily.objects.filter(stat_date__gte=ms, stat_date__lte=me)
                  .values('eleven_id', 'product_no').annotate(c=Sum('cost'))):
            pno = _re.sub(r'\s*\(.*?\)\s*', '', str(r['product_no'])).strip()
            if not pno.isdigit():
                continue
            cand[(r['eleven_id'], pno)] = cand.get((r['eleven_id'], pno), 0) + (r['c'] or 0)
        cand = {k: v for k, v in cand.items() if v >= min_cost}

        # 2) 판매자코드 매핑(전역) + 3) 매출(판매자코드별, 월)
        pnos = [int(p) for (_, p) in cand.keys()]
        code_map = {}
        if pnos:
            for p in (ElevenMyProduct.objects.filter(product_no__in=pnos).exclude(seller_product_code='')
                      .values('product_no', 'seller_product_code', 'product_name', 'status_type')):
                code_map.setdefault(str(p['product_no']),
                                    (p['seller_product_code'] or '', p['product_name'] or '', p['status_type'] or ''))
        codes = set()
        for sc, _nm, _st in code_map.values():
            if sc:
                codes.add(sc); codes.add(_bare_seller_code(sc))
        sales_by_code = {}
        if codes:
            for s in (SalesRecord.objects.filter(platform='11st', product_code__in=list(codes),
                                                 order_date__gte=ms, order_date__lte=me)
                      .values('product_code').annotate(s=Sum('total_price'))):
                sales_by_code[s['product_code']] = s['s'] or 0

        # 4) 매출 0인 것만 = 킬 대상
        rows = []
        for (eid, pno), cost in cand.items():
            info = code_map.get(pno)
            sc = info[0] if info else ''
            sales = sum(sales_by_code.get(x, 0) for x in {sc, _bare_seller_code(sc)}) if sc else 0
            if sales == 0:
                rows.append({'eleven_id': eid, 'product_no': pno, 'seller_code': sc,
                             'product_name': info[1] if info else '', 'cost': cost,
                             'status': (info[2] if info else '') or '미등록'})
        rows.sort(key=lambda x: (x['eleven_id'], -x['cost']))

        if request.query_params.get('export'):
            import csv as _csv
            from django.http import HttpResponse
            resp = HttpResponse(content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = f'attachment; filename="11st_killlist_{y}-{m:02d}.csv"'
            resp.write('﻿')
            w = _csv.writer(resp)
            w.writerow(['계정', '상품번호', '판매자코드', '상품명', '월광고비', '월매출', '상태'])
            for r in rows:
                w.writerow([r['eleven_id'], r['product_no'], r['seller_code'], r['product_name'], r['cost'], 0, r['status']])
            return resp

        import collections
        per = collections.defaultdict(lambda: {'count': 0, 'cost': 0})
        for r in rows:
            per[r['eleven_id']]['count'] += 1
            per[r['eleven_id']]['cost'] += r['cost']
        by_account = sorted(
            [{'eleven_id': k, 'count': v['count'], 'cost': v['cost']} for k, v in per.items()],
            key=lambda x: -x['cost'])
        return Response({
            'month': f'{y}-{m:02d}', 'min_cost': min_cost,
            'count': len(rows), 'total_cost': sum(r['cost'] for r in rows),
            'by_account': by_account, 'rows': rows[:3000],
        })


def _gmkt_roas_period(request):
    """ROAS 페이지 기간 파싱 → (months[(y,m)...], d0, d1).
    ym_from/ym_to('YYYY-MM') 범위 우선. 없으면 year/month 단일월(하위호환)."""
    import calendar as _cal
    from datetime import date as _d
    today = timezone.localdate()

    def _p(s):
        try:
            y, m = str(s).split('-')
            y, m = int(y), int(m)
            return (y, m) if 1 <= m <= 12 else None
        except Exception:
            return None
    a, b = _p(request.query_params.get('ym_from')), _p(request.query_params.get('ym_to'))
    if a and b:
        (y0, m0), (y1, m1) = a, b
    else:
        try:
            y0 = int(request.query_params.get('year') or today.year)
            m0 = int(request.query_params.get('month') or today.month)
        except (ValueError, TypeError):
            y0, m0 = today.year, today.month
        y1, m1 = y0, m0
    if (y0, m0) > (y1, m1):
        (y0, m0), (y1, m1) = (y1, m1), (y0, m0)
    months, yy, mm = [], y0, m0
    while (yy, mm) <= (y1, m1):
        months.append((yy, mm))
        mm += 1
        if mm > 12:
            mm, yy = 1, yy + 1
    # 집계 최대 2년(24개월) — 초과 시 가장 최근 24개월만
    if len(months) > 24:
        months = months[-24:]
        y0, m0 = months[0]
        y1, m1 = months[-1]
    d0 = _d(y0, m0, 1)
    d1 = _d(y1, m1, _cal.monthrange(y1, m1)[1])
    return months, d0, d1


def _gmkt_month_q(months):
    """[(y,m)...] → GmarketProductAdCost용 (year,month) 범위 Q.
    기존엔 월마다 Q(year,month) OR 체인(최대24개)을 만들어 옵티마이저가 (year,month)
    복합인덱스 range를 못 쓰고 198만행 풀스캔(연단위 ~178초)했음. months가 연속이므로
    최대 3개 범위항(첫해>=m0 | 중간연 전체 | 끝해<=m1)으로 줄임."""
    from django.db.models import Q as _Q
    if not months:
        return _Q(pk__in=[])
    (y0, m0), (y1, m1) = months[0], months[-1]
    if y0 == y1:
        return _Q(year=y0, month__gte=m0, month__lte=m1)
    q = _Q(year=y0, month__gte=m0) | _Q(year=y1, month__lte=m1)
    if y1 - y0 >= 2:
        q |= _Q(year__gt=y0, year__lt=y1)
    return q


def _gmkt_cache_key(prefix, request):
    """쿼리파라미터 기반 결정적 캐시키. ROAS/대시보드 집계는 일1회 크롤 때만 변하므로 캐시 적합."""
    items = sorted((k, v) for k, v in request.query_params.items() if k != 'export')
    return prefix + ':' + '&'.join(f'{k}={v}' for k, v in items)


def _gmkt_realsales_window(adcost_qs):
    """광고데이터(GmarketProductAdCost 쿼리셋)가 실제 존재하는 월의 최소~최대 범위(d0,d1).
    실매출 매칭을 광고비가 있는 기간에 맞춰, 기간 불일치로 ROAS가 부풀려지는 것 방지.
    (예: 년간 선택해도 6월만 광고데이터면 실매출도 6월만 집계)"""
    import calendar as _cal
    from datetime import date as _d
    present = sorted(adcost_qs.values_list('year', 'month').distinct())
    if not present:
        return None, None
    (y0, m0), (y1, m1) = present[0], present[-1]
    return _d(y0, m0, 1), _d(y1, m1, _cal.monthrange(y1, m1)[1])


# ESM sellStatus 코드/라벨 → 표시 상태. '11'/'21'=판매중(실측: 11은 100% 재고>0).
_GMKT_STATUS = {
    '11': '판매중', '21': '판매중', '판매중': '판매중',
    '22': '판매중지', '판매중지': '판매중지',
    '23': '품절', '품절': '품절',
    '24': '판매종료', '판매종료': '판매종료',
    '25': '판매불가', '판매불가': '판매불가',
}


def _gmkt_status_label(raw, in_catalog=True):
    """상품상태 표시값. 카탈로그(MyProduct)에 없으면 '삭제'. 코드는 라벨로 매핑(미지정은 원본)."""
    if not in_catalog:
        return '삭제'
    raw = (raw or '').strip()
    if not raw:
        return '미상'
    return _GMKT_STATUS.get(raw, raw)


def _gmarket_realsales(d0, d1, product_nos):
    """광고 상품번호 → (판매자코드, 실매출, 상품상태, 실구매건수) 매핑.
    다리: GmarketMyProduct.product_no→seller_product_code(자체/판매자코드) ↔ SalesRecord.product_code.
    실매출 = 해당 판매자코드의 d0~d1 매출(total_price) 전역합, 실구매건수 = 매칭 주문 건수(Count).
    (지마켓+옥션, 판매자ID 무관 전역매칭). status_by_pno: 최신 MyProduct status_type(키없음=삭제)."""
    if not d0 or not d1:
        return {}, {}, {}, {}
    from django.db.models import Sum as _Sum, Count as _Count
    from apps.cpc.models import GmarketMyProduct
    from django.db.models import Max as _Max
    from datetime import timedelta as _td
    pnos = {p for p in product_nos if p}
    if not pnos:
        return {}, {}, {}, {}
    # 계정별 최신 크롤 시각: 판매불가 상품은 goods/search API에서 통째로 빠지므로
    # 마지막 "판매중" 스냅샷이 박제된다. 그 계정 최신 크롤보다 synced_at이 한참 이전이면
    # = 최신 크롤에서 누락 = 더이상 판매목록에 없음(판매불가/삭제) → 비고를 '판매불가'로 표시.
    acct_latest = {r['account_id']: r['mx'] for r in (
        GmarketMyProduct.objects.values('account_id').annotate(mx=_Max('synced_at')))}
    # 3일+ 연속 누락만 판매불가로 판정(12h/1일은 크롤 변동·부분실패로 멀쩡한 상품을 오판 → 검증결과 3일이 안전).
    STALE = _td(days=3)
    code_by_pno = {}
    status_by_pno = {}
    for p in (GmarketMyProduct.objects.filter(product_no__in=pnos)
              .order_by('-synced_at')
              .values('product_no', 'seller_product_code', 'status_type', 'synced_at', 'account_id')):
        pno = p['product_no']
        if pno not in status_by_pno:
            st = p['status_type'] or ''
            latest = acct_latest.get(p['account_id'])
            if latest and p['synced_at'] and (latest - p['synced_at']) > STALE:
                st = '판매불가'   # 최신 크롤 누락 = 판매목록에서 제외됨(판매불가/삭제)
            status_by_pno[pno] = st
        if p['seller_product_code'] and pno not in code_by_pno:
            code_by_pno[pno] = p['seller_product_code']
    # 카탈로그에서 삭제된 상품은 영구보존고(ProductCodeArchive)에서 판매자코드 보충 → 빈칸 자동채움
    miss = [p for p in product_nos if p not in code_by_pno]
    if miss:
        from apps.cpc.models import ProductCodeArchive
        arch = {a['product_no']: a['seller_code'] for a in (
            ProductCodeArchive.objects.filter(platform='gmarket', product_no__in=[str(x) for x in miss])
            .exclude(seller_code='').values('product_no', 'seller_code'))}
        for p in miss:
            sc = arch.get(str(p))
            if sc:
                code_by_pno[p] = sc
                status_by_pno.setdefault(p, '삭제(코드보존)')
    codes = set(code_by_pno.values())
    sales_by_code = {}
    orders_by_code = {}
    if codes:
        from apps.sales.models import SalesRecord
        for r in (SalesRecord.objects
                  .filter(platform__in=['gmarket', 'auction'],
                          order_date__gte=d0, order_date__lte=d1,
                          product_code__in=codes)
                  .values('product_code').annotate(s=_Sum('total_price'), n=_Count('id'))):
            sales_by_code[r['product_code']] = r['s'] or 0
            orders_by_code[r['product_code']] = r['n'] or 0
    real_by_pno = {pno: sales_by_code.get(code, 0) for pno, code in code_by_pno.items()}
    realorders_by_pno = {pno: orders_by_code.get(code, 0) for pno, code in code_by_pno.items()}
    return code_by_pno, real_by_pno, status_by_pno, realorders_by_pno


class GmarketProductRoasView(views.APIView):
    """지마켓 상품별 광고비/ROAS — CPC / AI매출업 구분. (GmarketProductAdCost)
    params: year, month(기본 당월), eid(login_id 선택), ad_type(cpc|ai|선택)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count
        from django.core.cache import cache as _cache
        from apps.cpc.models import GmarketProductAdCost, CrawlerAccount
        ck = _gmkt_cache_key('gmkt_prodroas', request)
        if not request.query_params.get('export'):
            cached = _cache.get(ck)
            if cached is not None:
                return Response(cached)
        months, d0, d1 = _gmkt_roas_period(request)
        mq = _gmkt_month_q(months)
        eid = request.query_params.get('eid') or ''
        only = request.query_params.get('ad_type') or ''

        base = GmarketProductAdCost.objects.filter(mq)
        if eid:
            base = base.filter(login_id=eid)
        if only in ('cpc', 'ai'):
            base = base.filter(ad_type=only)

        name_map = {a.login_id: (a.seller_name or a.login_id)
                    for a in CrawlerAccount.objects.filter(platform='gmarket')}

        # 상품번호 → 판매자코드 + 실매출(매출자료 전역매칭, 광고데이터 있는 월로 한정)
        all_pno = set(base.values_list('product_no', flat=True))
        rd0, rd1 = _gmkt_realsales_window(base)
        code_by_pno, real_by_pno, status_by_pno, realorders_by_pno = _gmarket_realsales(rd0, rd1, all_pno)

        # 상품별 수집 키워드(GmarketKeywordReport, 기간 연도 집계) — 상세 표 키워드 컬럼용
        from collections import defaultdict as _dd
        from apps.cpc.models import GmarketKeywordReport
        _years = sorted({yy for (yy, mm) in months})
        _kagg = _dd(lambda: {'cost': 0, 'clicks': 0, 'conv_amount': 0})
        for k in (GmarketKeywordReport.objects.filter(product_no__in=all_pno, year__in=_years)
                  .values('product_no', 'keyword', 'cost', 'clicks', 'conv_amount')):
            _a = _kagg[(k['product_no'], k['keyword'])]
            _a['cost'] += k['cost'] or 0; _a['clicks'] += k['clicks'] or 0; _a['conv_amount'] += k['conv_amount'] or 0
        # 키워드는 효율(ROAS) 있는 것만 표시 — 기본 100% 이상(전체 나열 금지). ?kw_roas_min 로 조정.
        _kw_rmin = request.query_params.get('kw_roas_min')
        _kw_rmin = float(_kw_rmin) if _kw_rmin not in (None, '') else 100.0
        kw_by_pno = _dd(list)
        for (_pno, _kw), _a in _kagg.items():
            _kr = round(_a['conv_amount'] * 100.0 / _a['cost'], 1) if _a['cost'] else 0
            if _kr < _kw_rmin:
                continue
            kw_by_pno[_pno].append({'keyword': _kw, 'cost': _a['cost'], 'clicks': _a['clicks'],
                                    'conv_amount': _a['conv_amount'], 'roas': _kr})
        for _pno in kw_by_pno:
            kw_by_pno[_pno].sort(key=lambda x: -x['cost'])

        if request.query_params.get('export'):
            import csv as _csv
            from django.http import HttpResponse
            from django.db.models import Q as _Q
            q = base.order_by('login_id', 'ad_type', '-cost')
            if request.query_params.get('loss'):
                cmin = int(request.query_params.get('cost_min') or 3000)
                rmax = float(request.query_params.get('roas_max') or 100)
                clkmin = int(request.query_params.get('clicks_min') or 15)
                # 적자: 광고비≥cmin & 클릭≥clkmin & (구매금액0 또는 ROAS≤rmax)
                q = (q.filter(cost__gte=cmin, clicks__gte=clkmin)
                     .filter(_Q(conv_amount=0) | _Q(roas__lte=rmax)))
            elif request.query_params.get('roas_min'):
                rmin = float(request.query_params.get('roas_min') or 200)
                cmin = int(request.query_params.get('cost_min') or 0)
                clkmin = int(request.query_params.get('clicks_min') or 0)
                # 고효율 대상: ROAS≥rmin (구매금액>0). cost_min/clicks_min은 소액 우연전환 노이즈 제거용(선택)
                q = q.filter(roas__gte=rmin, conv_amount__gt=0)
                if cmin:
                    q = q.filter(cost__gte=cmin)
                if clkmin:
                    q = q.filter(clicks__gte=clkmin)
                q = q.order_by('-roas', '-cost')
            _yf = f'{months[0][0]}{months[0][1]:02d}' if months else ''
            _yt = f'{months[-1][0]}{months[-1][1]:02d}' if months else ''
            fname = request.query_params.get('fname') or f'gmarket_product_roas_{_yf}_{_yt}.csv'
            resp = HttpResponse(content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = f'attachment; filename="{fname}"'
            resp.write('﻿')
            w = _csv.writer(resp)
            w.writerow(['계정', '상호', '판매자ID', '광고유형', '상품번호', '판매자코드', '그룹명', '사이트',
                        '노출수', '클릭수', '평균클릭비용', '광고비', '구매수', '구매금액', '전환율', '광고수익률',
                        '실매출', '실구매건수', '실ROAS', '비고(상품상태)'])
            for r in q[:50000]:
                _rs = real_by_pno.get(r.product_no, 0)
                _ro = realorders_by_pno.get(r.product_no, 0)
                _st = _gmkt_status_label(status_by_pno.get(r.product_no), r.product_no in status_by_pno)
                w.writerow([r.login_id, name_map.get(r.login_id, r.login_id), r.seller_id,
                            'CPC' if r.ad_type == 'cpc' else 'AI매출업', r.product_no,
                            code_by_pno.get(r.product_no, ''), r.group_name, r.site,
                            r.impressions, r.clicks, r.avg_click_cost, r.cost, r.orders,
                            r.conv_amount, float(r.conv_rate), float(r.roas),
                            _rs, _ro, round(_rs * 100.0 / r.cost, 1) if r.cost else 0, _st])
            return resp

        def pack(ad_type):
            qs = base.filter(ad_type=ad_type)
            # 상품번호 단위 집계(여러 달 선택 시 동일 상품 누적). 대표 속성은 임의 1행.
            grp = list(qs.values('product_no')
                       .annotate(cost=Sum('cost'), conv=Sum('conv_amount'),
                                 clicks=Sum('clicks'), impressions=Sum('impressions'),
                                 orders=Sum('orders'))
                       .order_by('-cost')[:5000])
            attr = {}
            for a in qs.values('product_no', 'login_id', 'seller_id', 'group_name', 'site'):
                attr.setdefault(a['product_no'], a)
            rows = []
            real_total = 0
            for g in grp:
                pno = g['product_no']
                a = attr.get(pno, {})
                cost = g['cost'] or 0
                conv = g['conv'] or 0
                clk = g['clicks'] or 0
                rs = real_by_pno.get(pno, 0)
                real_total += rs
                rows.append({
                    'login_id': a.get('login_id', ''), 'seller_id': a.get('seller_id', ''),
                    'seller_name': name_map.get(a.get('login_id', ''), a.get('login_id', '')),
                    'product_no': pno, 'seller_code': code_by_pno.get(pno, ''),
                    'group_name': a.get('group_name', ''), 'site': a.get('site', ''),
                    'impressions': g['impressions'] or 0, 'clicks': clk,
                    'avg_click_cost': round(cost / clk) if clk else 0, 'cost': cost,
                    'orders': g['orders'] or 0, 'conv_amount': conv,
                    'conv_rate': round(conv * 100.0 / cost, 2) if cost else 0.0,
                    'roas': round(conv * 100.0 / cost, 1) if cost else 0,
                    'real_sales': rs, 'real_roas': round(rs * 100.0 / cost, 1) if cost else 0,
                    'real_orders': realorders_by_pno.get(pno, 0),
                    'status': _gmkt_status_label(status_by_pno.get(pno), pno in status_by_pno),
                    'keywords': kw_by_pno.get(pno, [])[:30],   # 수집된 키워드(광고비순)
                })
            agg = qs.aggregate(cost=Sum('cost'), conv=Sum('conv_amount'),
                               clk=Sum('clicks'), imp=Sum('impressions'))
            cost = agg['cost'] or 0
            conv = agg['conv'] or 0
            return {
                'products': len(grp),
                'totals': {
                    'cost': cost, 'conv_amount': conv, 'clicks': agg['clk'] or 0,
                    'impressions': agg['imp'] or 0,
                    'roas': round(conv * 100.0 / cost, 1) if cost else 0,
                    'real_sales': real_total,
                    'real_roas': round(real_total * 100.0 / cost, 1) if cost else 0,
                },
                'rows': rows,
            }

        out = {'ym_from': f'{months[0][0]}-{months[0][1]:02d}' if months else '',
               'ym_to': f'{months[-1][0]}-{months[-1][1]:02d}' if months else '', 'eid': eid}
        if only in ('cpc', 'ai'):
            out[only] = pack(only)
        else:
            out['cpc'] = pack('cpc')
            out['ai'] = pack('ai')
        _cache.set(ck, out, 300)
        return Response(out)


class GmarketRoasAccountsView(views.APIView):
    """지마켓 상품ROAS — 계정별 요약(대시보드 진입 목록). CPC/AI 광고비·구매금액·ROAS 합."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum, Count
        from django.core.cache import cache as _cache
        from apps.cpc.models import GmarketProductAdCost, CrawlerAccount
        ck = _gmkt_cache_key('gmkt_roasacc', request)
        cached = _cache.get(ck)
        if cached is not None:
            return Response(cached)
        months, d0, d1 = _gmkt_roas_period(request)
        mq = _gmkt_month_q(months)
        name_map = {a.login_id: (a.seller_name or a.login_id)
                    for a in CrawlerAccount.objects.filter(platform='gmarket')}
        # 계정별 광고 상품번호 → 실매출(매출자료 전역매칭)
        from collections import defaultdict as _dd
        pno_by_lid = _dd(set)
        for r in (GmarketProductAdCost.objects.filter(mq)
                  .values('login_id', 'product_no')):
            pno_by_lid[r['login_id']].add(r['product_no'])
        all_pno = set().union(*pno_by_lid.values()) if pno_by_lid else set()
        rd0, rd1 = _gmkt_realsales_window(GmarketProductAdCost.objects.filter(mq))   # 실매출은 광고데이터 있는 월로 한정
        _code_by_pno, real_by_pno, _status_by_pno, _ord_by_pno = _gmarket_realsales(rd0, rd1, all_pno)
        rows = {}
        for r in (GmarketProductAdCost.objects.filter(mq)
                  .values('login_id', 'ad_type')
                  .annotate(cost=Sum('cost'), conv=Sum('conv_amount'), n=Count('id'))):
            d = rows.setdefault(r['login_id'], {
                'login_id': r['login_id'], 'seller_name': name_map.get(r['login_id'], r['login_id']),
                'cpc_cost': 0, 'ai_cost': 0, 'cpc_conv': 0, 'ai_conv': 0,
                'cpc_products': 0, 'ai_products': 0})
            d[f"{r['ad_type']}_cost"] = r['cost'] or 0
            d[f"{r['ad_type']}_conv"] = r['conv'] or 0
            d[f"{r['ad_type']}_products"] = r['n'] or 0
        def _roas(conv, cost):
            return round(conv * 100.0 / cost, 1) if cost else 0
        out = []
        for d in rows.values():
            cost = d['cpc_cost'] + d['ai_cost']
            conv = d['cpc_conv'] + d['ai_conv']
            d['total_cost'] = cost
            d['total_conv'] = conv
            d['cpc_roas'] = _roas(d['cpc_conv'], d['cpc_cost'])   # CPC 단독 ROAS
            d['ai_roas'] = _roas(d['ai_conv'], d['ai_cost'])      # AI 단독 ROAS
            d['roas'] = _roas(conv, cost)                          # 합계 ROAS(광고전환매출)
            # 실매출 = 광고상품번호 전역매칭(상품번호당 1회) → 실ROAS
            real = sum(real_by_pno.get(p, 0) for p in pno_by_lid.get(d['login_id'], ()))
            d['real_sales'] = real
            d['real_roas'] = _roas(real, cost)
            out.append(d)
        out.sort(key=lambda x: -x['total_cost'])
        t_cpc_cost = sum(x['cpc_cost'] for x in out)
        t_ai_cost = sum(x['ai_cost'] for x in out)
        t_cpc_conv = sum(x['cpc_conv'] for x in out)
        t_ai_conv = sum(x['ai_conv'] for x in out)
        t_real = sum(x['real_sales'] for x in out)
        payload = {'ym_from': f'{months[0][0]}-{months[0][1]:02d}' if months else '',
                   'ym_to': f'{months[-1][0]}-{months[-1][1]:02d}' if months else '',
                   'months': len(months), 'accounts': out,
                   'totals': {
                       'cpc_cost': t_cpc_cost, 'ai_cost': t_ai_cost,
                       'cpc_conv': t_cpc_conv, 'ai_conv': t_ai_conv,
                       'cost': t_cpc_cost + t_ai_cost,
                       'conv_amount': t_cpc_conv + t_ai_conv,
                       'cpc_roas': _roas(t_cpc_conv, t_cpc_cost),   # CPC 단독
                       'ai_roas': _roas(t_ai_conv, t_ai_cost),      # AI 단독
                       'roas': _roas(t_cpc_conv + t_ai_conv, t_cpc_cost + t_ai_cost),  # 합계(광고전환)
                       'real_sales': t_real,
                       'real_roas': _roas(t_real, t_cpc_cost + t_ai_cost),   # 실매출 기준
                   }}
        _cache.set(ck, payload, 300)
        return Response(payload)


def _gmkt_product_rows(request):
    """지마켓 상품목록(기간 집계, 상품번호 단위) 행 리스트 — 적자/ROAS200%↑/전체 모달 공통.
    필터(모두 선택): cost_min, clicks_min, roas_min, roas_max(미지정=무제한). ROAS=광고전환(conv/cost).
    status: 삭제완료(GmarketLossDeleted) > 삭제(카탈로그없음) > 상품상태(판매중/중지/불가...)."""
    from django.db.models import Sum, Max
    from apps.cpc.models import GmarketProductAdCost, CrawlerAccount, GmarketLossDeleted, GmarketMyProduct
    months, d0, d1 = _gmkt_roas_period(request)
    mq = _gmkt_month_q(months)
    eid = request.query_params.get('eid') or ''
    ad_type = request.query_params.get('ad_type') or ''   # cpc(=키워드광고)/ai 한정
    cost_min = int(request.query_params.get('cost_min') or 0)
    clicks_min = int(request.query_params.get('clicks_min') or 0)
    _rmax = request.query_params.get('roas_max')
    _rmin = request.query_params.get('roas_min')
    roas_max = float(_rmax) if _rmax not in (None, '') else None
    roas_min = float(_rmin) if _rmin not in (None, '') else None
    CAP = 5000

    base = GmarketProductAdCost.objects.filter(mq)
    if eid:
        base = base.filter(login_id=eid)
    if ad_type in ('cpc', 'ai'):
        base = base.filter(ad_type=ad_type)
    name_map = {a.login_id: (a.seller_name or a.login_id)
                for a in CrawlerAccount.objects.filter(platform='gmarket')}
    grp = (base.values('product_no')
           .annotate(cost=Sum('cost'), conv=Sum('conv_amount'), clicks=Sum('clicks'), orders=Sum('orders'))
           .filter(cost__gte=cost_min, clicks__gte=clicks_min)
           .order_by('-cost'))
    attr = {}
    for a in base.values('product_no', 'login_id', 'site'):
        attr.setdefault(a['product_no'], a)
    grp_list = list(grp)
    all_pno = {g['product_no'] for g in grp_list}
    rd0, rd1 = _gmkt_realsales_window(base)
    code_by_pno, real_by_pno, status_by_pno, realorders_by_pno = _gmarket_realsales(rd0, rd1, all_pno)
    # 누적 판매수량(2025-01-01 ~ 현재) — 판매자코드 전역매칭(지마켓+옥션). 평균단가 왼쪽 표기용.
    from datetime import date as _cum_date
    cum_qty_by_pno = {}
    _cum_codes = {c for c in code_by_pno.values() if c}
    if _cum_codes:
        from apps.sales.models import SalesRecord as _SR
        _cum_by_code = {}
        for _r in (_SR.objects.filter(platform__in=['gmarket', 'auction'],
                                      order_date__gte=_cum_date(2025, 1, 1),
                                      product_code__in=_cum_codes)
                   .values('product_code').annotate(q=Sum('quantity'))):
            _cum_by_code[_r['product_code']] = _r['q'] or 0
        cum_qty_by_pno = {pno: _cum_by_code.get(code, 0) for pno, code in code_by_pno.items()}
    # 상태 신선도 — '판매중' 스냅샷이 오래됐으면(STALE_DAYS+) 실제 판매중지일 수 있어 '확인필요' 경고.
    from django.utils import timezone as _tz
    STALE_DAYS = 2
    synced_by_pno = {r['product_no']: r['m'] for r in
                     (GmarketMyProduct.objects.filter(product_no__in=all_pno)
                      .values('product_no').annotate(m=Max('synced_at'))) if r['m']}
    _now = _tz.now()
    deleted = set(GmarketLossDeleted.objects.values_list('login_id', 'product_no'))
    # 상품별 수집 키워드(있으면) — 기간 집계(키워드 단위 합), 광고비 오른쪽에 표기용
    from collections import defaultdict as _dd
    from apps.cpc.models import GmarketKeywordReport
    _years = sorted({yy for (yy, mm) in months})   # 연도-버킷(month=0) 포함 위해 연도기준 매칭
    # 키워드 엑셀 상세양식용 전체 필드 집계(노출/구매수/평균순위/평균클릭비용/전환율 등)
    # (login_id, 상품번호, 키워드) 단위 집계 — 키워드의 실제 수집 계정(login_id)을 보존(공유ESM 정확).
    _kagg = _dd(lambda: {'cost': 0, 'clicks': 0, 'conv_amount': 0, 'impressions': 0, 'orders': 0, 'avg_rank': ''})
    for k in (GmarketKeywordReport.objects.filter(product_no__in=all_pno, year__in=_years)
              .values('login_id', 'product_no', 'keyword', 'cost', 'clicks', 'conv_amount', 'impressions', 'orders', 'avg_rank')):
        a = _kagg[(k['login_id'], k['product_no'], k['keyword'])]
        a['cost'] += k['cost'] or 0; a['clicks'] += k['clicks'] or 0; a['conv_amount'] += k['conv_amount'] or 0
        a['impressions'] += k['impressions'] or 0; a['orders'] += k['orders'] or 0
        if k.get('avg_rank') not in (None, ''):
            a['avg_rank'] = k['avg_rank']
    _kw_rmin = request.query_params.get('kw_roas_min')   # 키워드 ROAS 하한(키워드 화면=200, 적자=미적용)
    _kw_rmin = float(_kw_rmin) if _kw_rmin not in (None, '') else None
    kw_by_pno = _dd(list)
    for (lid, pno, kw), a in _kagg.items():
        _kr = round(a['conv_amount'] * 100.0 / a['cost'], 1) if a['cost'] else 0
        if _kw_rmin is not None and _kr < _kw_rmin:
            continue   # ROAS 하한 미만 키워드 제외(ROAS200%↑ 키워드만 표기)
        _imp = a['impressions']; _clk = a['clicks']; _ord = a['orders']
        kw_by_pno[pno].append({'login_id': lid, 'keyword': kw, 'cost': a['cost'], 'clicks': _clk,
                               'conv_amount': a['conv_amount'], 'roas': _kr,
                               'impressions': _imp, 'orders': _ord,
                               'click_rate': round(_clk * 100.0 / _imp, 2) if _imp else 0,
                               'avg_click_cost': round(a['cost'] / _clk) if _clk else 0,
                               'avg_rank': a['avg_rank'],
                               'conv_rate': round(_ord * 100.0 / _clk, 2) if _clk else 0})
    for pno in kw_by_pno:
        kw_by_pno[pno].sort(key=lambda x: -x['cost'])
    rows = []
    for g in grp_list:
        pno = g['product_no']
        cost = g['cost'] or 0
        conv = g['conv'] or 0
        roas = round(conv * 100.0 / cost, 1) if cost else 0
        if roas_max is not None and roas > roas_max:
            continue
        if roas_min is not None and roas < roas_min:
            continue
        a = attr.get(pno, {})
        lid = a.get('login_id', '')
        st = '삭제완료' if (lid, pno) in deleted else _gmkt_status_label(status_by_pno.get(pno), pno in status_by_pno)
        _sy = synced_by_pno.get(pno)
        # '판매중'인데 상태수집이 STALE_DAYS+ 지난 경우만 확인필요(중지/품절/삭제는 이미 확정상태라 제외)
        _stale = bool(st == '판매중' and _sy and (_now - _sy).days >= STALE_DAYS)
        rows.append({
            'login_id': lid, 'seller_name': name_map.get(lid, lid),
            'product_no': pno, 'seller_code': code_by_pno.get(pno, ''),
            'site': a.get('site', ''),
            'cost': cost, 'clicks': g['clicks'] or 0,
            'cum_sold_qty': cum_qty_by_pno.get(pno, 0),  # 누적 판매수량(2025~현재)
            'avg_click_cost': round(cost / (g['clicks'] or 1)) if (g['clicks'] or 0) else 0,  # 평균단가(광고비/클릭)
            'ad_orders': g['orders'] or 0, 'conv_amount': conv,   # 광고센터 구매수/구매금액
            'real_sales': real_by_pno.get(pno, 0), 'real_orders': realorders_by_pno.get(pno, 0),  # 매출자료(참고)
            'real_roas': round(real_by_pno.get(pno, 0) * 100.0 / cost, 1) if cost else 0,  # 실매출 효율(실매출÷광고비)
            'roas': roas, 'status': st,
            'status_stale': _stale,   # 판매중 표시가 오래된 스냅샷(확인필요)
            'status_synced_at': _sy.isoformat() if _sy else None,
            'keywords': kw_by_pno.get(pno, [])[:30],   # 수집된 키워드(광고비순)
        })
        if len(rows) >= CAP:
            break
    return {
        'count': len(rows), 'rows': rows, 'capped': len(rows) >= CAP,
        'ym_from': f'{months[0][0]}-{months[0][1]:02d}' if months else '',
        'ym_to': f'{months[-1][0]}-{months[-1][1]:02d}' if months else '',
        'cost_min': cost_min, 'roas_max': roas_max, 'roas_min': roas_min, 'clicks_min': clicks_min,
    }


class GmarketLossProductsView(views.APIView):
    """지마켓 상품목록(JSON) — 적자/ROAS200%↑/전체 모달 공통. params: ym_from, ym_to, eid?, cost_min, roas_max, roas_min, clicks_min."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(_gmkt_product_rows(request))


class GmarketLossMarkDeletedView(views.APIView):
    """지마켓 적자상품 삭제완료 표시 — body: {items:[{login_id,product_no,seller_code}]}."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from apps.cpc.models import GmarketLossDeleted
        items = request.data.get('items') or []
        marked = 0
        for it in items:
            lid = it.get('login_id') or ''
            pno = str(it.get('product_no') or '')
            if not lid or not pno:
                continue
            GmarketLossDeleted.objects.get_or_create(
                login_id=lid, product_no=pno,
                defaults={'seller_code': it.get('seller_code') or ''})
            marked += 1
        return Response({'status': 'ok', 'marked': marked,
                         'message': f'{marked}개 삭제완료 처리됨 (비고 표시)'})


class GmarketLossDeleteView(views.APIView):
    """지마켓 적자상품 자동 판매중지·삭제 (셀러오피스). 기본 dry-run(검증, 1상품).
    실삭제는 ?real=1 + 셀러오피스 삭제 플로우 검증 완료 후 활성화(VERIFIED)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import re, subprocess
        d = request.data
        ymf = (d.get('ym_from') or '2026-01').strip()
        ymt = (d.get('ym_to') or '2026-06').strip()
        if not re.match(r'^\d{4}-\d{2}$', ymf) or not re.match(r'^\d{4}-\d{2}$', ymt):
            return Response({'error': '기간(YYYY-MM) 형식 오류'}, status=400)
        real = str(d.get('real', '')).lower() in ('1', 'true', 'yes')
        # 안전장치: 지마켓 셀러오피스 삭제 플로우(전체선택·판매중지 드롭다운) 검증 전까지 실삭제 차단.
        VERIFIED = False
        if real and not VERIFIED:
            return Response({'status': 'blocked',
                             'message': '⚠️ 지마켓 실삭제는 아직 비활성화 상태입니다. 먼저 검증(dry-run)으로 셀러오피스 플로우를 확인한 뒤 활성화됩니다.'}, status=400)
        # ★ 상품번호 지정 삭제(나의상품 선택삭제)
        pnos = [re.sub(r'\D', '', str(p)) for p in (d.get('product_nos') or [])]
        pnos = [p for p in pnos if p]
        if pnos:
            eid = str(d.get('eid') or '').strip()
            if not eid or not re.match(r'^[A-Za-z0-9_]+$', eid):
                return Response({'error': '상품지정 삭제는 eid(계정) 필수'}, status=400)
            a = f'manage.py delete_loss_gmarket --eid {eid} --product-nos ' + ' '.join(pnos[:300])
            if real:
                a += ' --real'
            sc = (f'cd /home/rejoice888/Avengers/backend && /usr/bin/python3 {a} >> /tmp/delete_loss_gmarket.log 2>&1')
            try:
                subprocess.Popen(['bash', '-c', sc], start_new_session=True,
                                 stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception as e:
                return Response({'status': 'error', 'error': str(e)}, status=500)
            return Response({'status': 'started',
                             'message': f'🗑 지마켓 지정상품 {len(pnos)}개 {"실삭제" if real else "검증(dry-run)"} 시작 — 텔레그램/로그로 확인하세요.'})
        try:
            lim = int(d.get('limit')) if str(d.get('limit') or '').strip() else None
        except (ValueError, TypeError):
            lim = None
        limit = 1 if not real else lim
        eid = d.get('eid') or ''
        args = f'manage.py delete_loss_gmarket --ym-from {ymf} --ym-to {ymt}'
        if eid and re.match(r'^[A-Za-z0-9_]+$', str(eid)):
            args += f' --eid {eid}'
        if limit:
            args += f' --limit {limit}'
        if real:
            args += ' --real'
        script = (f'cd /home/rejoice888/Avengers/backend && /usr/bin/python3 {args} '
                  f'>> /tmp/delete_loss_gmarket.log 2>&1')
        try:
            subprocess.Popen(['bash', '-c', script], start_new_session=True,
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            return Response({'status': 'error', 'error': str(e)}, status=500)
        if real:
            scope = f'소량 {limit}개(테스트)' if limit else '전체'
            msg = f'🗑 지마켓 실삭제 시작 — {scope} 판매중지+삭제. 진행상황은 텔레그램/로그(/tmp/delete_loss_gmarket.log)로 확인하세요.'
        else:
            msg = '🔎 지마켓 검증(dry-run) 실행 — 1상품으로 셀러오피스 접속·입력·조회·셀렉터를 확인합니다(삭제 안 함).'
        return Response({'status': 'started', 'message': msg})


def _gmkt_keyword_launch(ym_from, ym_to, roas_min, eid, product_nos):
    """crawl_gmarket_keywords 관리명령을 별도 프로세스로 실행. 크롤러 내부 guard.preflight가 동시실행 차단."""
    import re, subprocess
    ymre = re.compile(r'^\d{4}-\d{2}$')
    if not ymre.match(ym_from or '') or not ymre.match(ym_to or ''):
        return False, '기간(YYYY-MM) 형식 오류'
    args = (f'manage.py crawl_gmarket_keywords --ym-from {ym_from} --ym-to {ym_to} '
            f'--roas-min {float(roas_min)}')
    if eid and re.match(r'^[A-Za-z0-9_]+$', str(eid)):
        args += f' --eid {eid}'
    pnos = [re.sub(r'\D', '', str(p)) for p in (product_nos or [])]
    pnos = [p for p in pnos if p][:2000]
    if pnos:
        args += ' --product-nos ' + ' '.join(pnos)
    script = (f'cd /home/rejoice888/Avengers/backend && /usr/bin/python3 {args} '
              f'>> /tmp/cron_gmkt_keywords.log 2>&1')
    try:
        subprocess.Popen(['bash', '-c', script], start_new_session=True,
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        return False, str(e)
    return True, ('지정 상품번호' if pnos else 'CPC ROAS≥200% 자동대상')


def _gmkt_keyword_launch_years(years, roas_min, eid):
    """연도-누적 버킷 모드(2025/2026/전체). 연도별 crawl_gmarket_keywords --year 를 순차 실행(&&).
    상품당 1회 범위조회라 월별 대비 빠름. 크롤러 preflight가 동시실행 차단."""
    import re, subprocess
    yrs = [int(y) for y in years if str(y).isdigit() and 2020 <= int(y) <= 2100]
    if not yrs:
        return False, '연도 오류'
    eid_arg = f' --eid {eid}' if (eid and re.match(r'^[A-Za-z0-9_]+$', str(eid))) else ''
    cmds = [f'/usr/bin/python3 manage.py crawl_gmarket_keywords --year {y} --roas-min {float(roas_min)}{eid_arg}'
            for y in sorted(set(yrs))]
    script = ('cd /home/rejoice888/Avengers/backend && ' + ' && '.join(cmds)
              + ' >> /tmp/cron_gmkt_keywords.log 2>&1')
    try:
        subprocess.Popen(['bash', '-c', script], start_new_session=True,
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        return False, str(e)
    return True, '·'.join(f'{y}년' for y in sorted(set(yrs)))


class GmarketKeywordCrawlView(views.APIView):
    """지마켓 CPC 키워드 수집 트리거. body: {ym_from, ym_to, eid?, roas_min=200, product_nos?[]}.
    product_nos 미지정시 해당 기간 CPC ROAS≥roas_min 상품을 자동 대상으로."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        d = request.data
        years = d.get('years')   # 연도-버킷 모드(2025/2026/전체)
        if years:
            ok, msg = _gmkt_keyword_launch_years(years, d.get('roas_min') or 200, d.get('eid') or '')
            if not ok:
                return Response({'status': 'error', 'error': msg}, status=400)
            return Response({'status': 'started', 'target': msg,
                             'message': f'연도 키워드 수집 시작({msg}) — 연 단위 누적, 상품당 1회. 완료 후 새로고침하면 표시됩니다.'})
        ok, msg = _gmkt_keyword_launch(
            (d.get('ym_from') or '').strip(), (d.get('ym_to') or '').strip(),
            d.get('roas_min') or 200, d.get('eid') or '', d.get('product_nos') or [])
        if not ok:
            return Response({'status': 'error', 'error': msg}, status=400)
        return Response({'status': 'started', 'target': msg,
                         'message': f'키워드 수집 시작({msg}) — 상품당 ~5초. 완료 후 새로고침하면 광고비 오른쪽에 키워드가 표시됩니다.'})


class GmarketKeywordUploadView(views.APIView):
    """엑셀(첫 열=상품번호) 업로드 → 해당 상품들 키워드 수집 트리거.
    multipart: file=xlsx, ym_from, ym_to, eid?(없으면 광고비데이터로 매핑)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'status': 'error', 'error': '파일 없음'}, status=400)
        try:
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
            pnos = []
            for row in wb.active.iter_rows(values_only=True):
                if not row:
                    continue
                import re as _re
                v = _re.sub(r'\D', '', str(row[0]) if row[0] is not None else '')
                if v:
                    pnos.append(v)
        except Exception as e:
            return Response({'status': 'error', 'error': f'엑셀 파싱 실패: {str(e)[:120]}'}, status=400)
        pnos = list(dict.fromkeys(pnos))   # 중복제거(순서유지)
        if not pnos:
            return Response({'status': 'error', 'error': '상품번호를 찾지 못했습니다(첫 열 확인)'}, status=400)
        ok, msg = _gmkt_keyword_launch(
            (request.data.get('ym_from') or '').strip(), (request.data.get('ym_to') or '').strip(),
            0, request.data.get('eid') or '', pnos)
        if not ok:
            return Response({'status': 'error', 'error': msg}, status=400)
        return Response({'status': 'started', 'count': len(pnos),
                         'message': f'엑셀 {len(pnos)}개 상품번호 키워드 수집 시작 — 완료 후 새로고침하세요.'})


class GmarketKeywordStatusView(views.APIView):
    """키워드 수집 진행상태 — 실행중 프로세스 여부 + 최근 수집시각."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import subprocess
        from apps.cpc.models import GmarketKeywordReport
        from django.db.models import Max
        try:
            running = subprocess.run(['pgrep', '-f', 'crawl_gmarket_keywords'],
                                     capture_output=True).returncode == 0
        except Exception:
            running = False
        latest = GmarketKeywordReport.objects.aggregate(m=Max('collected_at'))['m']
        return Response({'running': running, 'latest': latest,
                         'count': GmarketKeywordReport.objects.count()})


class GmarketKeywordCumulativeView(views.APIView):
    """수집 키워드 누계(중복제거) 조회. 연도버킷(month=0)만 합산해 월버킷 중복합산 방지.
    params: eid(계정), groupby=product(상품번호별, 기본) | seller(판매자코드별 합산).
    ROAS는 비가산이라 합산 후 재계산. 정렬은 프론트(상품번호→판매자코드→키워드 기본)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.cpc.models import GmarketKeywordReport, GmarketMyProduct
        from django.db.models import Sum
        eid = (request.query_params.get('eid') or '').strip()
        groupby = (request.query_params.get('groupby') or 'product').strip()

        qs = GmarketKeywordReport.objects.filter(month=0)   # 연도버킷만(중복합산 방지)
        if eid:
            qs = qs.filter(login_id=eid)

        # 상품번호→판매자코드 매핑(미매핑은 상품번호 그대로 표기)
        pnos = set(qs.values_list('product_no', flat=True))
        code_by_pno = {}
        for p in (GmarketMyProduct.objects.filter(product_no__in=pnos)
                  .values('product_no', 'seller_product_code')):
            code_by_pno.setdefault(p['product_no'], p['seller_product_code'] or '')

        # 1단계: (그룹키, 키워드) 누계 — 같은 키워드가 여러 버킷/상품에 있어도 합산(중복제거)
        kw_agg = {}
        for r in (qs.values('product_no', 'keyword')
                  .annotate(imp=Sum('impressions'), clk=Sum('clicks'),
                            cost=Sum('cost'), conv=Sum('conv_amount'), od=Sum('orders'))):
            pno = r['product_no']
            sc = code_by_pno.get(pno, '')
            gkey = (sc or f'(미매핑){pno}') if groupby == 'seller' else pno
            k = (gkey, r['keyword'])
            a = kw_agg.setdefault(k, {'gkey': gkey, 'pno': pno, 'sc': sc, 'kw': r['keyword'],
                                     'imp': 0, 'clk': 0, 'cost': 0, 'conv': 0, 'od': 0, 'pnos': set()})
            a['imp'] += r['imp'] or 0; a['clk'] += r['clk'] or 0; a['cost'] += r['cost'] or 0
            a['conv'] += r['conv'] or 0; a['od'] += r['od'] or 0; a['pnos'].add(pno)

        # 2단계: 그룹키(상품번호 or 판매자코드) 1행 — 키워드 모으기 + 지표 합산(코드 중복 없음)
        groups = {}
        for a in kw_agg.values():
            g = groups.setdefault(a['gkey'], {'product_no': a['pno'], 'seller_code': a['sc'],
                                              'imp': 0, 'clk': 0, 'cost': 0, 'conv': 0, 'od': 0,
                                              'kws': [], 'pnos': set()})
            kroas = round(a['conv'] * 100.0 / a['cost'], 1) if a['cost'] else 0
            kavg = round(a['cost'] / a['clk']) if a['clk'] else 0
            g['kws'].append({'keyword': a['kw'], 'cost': a['cost'], 'clicks': a['clk'],
                             'conv_amount': a['conv'], 'roas': kroas, 'avg_click_cost': kavg})
            g['imp'] += a['imp']; g['clk'] += a['clk']; g['cost'] += a['cost']
            g['conv'] += a['conv']; g['od'] += a['od']; g['pnos'] |= a['pnos']

        # 실매출(매출자료 전역매칭): product_no→판매자코드→SalesRecord 합. 기간=2025~현재(키워드 데이터 범위).
        from apps.sales.models import SalesRecord
        from django.utils import timezone as _tz
        from datetime import date as _date
        all_pnos = set()
        for g in groups.values():
            all_pnos |= g['pnos']
        codes = {code_by_pno.get(p) for p in all_pnos if code_by_pno.get(p)}
        sales_by_code = {}
        if codes:
            for r in (SalesRecord.objects
                      .filter(platform__in=['gmarket', 'auction'],
                              order_date__gte=_date(2025, 1, 1), order_date__lte=_tz.localdate(),
                              product_code__in=codes)
                      .values('product_code').annotate(s=Sum('total_price'))):
                sales_by_code[r['product_code']] = r['s'] or 0

        rows = []
        for g in groups.values():
            cost = g['cost']; conv = g['conv']; clk = g['clk']
            g['kws'].sort(key=lambda x: -x['cost'])
            gcodes = {code_by_pno.get(p) for p in g['pnos'] if code_by_pno.get(p)}
            real_sales = sum(sales_by_code.get(c, 0) for c in gcodes)   # 그룹 내 고유 판매자코드 매출 합(중복방지)
            rows.append({
                'product_no': g['product_no'] if groupby == 'product' else (f"{len(g['pnos'])}개 상품" if len(g['pnos']) > 1 else g['product_no']),
                'seller_code': g['seller_code'],
                'keyword_count': len(g['kws']),
                'impressions': g['imp'], 'clicks': clk, 'cost': cost,
                'avg_click_cost': round(cost / clk) if clk else 0,   # 평균광고비(=광고비합계/클릭합계)
                'orders': g['od'], 'conv_amount': conv,
                'roas': round(conv * 100.0 / cost, 1) if cost else 0,
                'real_sales': real_sales,                              # 실매출(매출자료)
                'real_roas': round(real_sales * 100.0 / cost, 1) if cost else 0,
                'keywords': g['kws'][:30], 'product_count': len(g['pnos']),
            })
        rows.sort(key=lambda x: (str(x['product_no']), str(x['seller_code'])))
        return Response({'count': len(rows), 'rows': rows, 'eid': eid, 'groupby': groupby})


class GmarketFocusTargetsView(views.APIView):
    """집중대상 리스트 — 실매출·고ROAS 결합. mode=focus(실매출>0 & 광고전환ROAS≥200) | hidden(잘팔리는데 광고적음).
    params: eid(계정), mode. 광고데이터=2026 CPC, 실매출=2026 판매자코드 전역매칭."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.cpc.models import GmarketProductAdCost, GmarketMyProduct
        from apps.sales.models import SalesRecord
        from django.db.models import Sum
        from django.utils import timezone as _tz
        from datetime import date as _date
        eid = (request.query_params.get('eid') or '').strip()
        mode = (request.query_params.get('mode') or 'focus').strip()
        ad_type = (request.query_params.get('ad_type') or 'cpc').strip()  # cpc/ai/'' (전체합산)
        d0 = _date(2026, 1, 1); d1 = _tz.localdate()

        base = GmarketProductAdCost.objects.filter(year=2026)
        if ad_type in ('cpc', 'ai'):
            base = base.filter(ad_type=ad_type)
        if eid:
            base = base.filter(login_id=eid)
        adp = {}
        for r in base.values('product_no', 'login_id').annotate(c=Sum('cost'), v=Sum('conv_amount'), clk=Sum('clicks')):
            p = r['product_no']
            a = adp.setdefault(p, {'c': 0, 'v': 0, 'clk': 0, 'lid': r['login_id']})
            a['c'] += r['c'] or 0; a['v'] += r['v'] or 0; a['clk'] += r['clk'] or 0
        site_by_pno = {}
        for r in base.values('product_no', 'site'):
            if r.get('site') and r['product_no'] not in site_by_pno:
                site_by_pno[r['product_no']] = r['site']
        meta = {p['product_no']: (p['seller_product_code'] or '', p['product_name'] or '')
                for p in GmarketMyProduct.objects.filter(product_no__in=list(adp)).values('product_no', 'seller_product_code', 'product_name')}
        codes = {meta.get(p, ('', ''))[0] for p in adp if meta.get(p, ('', ''))[0]}
        sbc = {}
        if codes:
            for r in (SalesRecord.objects
                      .filter(platform__in=['gmarket', 'auction'], order_date__gte=d0, order_date__lte=d1, product_code__in=codes)
                      .values('product_code').annotate(s=Sum('total_price'))):
                sbc[r['product_code']] = r['s'] or 0
        rows = []
        for p, a in adp.items():
            sc, nm = meta.get(p, ('', ''))
            real = sbc.get(sc, 0)
            rows.append({'product_no': p, 'login_id': a['lid'], 'seller_code': sc, 'product_name': nm[:40],
                         'site': site_by_pno.get(p, ''),
                         'cost': a['c'], 'conv_roas': round(a['v'] * 100.0 / a['c'], 1) if a['c'] else 0,
                         'real_sales': real, 'real_roas': round(real * 100.0 / a['c'], 1) if a['c'] else 0,
                         'clicks': a['clk']})
        if mode == 'hidden':
            rows = [r for r in rows if r['real_sales'] >= 100000 and r['cost'] < 3000]
        else:
            rows = [r for r in rows if r['real_sales'] > 0 and r['conv_roas'] >= 200]
        rows.sort(key=lambda x: -x['real_sales'])
        rows = rows[:3000]
        # 키워드 첨부 — 최종 대상 상품번호의 수집 키워드(상품번호+키워드 합산)
        from apps.cpc.models import GmarketKeywordReport
        from collections import defaultdict as _dd
        fpnos = [r['product_no'] for r in rows]
        kagg = _dd(lambda: {'cost': 0, 'clicks': 0, 'conv': 0})
        for k in (GmarketKeywordReport.objects.filter(product_no__in=fpnos, year=2026)
                  .values('product_no', 'keyword', 'cost', 'clicks', 'conv_amount')):
            x = kagg[(k['product_no'], k['keyword'])]
            x['cost'] += k['cost'] or 0; x['clicks'] += k['clicks'] or 0; x['conv'] += k['conv_amount'] or 0
        kw_by = _dd(list)
        for (pno, kw), x in kagg.items():
            kw_by[pno].append({'keyword': kw, 'cost': x['cost'], 'clicks': x['clicks'], 'conv_amount': x['conv'],
                               'roas': round(x['conv'] * 100.0 / x['cost'], 1) if x['cost'] else 0})
        for pno in kw_by:
            kw_by[pno].sort(key=lambda y: -y['cost'])
        for r in rows:
            r['keywords'] = kw_by.get(r['product_no'], [])[:30]
        return Response({'count': len(rows), 'rows': rows, 'mode': mode, 'eid': eid})
