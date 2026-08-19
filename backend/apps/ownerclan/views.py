import io

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse

from . import services
from .models import OwnerclanTask


class _WorkspaceMixin:
    """요청의 ?workspace= 값으로 대상 테이블(예비상품/상품가공)을 전환. 기본 reserve."""
    def initial(self, request, *args, **kwargs):
        services.set_workspace(request.query_params.get('workspace') or 'reserve')
        super().initial(request, *args, **kwargs)


class OwnerClanProductUploadView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)

        running = OwnerclanTask.objects.filter(
            task_type='ownerclan_upload', status__in=('pending', 'running')
        ).first()
        if running:
            return Response({
                'error': '이미 업로드 처리 중입니다.',
                'task_id': running.id,
            }, status=409)

        try:
            result = services.upload_excel_async(f, workspace=request.query_params.get('workspace') or 'reserve')
            return Response(result, status=202)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=400)
        try:
            task = OwnerclanTask.objects.get(pk=int(task_id))
        except OwnerclanTask.DoesNotExist:
            return Response({'error': 'not found'}, status=404)
        return Response({
            'task_id': task.id,
            'status': task.status,
            'result_data': task.result_data,
        })


class OwnerClanProductCsvUploadView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        try:
            result = services.upload_csv_status(f)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class OwnerClanSoldoutTxtUploadView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        try:
            result = services.upload_soldout_txt(f)
            return Response(result)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class OwnerClanProductListView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None
        sort = request.query_params.get('sort') or None
        order = request.query_params.get('order') or 'asc'
        filter_col = request.query_params.get('filter_col') or None
        filter_vals_raw = request.query_params.get('filter_vals') or ''
        filter_vals = [v for v in filter_vals_raw.split('|') if v != ''] if filter_vals_raw else None
        codes_raw = request.query_params.get('codes') or ''
        codes = [c.strip() for c in codes_raw.split(',') if c.strip()] if codes_raw else None
        result = services.get_products(
            page, per_page,
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
            sort=sort,
            order=order,
            filter_col=filter_col,
            filter_vals=filter_vals,
            codes=codes,
        )
        return Response(result)


class OwnerClanProductDetailView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        result = services.get_product_detail(pk)
        if not result:
            return Response({'error': '상품을 찾을 수 없습니다.'}, status=404)
        return Response(result)


class OwnerClanProductSyncView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        product_ids = request.data.get('product_ids')
        if product_ids and isinstance(product_ids, list):
            product_ids = [int(i) for i in product_ids]
        else:
            product_ids = None
        result = services.sync_products(product_ids)
        return Response(result)


class OwnerClanProductStatsView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(services.get_stats())


class OwnerClanProductChangedFieldsView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(services.get_changed_field_counts())


class OwnerClanProductExcelExportView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None

        rows = services.get_products_for_export(
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )

        STATUS_LABELS = {1: '판매중', 2: '품절', 3: '단종'}

        wb = Workbook()
        ws = wb.active
        ws.title = '오너클랜 상품대장'

        headers = ['W코드', '상태', '동기화', '상품명', '원본상품명',
                    '마켓상품명', '원본마켓상품명', '오너클랜가', '원본오너클랜가',
                    '마켓가', '원본마켓가', '배송비', '원본배송비',
                    '반품비', '원본반품비', '카테고리', '제조사', '원산지']
        col_widths = [12, 8, 8, 35, 35, 35, 35, 12, 12, 12, 12, 8, 8, 8, 8, 20, 15, 10]

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='F0F0F0')
        changed_fill = PatternFill('solid', fgColor='FFF3E0')
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
        money_fmt = '#,##0'

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        for i, r in enumerate(rows, 2):
            is_changed = r.get('is_synced') == 0
            fill = changed_fill if is_changed else None

            def _cell(col, val, fmt=None):
                c = ws.cell(row=i, column=col, value=val)
                c.border = thin_border
                if fmt:
                    c.number_format = fmt
                if fill:
                    c.fill = fill
                return c

            _cell(1, r.get('product_code'))
            _cell(2, STATUS_LABELS.get(r.get('sale_status'), '?'))
            _cell(3, '변경됨' if is_changed else '일치')
            _cell(4, r.get('product_name'))
            _cell(5, r.get('orig_product_name'))
            _cell(6, r.get('market_product_name'))
            _cell(7, r.get('orig_market_product_name'))
            _cell(8, r.get('ownerclan_price', 0), money_fmt)
            _cell(9, r.get('orig_ownerclan_price', 0), money_fmt)
            _cell(10, r.get('market_price', 0), money_fmt)
            _cell(11, r.get('orig_market_price', 0), money_fmt)
            _cell(12, r.get('shipping_fee', 0), money_fmt)
            _cell(13, r.get('orig_shipping_fee', 0), money_fmt)
            _cell(14, r.get('return_fee', 0), money_fmt)
            _cell(15, r.get('orig_return_fee', 0), money_fmt)
            _cell(16, r.get('category_name'))
            _cell(17, r.get('manufacturer'))
            _cell(18, r.get('origin'))

        ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ownerclan_products.xlsx"'
        return response


class OwnerClanProductDbExportView(_WorkspaceMixin, APIView):
    """오너클랜 상품 DB 전체 다운로드(CSV) — 필터 없이 전 건수.
    orig_*(원본대조용 내부컬럼)와 거대 HTML/공지 텍스트 필드는 파일이 열기 힘들 정도로 커져서 제외."""
    permission_classes = [IsAuthenticated]

    FIELDS = [
        'product_code', 'sale_status', 'is_synced',
        'seller_code1', 'seller_code2', 'category_code', 'category_name', 'market_category',
        'product_name', 'market_product_name',
        'ownerclan_price', 'consumer_price', 'market_price', 'shipping_fee', 'shipping_type',
        'min_qty', 'max_qty', 'return_fee', 'return_possible',
        'option1_name', 'option1_values', 'option2_name', 'option2_values',
        'combined_option', 'combined_option_detail', 'independent_option',
        'product_attribute', 'product_grade', 'tax_type', 'compliance', 'age_restriction',
        'manufacturer', 'brand', 'model_name', 'origin', 'keywords',
        'image_large', 'notice_code', 'notice_category',
        'market_gmarket', 'market_auction', 'market_11st', 'market_coupang',
        'market_smartstore', 'market_promo', 'market_gift',
        'certification_type', 'certification_info',
        'registered_at', 'modified_at', 'uploaded_at', 'synced_at',
    ]
    HEADERS = [
        'W코드', '판매상태', '동기화',
        '판매자코드1', '판매자코드2', '카테고리코드', '카테고리명', '마켓카테고리',
        '상품명', '마켓상품명',
        '오너클랜가', '소비자가', '마켓가', '배송비', '배송타입',
        '최소수량', '최대수량', '반품비', '반품가능',
        '옵션1명', '옵션1값', '옵션2명', '옵션2값',
        '조합옵션', '조합옵션상세', '독립옵션',
        '속성', '등급', '과세유형', '준수사항', '연령제한',
        '제조사', '브랜드', '모델명', '원산지', '키워드',
        '대표이미지', '고시코드', '고시분류',
        '지마켓', '옥션', '11번가', '쿠팡', '스마트스토어', '프로모션', '사은품',
        '인증유형', '인증정보',
        '등록일', '수정일', '업로드일', '동기화일',
    ]

    def get(self, request):
        import csv
        from django.http import StreamingHttpResponse
        from .models import OwnerclanProduct, ProcessingProduct

        is_processing = services._t() == 'processing_product'
        model = ProcessingProduct if is_processing else OwnerclanProduct

        def _stream():
            buf = io.StringIO()
            w = csv.writer(buf)
            w.writerow(self.HEADERS)
            yield ('﻿' + buf.getvalue()).encode('utf-8')
            qs = model.objects.order_by('id').values_list(*self.FIELDS).iterator(chunk_size=2000)
            for row in qs:
                buf = io.StringIO()
                w = csv.writer(buf)
                w.writerow(['' if v is None else v for v in row])
                yield buf.getvalue().encode('utf-8')

        fname = f'ownerclan_db_{"processing" if is_processing else "reserve"}.csv'
        response = StreamingHttpResponse(_stream(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response


class OwnerclanApiCrawlView(APIView):
    """오너클랜 정식 API 신규상품 수집 — 백그라운드 실행 + 상태 폴링. /blog(오너클랜크롤러 메뉴) 페이지에서 사용."""
    permission_classes = [IsAuthenticated]
    LOG_FILE = '/tmp/ownerclan_api_crawl.log'

    def get(self, request):
        import os
        from .models import OwnerclanApiAccount
        task = OwnerclanTask.objects.filter(task_type='api_crawl').order_by('-created_at').first()
        busy = False
        if task and task.status == 'running' and task.pid:
            try:
                os.kill(task.pid, 0)
                busy = True
            except (ProcessLookupError, PermissionError):
                task.status = 'done'
                task.save(update_fields=['status'])

        log_tail = ''
        try:
            with open(self.LOG_FILE, encoding='utf-8', errors='ignore') as f:
                log_tail = ''.join(f.readlines()[-40:])
        except FileNotFoundError:
            pass

        accounts = [{'login_id': a.login_id, 'last_synced_at': a.last_synced_at,
                     'last_new_count': a.last_new_count,
                     'balance': a.balance, 'order_stats': a.order_stats,
                     'subscription_info': a.subscription_info,
                     'lowest_price_quota': a.lowest_price_quota,
                     'info_synced_at': a.info_synced_at}
                    for a in OwnerclanApiAccount.objects.filter(is_active=True)]
        return Response({'busy': busy, 'log': log_tail, 'accounts': accounts})

    def post(self, request):
        import subprocess
        running = OwnerclanTask.objects.filter(task_type='api_crawl', status='running').first()
        if running and running.pid:
            import os
            try:
                os.kill(running.pid, 0)
                return Response({'error': '이미 수집 중입니다.'}, status=409)
            except (ProcessLookupError, PermissionError):
                pass

        task = OwnerclanTask.objects.create(task_type='api_crawl', status='running')
        cmd = (f'cd /home/rejoice888/Avengers/backend && '
               f'python3 manage.py crawl_ownerclan_api > {self.LOG_FILE} 2>&1')
        proc = subprocess.Popen(['bash', '-c', cmd], start_new_session=True)
        task.pid = proc.pid
        task.save(update_fields=['pid'])
        return Response({'status': 'started', 'task_id': task.id})


class OwnerclanWeeklyPopularView(APIView):
    """오너클랜 '주간 인기 상품' 다운로드(db저장창고) — 파일 목록 조회/수동 실행.
    매일 09:00 크론(cron_ownerclan_weekly_popular.sh)으로도 자동 저장됨."""
    permission_classes = [IsAuthenticated]
    LOG_FILE = '/tmp/cron_ownerclan_weekly.log'

    def get(self, request):
        import os
        from django.conf import settings
        storage_dir = os.path.join(settings.BASE_DIR, 'media', 'ownerclan_weekly_popular')
        files = []
        if os.path.isdir(storage_dir):
            for name in os.listdir(storage_dir):
                path = os.path.join(storage_dir, name)
                if os.path.isfile(path):
                    stat = os.stat(path)
                    files.append({
                        'filename': name,
                        'size': stat.st_size,
                        'saved_at': stat.st_mtime,
                    })
        files.sort(key=lambda f: f['saved_at'], reverse=True)

        task = OwnerclanTask.objects.filter(task_type='weekly_popular').order_by('-created_at').first()
        busy = False
        if task and task.status == 'running' and task.pid:
            try:
                os.kill(task.pid, 0)
                busy = True
            except (ProcessLookupError, PermissionError):
                task.status = 'done'
                task.save(update_fields=['status'])

        return Response({'files': files, 'storage_dir': storage_dir, 'busy': busy})

    def post(self, request):
        import os
        import subprocess
        running = OwnerclanTask.objects.filter(task_type='weekly_popular', status='running').first()
        if running and running.pid:
            try:
                os.kill(running.pid, 0)
                return Response({'error': '이미 수집 중입니다.'}, status=409)
            except (ProcessLookupError, PermissionError):
                pass

        task = OwnerclanTask.objects.create(task_type='weekly_popular', status='running')
        cmd = (f'cd /home/rejoice888/Avengers/backend && '
               f'python3 manage.py crawl_ownerclan_weekly_popular > {self.LOG_FILE} 2>&1; '
               f'echo DONE >> {self.LOG_FILE}')
        proc = subprocess.Popen(['bash', '-c', cmd], start_new_session=True)
        task.pid = proc.pid
        task.status = 'running'
        task.save(update_fields=['pid', 'status'])
        return Response({'status': 'started', 'task_id': task.id})


class OwnerclanWeeklyPopularDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import os
        from django.conf import settings
        from django.http import FileResponse, Http404

        filename = request.query_params.get('filename', '')
        # 경로 조작 방지 — 순수 파일명만 허용
        if not filename or os.path.basename(filename) != filename:
            raise Http404()
        storage_dir = os.path.join(settings.BASE_DIR, 'media', 'ownerclan_weekly_popular')
        path = os.path.join(storage_dir, filename)
        if not os.path.isfile(path):
            raise Http404()
        return FileResponse(open(path, 'rb'), as_attachment=True, filename=filename)


class OwnerclanAccountInfoCrawlView(APIView):
    """오너클랜 마이페이지 계정정보(예치금/주문현황/구독서비스/최저가선점권) 새로고침 — 백그라운드 실행."""
    permission_classes = [IsAuthenticated]
    LOG_FILE = '/tmp/ownerclan_account_info_crawl.log'

    def post(self, request):
        import os
        import subprocess
        running = OwnerclanTask.objects.filter(task_type='account_info', status='running').first()
        if running and running.pid:
            try:
                os.kill(running.pid, 0)
                return Response({'error': '이미 수집 중입니다.'}, status=409)
            except (ProcessLookupError, PermissionError):
                pass

        task = OwnerclanTask.objects.create(task_type='account_info', status='running')
        cmd = (f'cd /home/rejoice888/Avengers/backend && '
               f'python3 manage.py crawl_ownerclan_account_info > {self.LOG_FILE} 2>&1')
        proc = subprocess.Popen(['bash', '-c', cmd], start_new_session=True)
        task.pid = proc.pid
        task.save(update_fields=['pid'])
        return Response({'status': 'started', 'task_id': task.id})


class OwnerClanProductWCodesView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sale_status = request.query_params.get('sale_status')
        is_synced = request.query_params.get('is_synced')
        search = request.query_params.get('search') or None
        changed_field = request.query_params.get('changed_field') or None

        codes = services.get_w_codes(
            sale_status=int(sale_status) if sale_status else None,
            is_synced=int(is_synced) if is_synced is not None and is_synced != '' else None,
            search=search,
            changed_field=changed_field,
        )
        return Response({'codes': codes, 'count': len(codes)})


class OwnerClanProductDeleteAllView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        confirm = request.data.get('confirm')
        if confirm != 'DELETE_ALL':
            return Response({'error': "확인 토큰 누락 (confirm='DELETE_ALL' 필요)"}, status=400)
        result = services.delete_all_products()
        return Response(result)


class OwnerClanProductDeleteByIdsView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'error': '삭제할 id 리스트 필요'}, status=400)
        result = services.delete_products_by_ids(ids)
        return Response(result)


class OwnerClanProductDedupeView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        result = services.dedupe_by_product_name()
        return Response(result)


class OwnerClanApplyElevenNameView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        items = request.data.get('items') or []
        if not isinstance(items, list) or not items:
            return Response({'error': 'items 배열 필요 ([{code,name}])'}, status=400)
        return Response(services.apply_eleven_names(items))


class OwnerClanDistinctValuesView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        column = request.query_params.get('column')
        if not column:
            return Response({'error': 'column 파라미터 필요'}, status=400)
        try:
            values = services.get_distinct_values(column)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        return Response({'column': column, 'values': values})


class MyProductCopyView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        codes = request.data.get('source_product_codes') or []
        if not isinstance(codes, list) or not codes:
            return Response({'error': 'source_product_codes 배열 필요'}, status=400)
        result = services.copy_to_my_product(codes)
        return Response(result)


class MyProductListView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        page = int(request.query_params.get('page', 1))
        per_page = int(request.query_params.get('per_page', 50))
        search = request.query_params.get('search') or None
        is_modified = request.query_params.get('is_modified')
        sort = request.query_params.get('sort') or None
        order = request.query_params.get('order') or 'asc'
        filter_col = request.query_params.get('filter_col') or None
        filter_vals_raw = request.query_params.get('filter_vals') or ''
        filter_vals = [v for v in filter_vals_raw.split('|') if v != ''] if filter_vals_raw else None
        codes_raw = request.query_params.get('codes') or ''
        codes = [c.strip() for c in codes_raw.split(',') if c.strip()] if codes_raw else None
        result = services.get_my_products(
            page, per_page, search=search,
            is_modified=int(is_modified) if is_modified is not None and is_modified != '' else None,
            sort=sort, order=order,
            filter_col=filter_col, filter_vals=filter_vals, codes=codes,
        )
        return Response(result)


class MyProductDetailView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        result = services.get_my_product_detail(pk)
        if not result:
            return Response({'error': '나의 상품을 찾을 수 없습니다.'}, status=404)
        return Response(result)

    def patch(self, request, pk):
        fields_dict = request.data or {}
        result = services.update_my_product(pk, fields_dict)
        return Response(result)

    def delete(self, request, pk):
        result = services.delete_my_products_by_ids([pk])
        return Response(result)


class MyProductDeleteAllView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        confirm = request.data.get('confirm')
        if confirm != 'DELETE_ALL':
            return Response({'error': "확인 토큰 누락 (confirm='DELETE_ALL' 필요)"}, status=400)
        result = services.delete_all_my_products()
        return Response(result)


class MyProductDeleteByIdsView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'error': '삭제할 id 리스트 필요'}, status=400)
        result = services.delete_my_products_by_ids(ids)
        return Response(result)


class MyProductDedupeView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        result = services.dedupe_my_by_product_name()
        return Response(result)


class MyProductDistinctValuesView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        column = request.query_params.get('column')
        if not column:
            return Response({'error': 'column 파라미터 필요'}, status=400)
        try:
            values = services.get_my_distinct_values(column)
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        return Response({'column': column, 'values': values})


class MyProductWCodesView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = request.query_params.get('search') or None
        is_modified = request.query_params.get('is_modified')
        filter_col = request.query_params.get('filter_col') or None
        filter_vals_raw = request.query_params.get('filter_vals') or ''
        filter_vals = [v for v in filter_vals_raw.split('|') if v != ''] if filter_vals_raw else None
        codes = services.get_my_w_codes(
            search=search,
            is_modified=int(is_modified) if is_modified is not None and is_modified != '' else None,
            filter_col=filter_col, filter_vals=filter_vals,
        )
        return Response({'codes': codes, 'count': len(codes)})


class MyProductExcelExportView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        search = request.query_params.get('search') or None
        is_modified = request.query_params.get('is_modified')
        filter_col = request.query_params.get('filter_col') or None
        filter_vals_raw = request.query_params.get('filter_vals') or ''
        filter_vals = [v for v in filter_vals_raw.split('|') if v != ''] if filter_vals_raw else None

        rows = services.get_my_products_for_export(
            search=search,
            is_modified=int(is_modified) if is_modified is not None and is_modified != '' else None,
            filter_col=filter_col, filter_vals=filter_vals,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = '나의 상품'

        headers = ['나의W코드', '원본W코드', '수정', '상품명', '마켓상품명',
                    '오너클랜가', '마켓가', '배송비', '반품비', '카테고리', '제조사', '원산지', '복사일']
        col_widths = [16, 14, 8, 35, 35, 12, 12, 8, 8, 20, 15, 10, 18]

        header_font = Font(bold=True, size=10)
        header_fill = PatternFill('solid', fgColor='F0F0F0')
        modified_fill = PatternFill('solid', fgColor='E8F5E9')
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
        money_fmt = '#,##0'

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        for i, r in enumerate(rows, 2):
            mod = bool(r.get('is_modified'))
            fill = modified_fill if mod else None

            def _cell(col, val, fmt=None):
                c = ws.cell(row=i, column=col, value=val)
                c.border = thin_border
                if fmt:
                    c.number_format = fmt
                if fill:
                    c.fill = fill
                return c

            _cell(1, r.get('my_product_code'))
            _cell(2, r.get('source_product_code'))
            _cell(3, '수정됨' if mod else '원본')
            _cell(4, r.get('product_name'))
            _cell(5, r.get('market_product_name'))
            _cell(6, r.get('ownerclan_price', 0), money_fmt)
            _cell(7, r.get('market_price', 0), money_fmt)
            _cell(8, r.get('shipping_fee', 0), money_fmt)
            _cell(9, r.get('return_fee', 0), money_fmt)
            _cell(10, r.get('category_name'))
            _cell(11, r.get('manufacturer'))
            _cell(12, r.get('origin'))
            cd = r.get('copied_at')
            _cell(13, cd.strftime('%Y-%m-%d %H:%M') if cd else '')

        ws.auto_filter.ref = ws.dimensions

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="my_products.xlsx"'
        return response


class MyProductUploadView(_WorkspaceMixin, APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'file required'}, status=400)
        running = OwnerclanTask.objects.filter(
            task_type='my_product_upload', status__in=('pending', 'running')
        ).first()
        if running:
            return Response({'error': '이미 업로드 처리 중입니다.', 'task_id': running.id}, status=409)
        try:
            result = services.upload_my_excel_async(f)
            return Response(result, status=202)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def get(self, request):
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({'error': 'task_id required'}, status=400)
        try:
            task = OwnerclanTask.objects.get(pk=int(task_id))
        except OwnerclanTask.DoesNotExist:
            return Response({'error': 'not found'}, status=404)
        return Response({
            'task_id': task.id,
            'status': task.status,
            'result_data': task.result_data,
        })
