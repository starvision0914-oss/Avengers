import io
import os
import traceback
import zipfile
from datetime import datetime

import openpyxl
from django.core.management.base import BaseCommand
from django.db import connections

from apps.ownerclan.models import OwnerclanTask
from apps.ownerclan.services import (
    EXCEL_COL_MAP, INT_FIELDS, DATETIME_FIELDS,
    _parse_excel_row, _safe_str, _field_changed,
    set_workspace, _t,
)


class Command(BaseCommand):
    help = '오너클랜 상품대장 비동기 업로드 워커'

    def add_arguments(self, parser):
        parser.add_argument('task_id', type=int)

    def handle(self, *args, **options):
        task_id = options['task_id']
        try:
            task = OwnerclanTask.objects.get(pk=task_id)
        except OwnerclanTask.DoesNotExist:
            return

        task.status = 'running'
        task.pid = os.getpid()
        task.save(update_fields=['status', 'pid'])

        set_workspace(task.input_data.get('workspace', 'reserve'))
        file_path = task.input_data.get('file_path', '')

        try:
            result = _process_upload(file_path, task)
            task.result_data = result
            task.status = 'done'
            # 예비상품 적재 직후 11번가/지마켓 나의상품 '구매원가/차이' 정렬값 자동 갱신 (업로드된 코드만, 빠름)
            try:
                from apps.cpc.eleven_my_product_service import refresh_purchase_costs, refresh_gmarket_purchase_costs
                n = refresh_purchase_costs(codes=result.get('codes'))
                result['purchase_cost_refreshed'] = n
                gn = refresh_gmarket_purchase_costs(codes=result.get('codes'))
                result['gmarket_purchase_cost_refreshed'] = gn
            except Exception:
                result['purchase_cost_refreshed'] = f'skip: {traceback.format_exc().splitlines()[-1]}'
        except Exception:
            task.result_data = {'error': traceback.format_exc()}
            task.status = 'error'
        finally:
            task.save(update_fields=['status', 'result_data'])
            try:
                os.unlink(file_path)
            except Exception:
                pass


def _is_playauto(file_path):
    """PlayAuto '11번가 등록현황'(.xls / zip내 .xls) 여부 판별."""
    name = file_path.lower()
    if name.endswith('.xls'):
        return True
    if name.endswith('.zip'):
        with open(file_path, 'rb') as f:
            zf = zipfile.ZipFile(io.BytesIO(f.read()))
        names = zf.namelist()
        has_xls = any(n.lower().endswith('.xls') for n in names)
        has_xlsx = any(n.lower().endswith(('.xlsx', '.xlsm')) for n in names)
        return has_xls and not has_xlsx
    return False


def _load_workbooks(file_path):
    name = file_path.lower()
    if name.endswith('.zip'):
        with open(file_path, 'rb') as f:
            zf = zipfile.ZipFile(io.BytesIO(f.read()))
        all_names = zf.namelist()
        xlsx_names = sorted(n for n in all_names if n.lower().endswith(('.xlsx', '.xlsm')))
        if not xlsx_names:
            preview = ', '.join(all_names[:10]) or '(empty)'
            raise ValueError(
                f'ZIP 안에 .xlsx/.xlsm 파일이 없습니다. '
                f'발견된 파일({len(all_names)}개): {preview}'
            )
        wbs = []
        for xn in xlsx_names:
            xlsx_bytes = zf.read(xn)
            wbs.append((xn, openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)))
        return wbs
    elif name.endswith(('.xlsx', '.xlsm')):
        return [(os.path.basename(file_path), openpyxl.load_workbook(file_path, read_only=True))]
    else:
        raise ValueError(
            f'xlsx 또는 zip 파일만 업로드 가능합니다. (받은 파일: {os.path.basename(file_path)})'
        )


def _process_upload(file_path, task):
    # PlayAuto(.xls/zip) = 값 가공 없이 있는 그대로 적재 (별도 매핑·xlrd 경로)
    if _is_playauto(file_path):
        from apps.ownerclan import services
        return services.ingest_playauto(
            path=file_path,
            log_fn=lambda m: None,
        )
    workbooks = _load_workbooks(file_path)
    now = datetime.now()

    fields = list(EXCEL_COL_MAP.values())
    orig_fields = [f'orig_{f}' for f in fields]

    rows_to_process = []
    for wb_name, wb in workbooks:
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 3:
                continue
            product_code = _safe_str(row[2])
            if not product_code:
                continue
            data = _parse_excel_row(list(row))
            rows_to_process.append((product_code, data))
        wb.close()

    # 같은 파일 안에서 동일 product_code가 중복되면 마지막 값으로 통일(= 순차처리 최종상태와 동일 결과).
    dedup = {}
    for pc, data in rows_to_process:
        dedup[pc] = data
    codes = list(dedup.keys())
    total_rows = len(rows_to_process)

    task.result_data = {
        'progress': 0, 'inserted': 0, 'updated': 0,
        'skipped': 0, 'total_rows': total_rows,
    }
    task.save(update_fields=['result_data'])

    # 1) 기존 행(현재값+원본값)을 청크 단위 IN 조회로 일괄 로딩 — 건당 SELECT 왕복 제거.
    select_cols = ['id', 'product_code'] + fields + orig_fields
    existing = {}
    with connections['default'].cursor() as cur:
        FETCH_CHUNK = 3000
        for i in range(0, len(codes), FETCH_CHUNK):
            chunk = codes[i:i + FETCH_CHUNK]
            ph = ', '.join(['%s'] * len(chunk))
            cur.execute(f"SELECT {', '.join(select_cols)} FROM {_t()} WHERE product_code IN ({ph})", chunk)
            for row in cur.fetchall():
                d = dict(zip(select_cols, row))
                existing[d['product_code']] = {
                    'id': d['id'],
                    'cur': {f: d[f] for f in fields},
                    'orig': {f: d[f'orig_{f}'] for f in fields},
                }

    # 2) 삽입/갱신 대상 분류 (파이썬 메모리 비교, DB 왕복 없음)
    insert_batch = []   # (product_code, data)
    update_batch = []   # (data, is_synced, id)
    inserted = updated = skipped = 0
    for pc in codes:
        data = dedup[pc]
        ex = existing.get(pc)
        if ex is None:
            insert_batch.append((pc, data))
            continue
        old_data = ex['cur']
        if not any(_field_changed(old_data[f], data[f], f) for f in fields):
            skipped += 1
            continue
        orig_data = ex['orig']
        is_synced = 0 if any(_field_changed(orig_data[f], data[f], f) for f in fields) else 1
        update_batch.append((data, is_synced, ex['id']))

    def _save_progress(done):
        progress = int(done * 100 / total_rows) if total_rows else 100
        task.result_data = {
            'progress': progress, 'inserted': inserted, 'updated': updated,
            'skipped': skipped, 'total_rows': total_rows,
        }
        task.save(update_fields=['result_data'])

    WRITE_CHUNK = 2000
    with connections['default'].cursor() as cur:
        # INSERT — executemany로 일괄 처리
        insert_sql = (
            f"INSERT INTO {_t()} (product_code, {', '.join(fields)}, {', '.join(orig_fields)}, "
            f"sale_status, is_synced, uploaded_at) "
            f"VALUES ({', '.join(['%s'] * (1 + len(fields) * 2 + 3))})"
        )
        for i in range(0, len(insert_batch), WRITE_CHUNK):
            chunk = insert_batch[i:i + WRITE_CHUNK]
            params = [
                [pc] + [data[f] for f in fields] + [data[f] for f in fields] + [1, 1, now]
                for pc, data in chunk
            ]
            cur.executemany(insert_sql, params)
            inserted += len(chunk)
            _save_progress(inserted + updated + skipped)

        # UPDATE — 모든 변경행이 동일한 SET 구조(전체필드+is_synced+uploaded_at)라 executemany 가능.
        update_sql = (
            f"UPDATE {_t()} SET {', '.join(f'{f}=%s' for f in fields)}, "
            f"is_synced=%s, uploaded_at=%s WHERE id=%s"
        )
        for i in range(0, len(update_batch), WRITE_CHUNK):
            chunk = update_batch[i:i + WRITE_CHUNK]
            params = [
                [data[f] for f in fields] + [is_synced, now, pid]
                for data, is_synced, pid in chunk
            ]
            cur.executemany(update_sql, params)
            updated += len(chunk)
            _save_progress(inserted + updated + skipped)

    _save_progress(total_rows)

    return {
        'inserted': inserted,
        'updated': updated,
        'skipped': skipped,
        'total': inserted + updated + skipped,
        'codes': codes,
    }
