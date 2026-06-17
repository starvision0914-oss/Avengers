import io
import os
import traceback
import zipfile
from datetime import datetime

import openpyxl
from django.core.management.base import BaseCommand
from django.db import connections

from apps.ownerclan.models import OwnerclanTask
from apps.ownerclan.services import (
    EXCEL_COL_MAP, MY_PRODUCT_FIELDS, _parse_excel_row, _safe_str,
)

PROGRESS_INTERVAL = 500


class Command(BaseCommand):
    help = '나의 상품 엑셀 비동기 업로드 워커'

    def add_arguments(self, parser):
        parser.add_argument('task_id', type=int)

    def handle(self, *args, **options):
        task_id = options['task_id']
        try:
            task = OwnerclanTask.objects.get(pk=task_id)
        except OwnerclanTask.DoesNotExist:
            return

        task.status = 'running'
        task.pid = os.getpid()
        task.save(update_fields=['status', 'pid'])

        file_path = task.input_data.get('file_path', '')

        try:
            result = _process_upload(file_path, task)
            task.result_data = result
            task.status = 'done'
        except Exception:
            task.result_data = {'error': traceback.format_exc()}
            task.status = 'error'
        finally:
            task.save(update_fields=['status', 'result_data'])
            try:
                os.unlink(file_path)
            except Exception:
                pass


def _load_workbooks(file_path):
    name = file_path.lower()
    if name.endswith('.zip'):
        with open(file_path, 'rb') as f:
            zf = zipfile.ZipFile(io.BytesIO(f.read()))
        all_names = zf.namelist()
        xlsx_names = sorted(n for n in all_names if n.lower().endswith(('.xlsx', '.xlsm')))
        if not xlsx_names:
            preview = ', '.join(all_names[:10]) or '(empty)'
            raise ValueError(f'ZIP 안에 .xlsx/.xlsm 파일이 없습니다. 발견({len(all_names)}개): {preview}')
        wbs = []
        for xn in xlsx_names:
            xlsx_bytes = zf.read(xn)
            wbs.append((xn, openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)))
        return wbs
    elif name.endswith(('.xlsx', '.xlsm')):
        return [(os.path.basename(file_path), openpyxl.load_workbook(file_path, read_only=True))]
    else:
        raise ValueError(f'xlsx 또는 zip 파일만 업로드 가능합니다. (받은: {os.path.basename(file_path)})')


def _next_my_code(cur, source_code):
    cur.execute(
        "SELECT COUNT(*) FROM my_product WHERE source_product_code=%s",
        [source_code],
    )
    n = cur.fetchone()[0]
    return f'M_{source_code}_{n + 1}'


def _process_upload(file_path, task):
    workbooks = _load_workbooks(file_path)
    now = datetime.now()

    rows = []
    for wb_name, wb in workbooks:
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 3:
                continue
            product_code = _safe_str(row[2])
            if not product_code:
                continue
            data = _parse_excel_row(list(row))
            rows.append((product_code, data))
        wb.close()

    total_rows = len(rows)
    task.result_data = {
        'progress': 0, 'inserted': 0, 'updated': 0, 'skipped': 0, 'total_rows': total_rows,
    }
    task.save(update_fields=['result_data'])

    inserted = 0
    fields = MY_PRODUCT_FIELDS
    all_cols = ['source_product_code', 'my_product_code'] + fields + ['copied_at']
    placeholders = ', '.join(['%s'] * len(all_cols))

    with connections['default'].cursor() as cur:
        for idx, (source_code, data) in enumerate(rows, 1):
            my_code = _next_my_code(cur, source_code)
            vals = [source_code, my_code] + [data.get(f) for f in fields] + [now]
            cur.execute(
                f"INSERT INTO my_product ({', '.join(all_cols)}) VALUES ({placeholders})",
                vals,
            )
            inserted += 1

            if idx % PROGRESS_INTERVAL == 0 or idx == total_rows:
                progress = int(idx * 100 / total_rows) if total_rows else 100
                task.result_data = {
                    'progress': progress,
                    'inserted': inserted,
                    'updated': 0,
                    'skipped': 0,
                    'total_rows': total_rows,
                }
                task.save(update_fields=['result_data'])

    return {
        'inserted': inserted,
        'updated': 0,
        'skipped': 0,
        'total': inserted,
    }
