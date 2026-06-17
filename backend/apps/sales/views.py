import csv
import io
from rest_framework import viewsets, views, status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.db.models import Sum, Count
from .models import SalesRecord, SalesUploadLog
from .serializers import SalesRecordSerializer, SalesUploadLogSerializer
from apps.accounts.models import SellerAccount


def _matrix_to_rows(matrix):
    """2차원 배열(첫 행=헤더) → dict 목록"""
    rows = []
    if not matrix:
        return rows
    headers = [str(h if h is not None else '').strip() for h in matrix[0]]
    for r in matrix[1:]:
        rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return rows


def read_upload_rows(file):
    """업로드 파일을 형태 자동판별해서 dict 목록으로 변환.
    매직바이트로 실제 포맷 판단 → 확장자 위장(.xls인데 HTML/엑셀)에도 견고."""
    data = file.read()
    if not data:
        return []
    head = data[:8]

    # 1) 진짜 .xls (OLE2 매직)
    if head[:4] == b'\xd0\xcf\x11\xe0':
        import xlrd
        wk = xlrd.open_workbook(file_contents=data)
        ws = wk.sheet_by_index(0)
        matrix = []
        for ri in range(ws.nrows):
            rv = []
            for ci in range(ws.ncols):
                cell = ws.cell(ri, ci)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        rv.append(xlrd.xldate_as_datetime(cell.value, wk.datemode).date().isoformat())
                    except Exception:
                        rv.append(cell.value)
                else:
                    rv.append(cell.value)
            matrix.append(rv)
        return _matrix_to_rows(matrix)

    # 2) 진짜 .xlsx (zip 매직 PK)
    if head[:2] == b'PK':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        return _matrix_to_rows([list(r) for r in ws.iter_rows(values_only=True)])

    # 3) 텍스트 디코드 (한글 인코딩 대응)
    text = None
    for enc in ('utf-8-sig', 'cp949', 'euc-kr', 'utf-16'):
        try:
            text = data.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = data.decode('utf-8', errors='ignore')

    low = text.lower()
    # 4) HTML 위장 (.xls인데 실제 HTML 테이블 — 한국 마켓 정산자료에 흔함)
    if '<table' in low or '<html' in low[:1000] or '<tr' in low:
        import pandas as pd
        dfs = pd.read_html(io.StringIO(text))
        if dfs:
            df = max(dfs, key=lambda d: d.shape[0] * d.shape[1]).fillna('')
            cols = [str(c).strip() for c in df.columns]
            return [{cols[i]: row.iloc[i] for i in range(len(cols))} for _, row in df.iterrows()]
        return []

    # 5) CSV / TSV
    delim = '\t' if text.count('\t') > text.count(',') else ','
    return list(csv.DictReader(io.StringIO(text), delimiter=delim))


class SalesRecordViewSet(viewsets.ModelViewSet):
    queryset = SalesRecord.objects.select_related('seller').all()
    serializer_class = SalesRecordSerializer
    filterset_fields = ['seller', 'status', 'order_date']
    search_fields = ['product_name', 'product_code', 'order_number']
    ordering_fields = ['order_date', 'total_price']


class SalesUploadView(views.APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        seller_id = request.data.get('seller')   # 선택사항: 파일에 셀러 컬럼이 없을 때의 기본 셀러
        if not file:
            return Response({'error': '파일을 선택해주세요.'}, status=400)

        default_seller = SellerAccount.objects.filter(id=seller_id).first() if seller_id else None

        # 셀러 매칭 맵 (셀러명/아이디 → 계정) — 한 파일에 여러 셀러 섞여도 행별 자동 매칭
        by_name, by_sid, SHOP_ALIAS_MAP = {}, {}, {}
        for s in SellerAccount.objects.all():
            if s.seller_name:
                by_name[s.seller_name.strip().lower()] = s
            if s.seller_id:
                by_sid[s.seller_id.strip().lower()] = s
            # memo에 "쇼핑몰명:OOO" (또는 alias:/별칭:) 줄이 있으면 그 상점명을 이 계정으로 매핑
            for line in (s.memo or '').splitlines():
                if ':' in line and ('쇼핑몰' in line or 'alias' in line.lower() or '별칭' in line):
                    al = line.split(':', 1)[1].strip().lower()
                    if al:
                        SHOP_ALIAS_MAP[al] = s

        # CSV/엑셀(.xls/.xlsx)/HTML위장 자동판별 → 행(dict) 목록으로 통일
        try:
            rows = read_upload_rows(file)
        except Exception as ex:
            return Response({'error': f'파일 읽기 실패 ({(file.name or "")[-20:]}): {ex}'}, status=400)
        if not rows:
            return Response({'error': '파일에서 데이터를 못 읽었습니다. 형식/헤더를 확인하세요.'}, status=400)

        # 쇼핑몰id(로그인 아이디) 매칭 준비: 우리 크롤러 계정 전체(11번가+지마켓 등) 로그인 아이디 집합
        # (이전엔 platform='11st'만 넣어 지마켓 쇼핑몰id가 매칭 안 됐음 — 지마켓 매출 대량 매칭대기 원인)
        from apps.cpc.models import CrawlerAccount
        login_ids = set(by_sid.keys())
        _login_platform = {}   # login_id(lower) → platform : 자동생성 SellerAccount에 올바른 플랫폼 부여
        for _a in CrawlerAccount.objects.all():
            if _a.login_id:
                lk = _a.login_id.strip().lower()
                login_ids.add(lk)
                _login_platform.setdefault(lk, _a.platform)
        # 값이 로그인 아이디와 일치하는 컬럼을 '쇼핑몰id' 컬럼으로 자동 탐지(컬럼명에 의존 X)
        id_col = None
        try:
            sample = rows[:300]
            best = 0
            for col in list(rows[0].keys()):
                hit = sum(1 for r in sample if str(r.get(col, '')).strip().lower() in login_ids
                          and str(r.get(col, '')).strip())
                if hit > best:
                    best, id_col = hit, col
            if best < max(3, len(sample) // 10):
                id_col = None
        except Exception:
            id_col = None

        # 새 2026 양식: 항목 위치 고정 (B=쇼핑몰명, G=쇼핑몰=플랫폼, K=쇼핑몰ID).
        # 헤더명이 달라도/없어도 컬럼 위치로 잡는다(헤더명 매칭이 실패할 때의 폴백).
        ordered_cols = list(rows[0].keys()) if rows else []

        def _col_at(idx):
            return ordered_cols[idx] if 0 <= idx < len(ordered_cols) else None
        COL_SHOPNAME = _col_at(1)    # B열
        COL_PLATFORM = _col_at(6)    # G열
        COL_LOGINID = _col_at(10)    # K열
        # 쇼핑몰ID(K열) → SellerAccount. 매칭 키는 G열(플랫폼)+K열(로그인아이디).
        # K열 ID가 있으면 항상 그 계정으로 매칭/생성(쇼핑몰명으로 셀러를 만드는 옛 방식 폐기 →
        # 'starvis7783@gmail.com' / '아 이리스.' 같은 쓰레기 셀러 자동생성 차단)
        _seller_cache = {}

        def _seller_by_login(lid, platform='', name=''):
            lid = (lid or '').strip()
            if not lid:
                return None
            key = lid.lower()
            if key in by_sid:
                return by_sid[key]
            if key in _seller_cache:
                return _seller_cache[key]
            # 플랫폼: 크롤러계정에 등록된 플랫폼 우선, 없으면 G열로 받은 플랫폼
            plat = _login_platform.get(key) or platform or '11st'
            acc = (SellerAccount.objects.filter(seller_id__iexact=lid).first()
                   or SellerAccount.objects.create(seller_id=lid,
                                                   seller_name=(name or lid)[:200],
                                                   platform=plat))
            by_sid[key] = acc
            _seller_cache[key] = acc
            return acc

        def _g(row, *keys, default=''):
            for k in keys:
                if k in row and row[k] not in (None, ''):
                    return row[k]
            return default

        def _gi(row, *keys):
            v = _g(row, *keys, default=0)
            try:
                return int(float(str(v).replace(',', '').replace('원', '').strip() or 0))
            except Exception:
                return 0

        import re as _re

        def _classify(t):
            if '11번가' in t or '11st' in t: return '11st'
            if '지마켓' in t or 'g마켓' in t or 'gmarket' in t: return 'gmarket'
            if '옥션' in t or 'auction' in t: return 'auction'
            if '쿠팡' in t or 'coupang' in t: return 'coupang'
            if '스마트' in t or 'smartstore' in t or '네이버' in t or 'naver' in t: return 'smartstore'
            if '롯데' in t or 'lotte' in t: return 'lotteon'
            if '에이블리' in t or 'ably' in t: return 'ably'
            return None

        def _platform(shop_col, shop_name):
            # G열(쇼핑몰: 01.지마켓 / 02.옥션 / 03.11번가 ...)을 가장 먼저 기준으로 구분
            c = str(shop_col).strip().lower()
            if c:
                p = _classify(c)
                if p:
                    return p
            # G열로 구분 안 되면 쇼핑몰명으로 보조 판정
            p = _classify((c + ' ' + str(shop_name)).lower())
            if p:
                return p
            return (str(shop_col).strip() or '기타')[:20]

        def _match_seller(skey):
            if not skey:
                return default_seller
            seller = by_name.get(skey.lower()) or by_sid.get(skey.lower())
            if seller is None:
                base = _re.sub(r'\s*(11번가|지마켓|g마켓|gmarket|11st|옥션|auction|쿠팡|coupang|스마트스토어|스토어|스스|smartstore|네이버|naver|톡스토어|esm)\s*$', '', skey, flags=_re.I).strip()
                if base and base != skey:
                    seller = by_name.get(base.lower()) or by_sid.get(base.lower())
                seller = seller or SHOP_ALIAS_MAP.get(skey.lower()) or SHOP_ALIAS_MAP.get(base.lower())
                if seller is None and base:
                    bl = base.lower()
                    cands = [(nm, acc) for nm, acc in by_name.items() if len(nm) >= 3 and bl.startswith(nm)]
                    if cands:
                        seller = max(cands, key=lambda x: len(x[0]))[1]
            return seller

        saved, matched, fail, errors_list = 0, 0, 0, []
        unmatched = {}
        to_create = []
        name_votes = {}   # seller.pk -> {B열쇼핑몰명: 횟수} : 업로드 후 셀러명 정정·별칭 등록용
        for i, row in enumerate(rows, 1):
            try:
                od_raw = _g(row, '결제일시', '주문일시', '주문일', '결제일', 'order_date', default='')
                odt = None                          # 주문일시(시간포함) — 중복판별용
                from datetime import datetime as _dt
                import pytz as _pytz
                _kst = _pytz.timezone('Asia/Seoul')
                if hasattr(od_raw, 'hour'):          # 엑셀 datetime 셀
                    odt = od_raw if getattr(od_raw, 'tzinfo', None) else _kst.localize(od_raw)
                    od = od_raw.date().isoformat()
                elif hasattr(od_raw, 'isoformat'):   # date 셀
                    od = od_raw.isoformat()
                else:
                    s = str(od_raw).strip().replace('/', '-')
                    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
                        try:
                            odt = _kst.localize(_dt.strptime(s[:19], fmt)); break
                        except Exception:
                            pass
                    od = s[:10]
                if not od:
                    continue                        # 날짜 없는 빈 행만 스킵
                shop_raw = str(_g(row, '쇼핑몰명', '셀러명', '셀러', '판매자', 'seller_name', 'seller')).strip()
                if not shop_raw and COL_SHOPNAME:
                    shop_raw = str(_g(row, COL_SHOPNAME)).strip()    # B열 위치 폴백
                id_raw = (str(_g(row, id_col)).strip() if id_col else
                          str(_g(row, '쇼핑몰id', '쇼핑몰ID', '쇼핑몰 ID', '쇼핑몰아이디', '쇼핑몰계정아이디',
                                 '로그인아이디', '판매아이디', '판매자아이디', '계정아이디', '아이디',
                                 'seller_id', 'login_id')).strip())
                if not id_raw and COL_LOGINID:
                    id_raw = str(_g(row, COL_LOGINID)).strip()       # K열 위치 폴백
                # 플랫폼: G열(쇼핑몰=11번가/지마켓/옥션/쿠팡/스마트스토어/롯데온)이 1차 기준
                platform = _platform(_g(row, '쇼핑몰', '마켓', 'platform') or
                                     (str(_g(row, COL_PLATFORM)) if COL_PLATFORM else ''), shop_raw)
                # 매칭 = G열(플랫폼) + K열(쇼핑몰ID). K열 ID 우선, 없을 때만 쇼핑몰명 보조매칭
                seller = _seller_by_login(id_raw, platform, shop_raw) or _match_seller(shop_raw)
                if seller and shop_raw and '@' not in shop_raw and shop_raw.lower() != (seller.seller_id or '').lower():
                    v = name_votes.setdefault(seller.pk, {})
                    v[shop_raw] = v.get(shop_raw, 0) + 1
                qty = _gi(row, '수량', 'quantity') or 1
                # 매출 = 정산받는금액 (없으면 합계/판매금액/판매가)
                revenue = _gi(row, '정산받는금액', '합계', '매출', '판매금액', 'total_price') or (_gi(row, '판매가') * qty)
                # 구매가(원가) = 판매사 주문관리 메모
                cost = _gi(row, '판매사 주문관리 메모', '판매사주문관리메모', '구매가', '원가', 'cost')
                net = revenue - cost   # 순익 = 매출 - 구매가
                # 매칭 안 돼도 저장(seller=None) → 나중에 하나씩 매칭
                to_create.append(SalesRecord(
                    seller=seller,
                    shop_name=(shop_raw or id_raw)[:200],
                    platform=platform,
                    order_date=od,
                    order_datetime=odt,
                    order_number=str(_g(row, '주문번호', 'order_number'))[:100],
                    product_name=str(_g(row, '상품명', '주문내역', 'product_name'))[:500],
                    product_code=str(_g(row, '상품코드', '오너클랜상품코드', '상품번호', 'product_code'))[:100],
                    quantity=qty,
                    unit_price=_gi(row, '판매가', '단가', 'unit_price'),
                    total_price=revenue,
                    cost=cost,
                    commission=_gi(row, '수수료', '마켓수수료', 'commission'),
                    shipping_fee=_gi(row, '배송비', 'shipping_fee'),
                    net_profit=net,
                ))
                saved += 1
                if seller:
                    matched += 1
                else:
                    k = shop_raw or id_raw or '(미지정)'
                    unmatched[k] = unmatched.get(k, 0) + 1
            except Exception as e:
                fail += 1
                if len(errors_list) < 50:
                    errors_list.append(f'행 {i}: {str(e)[:80]}')

        # 쇼핑몰별 기간 교체 (스냅샷, 계정 안전): 파일에 '있는 쇼핑몰'만 그 쇼핑몰의 기간[min~max]을
        # 삭제 후 재삽입. → ① 그 쇼핑몰의 취소/반품 건은 자동 제거(범위 통째 교체)
        #   ② 파일에 없는 다른 쇼핑몰/계정 데이터는 그대로 보존(계정별 파일 올려도 손실 0)
        #   매출=정산받는금액.
        replaced = 0
        if to_create:
            from collections import defaultdict
            # 셀러(계정) 기준 기간 교체: 매칭된 계정은 (platform, seller)로, 미매칭은 (platform, shop_name)로.
            grp = defaultdict(list)
            for r in to_create:
                key = ('S', r.platform, r.seller_id) if r.seller_id else ('N', r.platform, r.shop_name)
                grp[key].append(r.order_date)
            for key, ds in grp.items():
                if key[0] == 'S':
                    old = SalesRecord.objects.filter(platform=key[1], seller_id=key[2],
                                                     order_date__gte=min(ds), order_date__lte=max(ds))
                else:
                    old = SalesRecord.objects.filter(platform=key[1], seller__isnull=True, shop_name=key[2],
                                                     order_date__gte=min(ds), order_date__lte=max(ds))
                replaced += old.count()
                old.delete()

            # 옛 양식 정리: 쇼핑몰'명'으로 잘못 생성된 셀러(seller_id에 공백/한글 — 로그인아이디로
            # 절대 나올 수 없는 꼴)의 동일기간 레코드를 제거 → 재업로드 시 이중집계 방지.
            # (이메일 키는 스마트스토어 실제 로그인일 수 있어 제외 — 위 계정별 교체로 처리됨)
            junk_ids = [s.id for s in SellerAccount.objects.all()
                        if (' ' in (s.seller_id or '')) or any(ord(c) > 0x7f for c in (s.seller_id or ''))]
            if junk_ids:
                plat_range = defaultdict(list)
                for r in to_create:
                    plat_range[r.platform].append(r.order_date)
                for plat, ds in plat_range.items():
                    old = SalesRecord.objects.filter(platform=plat, seller_id__in=junk_ids,
                                                     order_date__gte=min(ds), order_date__lte=max(ds))
                    replaced += old.count()
                    old.delete()

        # 일괄 삽입 (7천건도 수초) — 한 건씩 create하면 느려서 502
        if to_create:
            SalesRecord.objects.bulk_create(to_create, batch_size=500)

        # 쇼핑몰명 정정·별칭 등록: 계정(K열 ID)별로 B열 쇼핑몰명을 정식 이름으로 채택.
        # seller_name이 placeholder(빈값/아이디·이메일과 동일)면 B열 실제 상점명으로 교체하고,
        # 다음 업로드부터 쇼핑몰명만으로도 매칭되게 memo에 별칭 등록.
        renamed = 0
        if name_votes:
            sellers = {s.pk: s for s in SellerAccount.objects.filter(pk__in=name_votes.keys())}
            for pk, votes in name_votes.items():
                s = sellers.get(pk)
                if not s:
                    continue
                best = max(votes.items(), key=lambda x: x[1])[0][:200]
                cur = (s.seller_name or '').strip()
                changed = False
                # placeholder 판정: 비었거나 / 아이디·이메일과 동일하거나 / @ 포함
                if (not cur or cur.lower() == (s.seller_id or '').lower() or '@' in cur) and best:
                    s.seller_name = best
                    changed = True
                # 별칭(memo 쇼핑몰명:OOO) 등록 — 다음 업로드 자동매칭
                alias_line = f'쇼핑몰명:{best}'
                if best and alias_line not in (s.memo or ''):
                    s.memo = ((s.memo or '') + ('\n' if s.memo else '') + alias_line)
                    changed = True
                if changed:
                    s.save(update_fields=['seller_name', 'memo'])
                    renamed += 1

        pending = saved - matched
        if unmatched:
            summary = ', '.join(f'{k}({v})' for k, v in sorted(unmatched.items(), key=lambda x: -x[1])[:25])
            errors_list.insert(0, f'📌 매칭대기 {pending}건 (셀러 미연결, 저장됨): {summary}')

        log = SalesUploadLog.objects.create(
            file_name=file.name,
            row_count=saved + fail,
            success_count=matched,
            error_count=pending,
            errors=errors_list,
        )
        resp = SalesUploadLogSerializer(log).data
        resp['saved_total'] = saved
        resp['matched'] = matched
        resp['pending_match'] = pending
        resp['parse_fail'] = fail
        resp['replaced'] = replaced
        resp['renamed'] = renamed
        if saved == 0 and rows:
            resp['detected_columns'] = [str(k) for k in rows[0].keys()]
            resp['hint'] = '저장 0건 — 컬럼명(특히 결제일시/주문일)을 인식 못했습니다.'
        return Response(resp, status=201)


class SalesUnmatchedView(views.APIView):
    """매칭 안 된(셀러 미연결) 매출의 쇼핑몰명 목록 + 추천 셀러"""
    def get(self, request):
        from django.db.models import Count, Sum
        groups = (SalesRecord.objects.filter(seller__isnull=True)
                  .values('shop_name', 'platform')
                  .annotate(cnt=Count('id'), sales=Sum('total_price'))
                  .order_by('-cnt'))
        accts = list(SellerAccount.objects.all())
        names = [(a.seller_name or '', a) for a in accts if a.seller_name]
        import re
        def suggest(shop):
            base = re.sub(r'\s*(11번가|지마켓|g마켓|gmarket|11st|옥션|auction|쿠팡|coupang|스마트스토어|스토어|스스|smartstore|네이버|naver|톡스토어|esm)\s*$', '', shop, flags=re.I).strip()
            bl = base.lower()
            # 양방향 부분일치 중 가장 비슷한 것
            cands = [a for nm, a in names if bl and (bl.startswith(nm.lower()) or nm.lower().startswith(bl) or bl in nm.lower() or nm.lower() in bl)]
            return cands[0] if cands else None
        out = []
        for g in groups:
            s = suggest(g['shop_name'] or '')
            out.append({
                'shop_name': g['shop_name'], 'platform': g['platform'],
                'count': g['cnt'], 'sales': g['sales'] or 0,
                'suggested_seller_id': s.id if s else None,
                'suggested_seller_name': s.seller_name if s else None,
            })
        return Response(out)


class SalesMatchView(views.APIView):
    """대기 매출(쇼핑몰명)을 셀러에 연결. remember=True면 다음 업로드부터 자동매칭되게 별칭 저장."""
    def post(self, request):
        shop_name = request.data.get('shop_name', '')
        seller_id = request.data.get('seller_id')
        remember = request.data.get('remember', True)
        if seller_id is None:
            return Response({'error': 'seller_id 필요'}, status=400)
        seller = SellerAccount.objects.filter(id=seller_id).first()
        if not seller:
            return Response({'error': '셀러 없음'}, status=404)
        n = SalesRecord.objects.filter(shop_name=shop_name, seller__isnull=True).update(seller=seller)
        if remember and shop_name:
            line = f'쇼핑몰명:{shop_name}'
            memo = seller.memo or ''
            if line not in memo:
                seller.memo = (memo + ('\n' if memo else '') + line)
                seller.save(update_fields=['memo'])
        return Response({'matched': n, 'seller': seller.seller_name})


class SalesSummaryView(views.APIView):
    def get(self, request):
        date_from = request.query_params.get('from')
        date_to = request.query_params.get('to')
        qs = SalesRecord.objects.filter(status='completed')
        if date_from:
            qs = qs.filter(order_date__gte=date_from)
        if date_to:
            qs = qs.filter(order_date__lte=date_to)
        summary = qs.aggregate(
            total_revenue=Sum('total_price'),
            total_profit=Sum('net_profit'),
            total_orders=Count('id'),
            total_quantity=Sum('quantity'),
        )
        # 마켓(쇼핑몰)별 집계 — 11번가 → 지마켓 → 나머지 순
        PRIORITY = {'11st': 1, 'gmarket': 2, 'auction': 3, 'coupang': 4, 'smartstore': 5, 'ably': 6}
        LABEL = {'11st': '11번가', 'gmarket': '지마켓', 'auction': '옥션', 'coupang': '쿠팡',
                 'smartstore': '스마트스토어', 'ably': '에이블리'}

        # 플랫폼별 광고비 — 11번가(ElevenCostHistory CPC) + 지마켓/옥션(GmarketCostHistory market별)
        ad_by_platform = self._ad_cost_by_platform(date_from, date_to)

        rows = qs.values('platform').annotate(
            revenue=Sum('total_price'), profit=Sum('net_profit'),
            orders=Count('id'), quantity=Sum('quantity'))
        by_platform = []
        for r in rows:
            p = r['platform']; rev = r['revenue'] or 0; prof = r['profit'] or 0
            ad = ad_by_platform.get(p, 0)
            by_platform.append({
                'platform': p, 'label': LABEL.get(p, p or '기타'),
                'revenue': rev, 'profit': prof, 'orders': r['orders'],
                'quantity': r['quantity'] or 0,
                'ad_cost': ad, 'net_after_ad': prof - ad,
                'margin': round(prof * 100.0 / rev, 1) if rev else 0,
                'real_margin': round((prof - ad) * 100.0 / rev, 1) if rev else 0})
        by_platform.sort(key=lambda x: PRIORITY.get(x['platform'], 99))
        summary['by_platform'] = by_platform
        summary['total_ad_cost'] = sum(ad_by_platform.values())
        summary['total_net_after_ad'] = (summary['total_profit'] or 0) - summary['total_ad_cost']

        # 셀러(쇼핑몰id)별 매출 상위 — 매출 대비 순수익
        shop_rows = (qs.values('seller__seller_id', 'platform')
                     .annotate(revenue=Sum('total_price'), profit=Sum('net_profit'), orders=Count('id'))
                     .order_by('-revenue')[:100])
        summary['by_shop'] = [
            {'shop_id': s['seller__seller_id'] or '(미연결)', 'platform': s['platform'],
             'label': LABEL.get(s['platform'], s['platform'] or '기타'),
             'revenue': s['revenue'] or 0, 'profit': s['profit'] or 0, 'orders': s['orders'],
             'margin': round((s['profit'] or 0) * 100.0 / s['revenue'], 1) if s['revenue'] else 0}
            for s in shop_rows]
        return Response(summary)

    def _ad_cost_by_platform(self, date_from, date_to):
        """플랫폼별 광고비 합계. 11st=ElevenCostHistory CPC, gmarket/auction=GmarketCostHistory market별."""
        import datetime as _dt
        import pytz
        from apps.cpc.models import GmarketCostHistory, ElevenCostHistory
        AD = ('CPC', 'AI매출업', '서버비용')
        out = {}
        gq = GmarketCostHistory.objects.filter(transaction_type__in=AD).exclude(comment__icontains='판매예치금')  # 판매예치금 송금 등 비광고 차감 제외
        if date_from: gq = gq.filter(use_date__gte=date_from)
        if date_to: gq = gq.filter(use_date__lte=date_to)
        for mk, plat in [('gmarket', 'gmarket'), ('auction', 'auction')]:
            out[plat] = abs(gq.filter(market=mk).aggregate(s=Sum('amount'))['s'] or 0)
        # 11번가 — transaction_datetime(KST 범위)
        eq = ElevenCostHistory.objects.filter(transaction_type='CPC')
        kst = pytz.timezone('Asia/Seoul')
        if date_from:
            d = _dt.datetime.strptime(date_from, '%Y-%m-%d')
            eq = eq.filter(transaction_datetime__gte=kst.localize(d))
        if date_to:
            d = _dt.datetime.strptime(date_to, '%Y-%m-%d') + _dt.timedelta(days=1)
            eq = eq.filter(transaction_datetime__lt=kst.localize(d))
        out['11st'] = abs(eq.aggregate(s=Sum('amount'))['s'] or 0)
        return out


class SalesUploadLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SalesUploadLog.objects.all().order_by('-uploaded_at')
    serializer_class = SalesUploadLogSerializer
